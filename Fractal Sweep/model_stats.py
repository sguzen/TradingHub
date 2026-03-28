#!/usr/bin/env python3
"""
model_stats.py  —  Multi-Timeframe Sweep + CISD Statistical Engine v6.0
========================================================================
Runs each of the four model pairs across TWO sweep modes × ONE CISD mode,
producing up to 8 output keys.

SWEEP MODE
  PREV     — sweep reference = immediately prior sweep-TF candle

CISD MODE
  CISD — Change in State of Delivery (body-only):
         • Wicks ignored — only opens and closes matter
         • Dojis (close == open) are neutral (skip, don't break run)
         • Step 1 — BACKWARD from the return bar: find the consecutive
           opposing delivery run that formed the high/low before the sweep.
             SHORT: find consecutive up-close run before return bar
             LONG:  find consecutive down-close run before return bar
         • cisd_level = open of the FIRST (earliest) candle in that run
             SHORT: lowest open in the ascending bullish run
             LONG:  highest open in the descending bearish run
         • Step 2 — FORWARD from the return bar: fire on close through level
             LONG:  close > cisd_level
             SHORT: close < cisd_level

Performance:
  • resolve_outcomes_vectorised: fully numpy, no iterrows — ~20-40× faster
  • detect_model: all inner loops use pre-built numpy arrays + searchsorted,
    no pandas .loc[] slicing in the hot path
  • find_unswept_level: uses passed-in numpy arrays, no list comprehensions
  • find_cisd: integer index arithmetic, no df.loc[] per call
  • All TF arrays built once, reused across all 8 model/mode combos

Usage:
    python3 model_stats.py                              # all 4 models
    python3 model_stats.py --models 1H_5M 1H_3M
    python3 model_stats.py --cisd-fast-bars 12
    python3 model_stats.py --table es_1m
"""

import argparse
import duckdb
import pandas as pd
import numpy as np
import json
from pathlib import Path
import scipy.stats as _scipy_stats

# ── PATHS ─────────────────────────────────────────────────────────────────────
DB_PATH   = Path(__file__).parent / 'candle_science.duckdb'
OUT_PATH  = Path(__file__).parent / 'model_stats.json'
TABLE     = 'nq_1m'
PHASE1_XL = Path(__file__).parent / 'fractal_phase1_results.xlsx'

# ── DATE → CLASSIFICATION MAP (from fractal_phase1_results.xlsx) ──────────────
def _load_date_classification():
    """Build {date_str: cls_key} from phase1 classification data sheets."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PHASE1_XL, read_only=True, data_only=True)
        mapping = {}
        for cls_key, sheet_name in [('DWP','DWP Data'),('DNP','DNP Data'),
                                     ('R1','R1 Data'),('R2','R2 Data'),
                                     ('Unclassified','Unclassified Data')]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            hdr = rows[2]
            ci  = {h: i for i, h in enumerate(hdr) if h}
            date_col = ci.get('Date')
            if date_col is None:
                continue
            for row in rows[3:]:
                if not row[0]:
                    continue
                try:
                    d = str(row[date_col])[:10]
                    if d and d not in mapping:
                        mapping[d] = cls_key
                except Exception:
                    continue
        return mapping
    except Exception as e:
        print(f'  [warn] Could not load classification map: {e}')
        return {}

DATE_CLASSIFICATION = _load_date_classification()
print(f'  Loaded {len(DATE_CLASSIFICATION)} date→classification mappings')

# ── GLOBAL CONSTANTS ──────────────────────────────────────────────────────────
ACCOUNT_SIZE     = 4500   # $ account size for risk metrics
RISK_PER_TRADE   = 225    # $ risk per trade
MIN_RISK_PTS     = 3.0
OUTCOME_MAX_BARS = 360
SWEEP_MIN_PCT    = 0.10
SWEEP_MAX_PCT    = 1.50
CISD_FAST_BARS   = 8
UNSWEPT_LOOKBACK = 10

# ── RISK PROFILES ─────────────────────────────────────────────────────────────
# Each tuple: (stop_val, target_val, display_key, profile_type)
#
# profile_type='mult'  → stop/target distances = val × base_risk
#                         base_risk = |entry_price − sweep_extreme|
#
# profile_type='pct'   → stop/target distances = entry_price × val / 100
#                         (fixed % of entry price, independent of sweep size)
RR_PROFILES = [
    # --- DDLI Weighted Top 10 (Sharpe 25% · PF 20% · EV 15% · SQN 15% · MaxDD 10% · RoR 10% · CE 5%) ---
    (0.26, 0.18, 'sl_026_tp_018', 'pct'),
    (0.26, 0.19, 'sl_026_tp_019', 'pct'),
    (0.21, 0.18, 'sl_021_tp_018', 'pct'),
    (0.20, 0.18, 'sl_020_tp_018', 'pct'),
    (0.18, 0.18, 'sl_018_tp_018', 'pct'),
    (0.19, 0.18, 'sl_019_tp_018', 'pct'),
    (0.20, 0.21, 'sl_020_tp_021', 'pct'),
    (0.20, 0.19, 'sl_020_tp_019', 'pct'),
    (0.18, 0.19, 'sl_018_tp_019', 'pct'),
    (0.19, 0.19, 'sl_019_tp_019', 'pct'),
    # --- Structural Dynamic: SL = sweep extreme (1×base_risk), TP1 = 1R, 50% exit; runner with BE stop ---
    (1.0, 1.0, 'structural_dynamic', 'structural'),
]
DEFAULT_PROFILE = 'sl_026_tp_018'

DOW_NAMES = {0:'Sun', 1:'Mon', 2:'Tue', 3:'Wed', 4:'Thu', 5:'Fri', 6:'Sat'}
HR_LABELS = {h: f"{h:02d}:00" for h in range(0, 24)}

# ── MODEL DEFINITIONS ─────────────────────────────────────────────────────────
MODELS = {
    '4H_15M': dict(
        label        = '4H Sweep · 15M CISD',
        sweep_tf_min = 4 * 60,
        cisd_tf_min  = 15,
        q1_min       = 60,
        min_range    = 30,
        session_hrs  = (7.0, 16.0),
    ),
    '1H_5M': dict(
        label        = '1H Sweep · 5M CISD',
        sweep_tf_min = 60,
        cisd_tf_min  = 5,
        q1_min       = 15,
        min_range    = 12,
        session_hrs  = (7.0, 16.0),
    ),
    '1H_3M': dict(
        label        = '1H Sweep · 3M CISD',
        sweep_tf_min = 60,
        cisd_tf_min  = 3,
        q1_min       = 15,
        min_range    = 12,
        session_hrs  = (7.0, 16.0),
    ),
    '30M_3M': dict(
        label        = '30M Sweep · 3M CISD',
        sweep_tf_min = 30,
        cisd_tf_min  = 3,
        q1_min       = 8,
        min_range    = 8,
        session_hrs  = (7.0, 16.0),
    ),
}
SWEEP_MODES = ['PREV']
CISD_MODES  = ['CISD']


# ── DB ────────────────────────────────────────────────────────────────────────
def connect(db_path):
    return duckdb.connect(str(db_path), read_only=True)


# ── LOAD 1m BARS ──────────────────────────────────────────────────────────────
def load_1m(con, table):
    print(f"[1] Loading {table} ...")
    raw = con.execute(f"""
        SELECT
            CAST(timezone('America/New_York', timestamp) AS DATE)  AS trade_date,
            CAST(timezone('America/New_York', timestamp) AS TIME)  AS bar_time,
            open::DOUBLE AS open, high::DOUBLE AS high,
            low::DOUBLE  AS low,  close::DOUBLE AS close,
            date_part('year',   timezone('America/New_York', timestamp)) AS yr,
            date_part('month',  timezone('America/New_York', timestamp)) AS mo,
            date_part('dow',    timezone('America/New_York', timestamp)) AS dow,
            date_part('hour',   timezone('America/New_York', timestamp)) AS hr,
            date_part('minute', timezone('America/New_York', timestamp)) AS mn
        FROM {table}
        ORDER BY timestamp
    """).df()
    raw['ts'] = pd.to_datetime(
        raw['trade_date'].astype(str) + ' ' + raw['bar_time'].astype(str)
    )
    raw = raw.sort_values('ts').reset_index(drop=True).set_index('ts')
    raw_rth = raw[
        (raw['hr'] >= 7) & ((raw['hr'] < 16) | ((raw['hr'] == 16) & (raw['mn'] == 0)))
    ].copy()
    print(f"   {len(raw):,} total bars  |  {len(raw_rth):,} RTH (07:00–16:00)")
    return raw, raw_rth


# ── RESAMPLE ──────────────────────────────────────────────────────────────────
def resample(df_1m, tf_min, label):
    print(f"   Building {label} bars ...")
    df2 = df_1m.copy()
    df2['ts_tf'] = df2.index.floor(f"{tf_min}min")
    agg = df2.groupby('ts_tf').agg(
        trade_date=('trade_date','first'), yr=('yr','first'), mo=('mo','first'),
        dow=('dow','first'), hr=('hr','first'),
        open_tf=('open','first'), high_tf=('high','max'),
        low_tf=('low','min'),    close_tf=('close','last'),
    ).sort_index()
    print(f"      {len(agg):,} {label} bars")
    return agg


# ── NUMPY ARRAY BUILDERS ──────────────────────────────────────────────────────
def df_to_arrays(df):
    """
    Convert a time-indexed resampled OHLC dataframe to numpy arrays.
    ts_ns is int64 nanoseconds — used with np.searchsorted for fast lookups.
    Called ONCE per timeframe; results reused across all model/mode combos.
    """
    return dict(
        ts_ns      = df.index.view('int64').copy(),
        open       = df['open_tf'].values.astype('float64'),
        high       = df['high_tf'].values.astype('float64'),
        low        = df['low_tf'].values.astype('float64'),
        close      = df['close_tf'].values.astype('float64'),
        trade_date = df['trade_date'].values.copy(),
        yr         = df['yr'].values.astype('int32'),
        dow        = df['dow'].values.astype('int32'),
        hr         = df['hr'].values.astype('int32'),
    )


def df_1m_to_arrays(df):
    """Convert 1m dataframe to numpy arrays for vectorised outcome resolution."""
    return dict(
        ts_ns      = df.index.view('int64').copy(),
        open       = df['open'].values.astype('float64'),
        high       = df['high'].values.astype('float64'),
        low        = df['low'].values.astype('float64'),
        close      = df['close'].values.astype('float64'),
        hr         = df['hr'].values.astype('int32'),
        mn         = df['mn'].values.astype('int32'),
        yr         = df['yr'].values.astype('int32'),
        dow        = df['dow'].values.astype('int32'),
        trade_date = df['trade_date'].values.copy(),
    )


# ── SESSION LABEL ─────────────────────────────────────────────────────────────
def get_session(hr_float):
    if 8.5  <= hr_float < 11.5: return 'NY1'
    if 11.5 <= hr_float < 16.0: return 'NY2'
    if 7.0  <= hr_float < 8.5:  return 'PRE'
    if 0    <= hr_float < 7.0:  return 'OVERNIGHT'
    return 'OTHER'



# ── CISD DETECTION ────────────────────────────────────────────────────────────
def _find_cisd(c_opens, c_closes, c_ts_ns, start_idx, n_bars, direction):
    """
    CISD — Change in State of Delivery (body-only).

    Step 1 — BACKWARD scan from start_idx (the return bar):
      Find the consecutive opposing delivery run that formed the high/low
      BEFORE the return to range — i.e. the candles that made the sweep.
        SHORT: look for consecutive up-close candles (bullish delivery run)
        LONG:  look for consecutive down-close candles (bearish delivery run)
      Scan back up to n_bars before start_idx.

    cisd_level = open of the FIRST (earliest in time) candle in that run:
        SHORT: first = lowest open of the ascending bullish run
        LONG:  first = highest open of the descending bearish run

    Step 2 — FORWARD scan from start_idx:
      Fire on the first bar whose close crosses cisd_level:
        LONG:  close > cisd_level
        SHORT: close < cisd_level

    Dojis (close == open) are neutral — skipped, do not break the run.
    """
    end  = min(start_idx + n_bars, len(c_closes))
    back = max(0, start_idx - n_bars * 4)      # backward scan: look much further back for run start

    if direction == 'LONG':
        # ── Step 1: backward — find consecutive bearish run before start_idx ──
        j = start_idx - 1
        while j >= back and c_closes[j] == c_opens[j]:
            j -= 1                             # skip dojis
        if j < back or c_closes[j] >= c_opens[j]:
            return None, None                  # no bearish candle found
        run_start = j
        k = j - 1
        while k >= back:
            if   c_closes[k] < c_opens[k]:  run_start = k; k -= 1
            elif c_closes[k] == c_opens[k]:  k -= 1        # doji neutral
            else: break
        cisd_level = float(c_opens[run_start])  # open of FIRST (highest) bearish candle

        # ── Step 2: forward — first bar that closes above cisd_level ──────────
        for i in range(start_idx, end):
            if c_closes[i] > cisd_level:
                return c_ts_ns[i], cisd_level

    else:  # SHORT
        # ── Step 1: backward — find consecutive bullish run before start_idx ──
        j = start_idx - 1
        while j >= back and c_closes[j] == c_opens[j]:
            j -= 1
        if j < back or c_closes[j] <= c_opens[j]:
            return None, None                  # no bullish candle found
        run_start = j
        k = j - 1
        while k >= back:
            if   c_closes[k] > c_opens[k]:  run_start = k; k -= 1
            elif c_closes[k] == c_opens[k]:  k -= 1
            else: break
        cisd_level = float(c_opens[run_start])  # open of FIRST (lowest) bullish candle

        # ── Step 2: forward — first bar that closes below cisd_level ──────────
        for i in range(start_idx, end):
            if c_closes[i] < cisd_level:
                return c_ts_ns[i], cisd_level

    return None, None


def find_cisd(c_arrs, return_ts_ns, direction, max_bars, cisd_mode):
    """
    Uses searchsorted to find the integer index — no df.loc[].
    Returns (fired_ts_ns, cisd_level) or (None, None).
    """
    start_idx = int(np.searchsorted(c_arrs['ts_ns'], return_ts_ns, side='left'))
    if start_idx >= len(c_arrs['ts_ns']):
        return None, None
    return _find_cisd(
        c_arrs['open'], c_arrs['close'], c_arrs['ts_ns'],
        start_idx, min(max_bars * 2, 40), direction
    )


# ── VECTORISED OUTCOME RESOLUTION ────────────────────────────────────────────
def resolve_outcomes_vectorised(m1_arrs, pending):
    """
    Resolve WIN / LOSS / EXPIRED / INVALID for all entries in one pass.
    No Python loops over bars — uses numpy cumulative min/max + argmax.

    pending: list of dicts with keys:
        idx, entry_ts_ns, entry_price, stop_price, target_price, direction
    Returns list of (outcome, r) in the same order as pending.
    """
    ts_ns  = m1_arrs['ts_ns']
    highs  = m1_arrs['high']
    lows   = m1_arrs['low']
    closes = m1_arrs['close']
    N      = len(ts_ns)

    results = []
    for e in pending:
        entry_ts_ns  = e['entry_ts_ns']
        entry_price  = e['entry_price']
        stop_price   = e['stop_price']
        target_price = e['target_price']
        direction    = e['direction']

        risk = abs(entry_price - stop_price)
        if risk < MIN_RISK_PTS:
            results.append(('INVALID', 0.0, 0.0, 0.0))
            continue

        # Target from the manipulation-range projection (2× manip range from sweep extreme)
        target = target_price

        # Start the scan one bar after the entry bar opens
        start = int(np.searchsorted(ts_ns, entry_ts_ns, side='right'))
        end   = min(start + OUTCOME_MAX_BARS, N)

        if start >= N:
            results.append(('EXPIRED', 0.0, 0.0, 0.0))
            continue

        h = highs[start:end]
        l = lows[start:end]

        if direction == 'LONG':
            t_hit = h >= target
            s_hit = l <= stop_price
        else:
            t_hit = l <= target
            s_hit = h >= stop_price

        t_any = t_hit.any()
        s_any = s_hit.any()

        actual_rr = round(abs(target - entry_price) / risk, 2) if risk > 0 else 0.0

        # Determine trade end index for MAE/MFE window
        if not t_any and not s_any:
            last_r = (closes[end - 1] - entry_price) / risk \
                     if direction == 'LONG' \
                     else (entry_price - closes[end - 1]) / risk
            outcome, r_val = 'EXPIRED', round(float(last_r), 2)
            trade_end = end - start
        elif not s_any:
            outcome, r_val = 'WIN', actual_rr
            trade_end = int(np.argmax(t_hit)) + 1
        elif not t_any:
            outcome, r_val = 'LOSS', -1.0
            trade_end = int(np.argmax(s_hit)) + 1
        else:
            t_idx = int(np.argmax(t_hit))
            s_idx = int(np.argmax(s_hit))
            outcome, r_val = ('WIN', actual_rr) if t_idx <= s_idx else ('LOSS', -1.0)
            trade_end = (t_idx if t_idx <= s_idx else s_idx) + 1

        # MAE/MFE as % of entry price over the trade window
        h_w = h[:trade_end]
        l_w = l[:trade_end]
        if direction == 'LONG':
            mae_pct = round(float(max(0.0, entry_price - l_w.min()) / entry_price * 100), 4)
            mfe_pct = round(float(max(0.0, h_w.max() - entry_price) / entry_price * 100), 4)
        else:
            mae_pct = round(float(max(0.0, h_w.max() - entry_price) / entry_price * 100), 4)
            mfe_pct = round(float(max(0.0, entry_price - l_w.min()) / entry_price * 100), 4)

        results.append((outcome, r_val, mae_pct, mfe_pct))

    return results


# ── STRUCTURAL-DYNAMIC OUTCOME RESOLUTION ─────────────────────────────────────
def resolve_outcomes_structural(m1_arrs, pending):
    """
    Structural Dynamic profile:
      - Phase 1: scan for SL (stop_price) or TP1 (target_price = entry ± 1R)
      - If SL first → LOSS, r = -1.0
      - If TP1 first → 50% exits at +1R; runner holds with BE (entry) stop
      - Phase 2 (runner): scan for BE stop or EOD; runner_exit_r in R units
      - net_r = 0.5 × 1.0 + 0.5 × runner_exit_r

    Returns list of (outcome, net_r, mae_pct, mfe_pct, tp1_hit, runner_exit_r)
    """
    ts_ns  = m1_arrs['ts_ns']
    highs  = m1_arrs['high']
    lows   = m1_arrs['low']
    closes = m1_arrs['close']
    N      = len(ts_ns)

    results = []
    for e in pending:
        entry_ts_ns  = e['entry_ts_ns']
        entry_price  = e['entry_price']
        stop_price   = e['stop_price']
        target_price = e['target_price']   # TP1 = entry ± 1R
        direction    = e['direction']

        risk = abs(entry_price - stop_price)
        if risk < MIN_RISK_PTS:
            results.append(('INVALID', 0.0, 0.0, 0.0, False, 0.0))
            continue

        start = int(np.searchsorted(ts_ns, entry_ts_ns, side='right'))
        end   = min(start + OUTCOME_MAX_BARS, N)

        if start >= N:
            results.append(('EXPIRED', 0.0, 0.0, 0.0, False, 0.0))
            continue

        h = highs[start:end]
        l = lows[start:end]

        if direction == 'LONG':
            tp1_hit_arr = h >= target_price
            sl_hit_arr  = l <= stop_price
        else:
            tp1_hit_arr = l <= target_price
            sl_hit_arr  = h >= stop_price

        tp1_any = tp1_hit_arr.any()
        sl_any  = sl_hit_arr.any()

        tp1_idx = int(np.argmax(tp1_hit_arr)) if tp1_any else len(h)
        sl_idx  = int(np.argmax(sl_hit_arr))  if sl_any  else len(h)

        # ── Compute MAE/MFE over full trade window ────────────────────────────
        if direction == 'LONG':
            mae_pct = round(float(max(0.0, entry_price - l.min()) / entry_price * 100), 4)
            mfe_pct = round(float(max(0.0, h.max() - entry_price) / entry_price * 100), 4)
        else:
            mae_pct = round(float(max(0.0, h.max() - entry_price) / entry_price * 100), 4)
            mfe_pct = round(float(max(0.0, entry_price - l.min()) / entry_price * 100), 4)

        # ── Outcome determination ─────────────────────────────────────────────
        if not tp1_any and not sl_any:
            # Expired — neither TP1 nor SL hit
            last_r = ((closes[end - 1] - entry_price) / risk
                      if direction == 'LONG'
                      else (entry_price - closes[end - 1]) / risk)
            results.append(('EXPIRED', round(float(last_r), 2), mae_pct, mfe_pct, False, 0.0))
            continue

        if sl_any and (not tp1_any or sl_idx < tp1_idx):
            # SL hit before TP1 → full LOSS
            results.append(('LOSS', -1.0, mae_pct, mfe_pct, False, 0.0))
            continue

        # ── TP1 hit first — run the runner with BE stop ───────────────────────
        runner_start = tp1_idx + 1
        runner_end   = len(h)
        runner_exit_r = 0.0  # default: runner reaches BE

        if runner_start < runner_end:
            rh = h[runner_start:runner_end]
            rl = l[runner_start:runner_end]
            if direction == 'LONG':
                be_hit = rl <= entry_price
            else:
                be_hit = rh >= entry_price

            if not be_hit.any():
                # Runner survived to EOD — mark to market at last close
                last_close = closes[start + runner_end - 1]
                if direction == 'LONG':
                    runner_exit_r = (last_close - entry_price) / risk
                else:
                    runner_exit_r = (entry_price - last_close) / risk
                runner_exit_r = max(0.0, round(float(runner_exit_r), 3))
            # else: runner stopped at BE → runner_exit_r stays 0.0

        net_r = round(0.5 * 1.0 + 0.5 * runner_exit_r, 3)
        results.append(('WIN', net_r, mae_pct, mfe_pct, True, runner_exit_r))

    return results


# ── FULL DISTRIBUTION STATS HELPER ────────────────────────────────────────────
def _dist_stats(vals_series, n_bins=30):
    """Return full percentile stats + histogram for a pandas Series."""
    vals = vals_series.dropna()
    vals = vals[vals > 0]
    if len(vals) == 0:
        return {}
    counts, edges = np.histogram(vals, bins=n_bins)
    hist = [{'lo': round(float(edges[i]), 4),
             'hi': round(float(edges[i+1]), 4),
             'n':  int(counts[i])} for i in range(len(counts))]
    # mode = midpoint of most-populated bucket
    mode_bucket = max(hist, key=lambda x: x['n'])
    mode_val = round((mode_bucket['lo'] + mode_bucket['hi']) / 2, 4)
    return dict(
        count = int(len(vals)),
        min   = round(float(vals.min()), 4),
        p10   = round(float(vals.quantile(.10)), 4),
        p25   = round(float(vals.quantile(.25)), 4),
        p50   = round(float(vals.quantile(.50)), 4),
        p75   = round(float(vals.quantile(.75)), 4),
        p90   = round(float(vals.quantile(.90)), 4),
        p95   = round(float(vals.quantile(.95)), 4),
        p99   = round(float(vals.quantile(.99)), 4),
        max   = round(float(vals.max()), 4),
        mean  = round(float(vals.mean()), 4),
        std   = round(float(vals.std(ddof=1)), 4) if len(vals) > 1 else 0.0,
        mode  = mode_val,
        hist  = hist,
    )


def _lognorm_fit(vals_np):
    """Return log-normal fit params for a positive numpy array."""
    lv = np.log(vals_np[vals_np > 0])
    if len(lv) < 5:
        return {}
    mu    = float(lv.mean())
    sigma = float(lv.std(ddof=1))
    n     = len(vals_np)
    probs = (np.arange(1, n + 1) - 0.5) / n
    theoretical = np.exp(mu + sigma * _scipy_stats.norm.ppf(probs))
    empirical   = np.sort(vals_np)
    r = float(np.corrcoef(empirical, theoretical)[0, 1]) if len(empirical) > 1 else 0.0
    return dict(
        mu             = round(mu, 4),
        sigma          = round(sigma, 4),
        implied_median = round(float(np.exp(mu)), 4),
        implied_mean   = round(float(np.exp(mu + sigma**2 / 2)), 4),
        implied_mode   = round(float(np.exp(mu - sigma**2)), 4),
        goodness       = round(r, 4),
    )


def _clusters(vals, n):
    """Return 3-tier cluster stats at p33 / p75 breakpoints."""
    p33 = float(vals.quantile(0.33))
    p75 = float(vals.quantile(0.75))
    tiers = [
        (0,   p33,          'Tight',    '0 → p33'),
        (p33, p75,          'Moderate', 'p33 → p75'),
        (p75, float('inf'), 'Wide',     'p75 → max'),
    ]
    out = []
    for lo, hi, label, rng in tiers:
        if hi < float('inf'):
            mask = (vals >= lo) & (vals < hi)
        else:
            mask = vals >= lo
        cv = vals[mask]
        out.append(dict(
            label        = label,
            range        = rng,
            n            = int(len(cv)),
            pct_of_trades= round(len(cv) / max(n, 1) * 100, 1),
            mean         = round(float(cv.mean()), 4)   if len(cv) else 0,
            median       = round(float(cv.median()), 4) if len(cv) else 0,
            max          = round(float(cv.max()), 4)    if len(cv) else 0,
        ))
    return out


def _full_mae_stats(wl, ce=None, n_bins=50):
    """Rich MAE distribution: log-normal fit, clusters, full percentiles, SL sweep."""
    vals = wl['mae_pct'].dropna()
    vals = vals[vals > 0]
    n = len(vals)
    if n < 20:
        return None

    p99_val  = float(vals.quantile(0.99))
    clipped  = vals[vals <= p99_val]
    counts_h, edges_h = np.histogram(clipped, bins=n_bins)
    mode_idx = int(np.argmax(counts_h))
    mode_v   = round((float(edges_h[mode_idx]) + float(edges_h[mode_idx + 1])) / 2, 4)

    pct_levels = [5,10,15,20,25,30,35,40,50,60,65,70,75,80,85,90,95,99]
    percentiles = {f'p{p}': round(float(vals.quantile(p / 100)), 4) for p in pct_levels}

    # SL sweep: for each MAE threshold, compute touch%, P(false stop), P(genuine)
    raw_thresholds = [float(vals.quantile(p / 100)) for p in [5,10,15,20,25,30,35,40,50,60,70,80,90,95]]
    thresholds = sorted({round(t, 4) for t in raw_thresholds})
    sl_sweep = []
    best_opt = None; best_score = -1.0
    for thr in thresholds:
        touched = wl[wl['mae_pct'] >= thr]
        if len(touched) == 0:
            continue
        nt = len(touched)
        n_win  = int((touched['outcome'] == 'WIN').sum())
        n_loss = int((touched['outcome'] == 'LOSS').sum())
        p_rec = round(n_win  / nt, 4)
        p_ko  = round(n_loss / nt, 4)
        exc   = round(nt / n * 100, 1)
        sl_sweep.append(dict(threshold=thr, exceed_pct=exc, p_recovered=p_rec, p_ko=p_ko))
        score = p_ko * exc
        if score > best_score:
            best_score = score; best_opt = thr

    return dict(
        n          = n,
        mean       = round(float(vals.mean()), 4),
        median     = round(float(vals.median()), 4),
        std        = round(float(vals.std(ddof=1)), 4),
        mode       = mode_v,
        skewness   = round(float(_scipy_stats.skew(vals)), 3),
        kurtosis   = round(float(_scipy_stats.kurtosis(vals)), 3),
        percentiles= percentiles,
        lognorm    = _lognorm_fit(vals.values),
        clusters   = _clusters(vals, n),
        histogram  = dict(
            edges  = [round(float(e), 4) for e in edges_h],
            counts = [int(c) for c in counts_h],
        ),
        sl_sweep   = sl_sweep,
        opt_sl     = best_opt,
        ce         = ce,
    )


def _full_mfe_stats(wl, n_bins=50):
    """Rich MFE distribution: log-normal fit, clusters, full percentiles, BE triggers, PTQ."""
    vals = wl['mfe_pct'].dropna()
    vals = vals[vals > 0]
    n = len(vals)
    if n < 20:
        return None

    p99_val  = float(vals.quantile(0.99))
    clipped  = vals[vals <= p99_val]
    counts_h, edges_h = np.histogram(clipped, bins=n_bins)
    mode_idx = int(np.argmax(counts_h))
    mode_v   = round((float(edges_h[mode_idx]) + float(edges_h[mode_idx + 1])) / 2, 4)

    pct_levels = [5,10,15,20,25,30,35,40,50,60,65,70,75,80,85,90,95,99]
    percentiles = {f'p{p}': round(float(vals.quantile(p / 100)), 4) for p in pct_levels}

    # Clusters for MFE use Small / Moderate / Large labels
    p33 = float(vals.quantile(0.33))
    p75 = float(vals.quantile(0.75))
    mfe_tiers = [
        (0,   p33,          'Small',    '0 → p33'),
        (p33, p75,          'Moderate', 'p33 → p75'),
        (p75, float('inf'), 'Large',    'p75 → max'),
    ]
    clusters = []
    for lo, hi, label, rng in mfe_tiers:
        mask = (vals >= lo) & (vals < hi) if hi < float('inf') else (vals >= lo)
        cv = vals[mask]
        clusters.append(dict(
            label=label, range=rng, n=int(len(cv)),
            pct_of_trades=round(len(cv) / max(n, 1) * 100, 1),
            mean=round(float(cv.mean()), 4) if len(cv) else 0,
            median=round(float(cv.median()), 4) if len(cv) else 0,
            max=round(float(cv.max()), 4) if len(cv) else 0,
        ))

    # BE trigger analysis
    net_r_col = 'net_r' if 'net_r' in wl.columns else None
    ev_base   = round(float(wl[net_r_col].mean()), 4) if net_r_col else None
    avg_loss  = float(wl.loc[wl[net_r_col] < 0, net_r_col].mean()) if net_r_col and (wl[net_r_col] < 0).any() else -1.0

    raw_trigs = [float(vals.quantile(p / 100)) for p in [5,10,15,20,25,30,35,40,50,60,70,80,90]]
    triggers  = sorted({round(t, 4) for t in raw_trigs})
    be_triggers = []
    ptq_level = None; ptq_reach_rate = None
    for thr in triggers:
        reached = wl[wl['mfe_pct'] >= thr]
        if len(reached) == 0:
            continue
        nr = len(reached)
        reach_rate = round(nr / n * 100, 1)
        if net_r_col:
            n_pos = int((reached[net_r_col] > 0).sum())
            n_neg = int((reached[net_r_col] <= 0).sum())
        else:
            n_pos = int((reached['outcome'] == 'WIN').sum())
            n_neg = int((reached['outcome'] == 'LOSS').sum())
        p_pos = round(n_pos / nr, 4)
        n_rescued = n_neg
        ev_delta  = round(n_rescued / n * abs(avg_loss), 4)
        new_ev    = round((ev_base or 0) + ev_delta, 4)
        be_triggers.append(dict(
            trigger_pct = thr,
            reach_rate  = reach_rate,
            p_pos_given = p_pos,
            n_rescued   = n_rescued,
            ev_delta    = ev_delta,
            new_ev      = new_ev,
        ))
        if ptq_level is None and p_pos >= 0.50:
            ptq_level = thr; ptq_reach_rate = reach_rate

    return dict(
        n           = n,
        mean        = round(float(vals.mean()), 4),
        median      = round(float(vals.median()), 4),
        std         = round(float(vals.std(ddof=1)), 4),
        mode        = mode_v,
        skewness    = round(float(_scipy_stats.skew(vals)), 3),
        kurtosis    = round(float(_scipy_stats.kurtosis(vals)), 3),
        percentiles = percentiles,
        lognorm     = _lognorm_fit(vals.values),
        clusters    = clusters,
        histogram   = dict(
            edges  = [round(float(e), 4) for e in edges_h],
            counts = [int(c) for c in counts_h],
        ),
        be_triggers    = be_triggers,
        ptq_level      = ptq_level,
        ptq_reach_rate = ptq_reach_rate,
    )


# ── SETUP DETECTOR  (profile-agnostic — no stop/target computed here) ─────────
def detect_setups_base(m1_arrs, s_arrs, c_arrs, model_key, model_cfg,
                       cisd_fast_bars=CISD_FAST_BARS):
    """
    Detect all sweep+CISD setups.  Returns (base_rows, base_pending).

    base_rows    — list of dicts with all metadata *except* stop/target/outcome.
                   Includes `base_risk` = |entry_price − sweep_extreme| which
                   apply_profile_and_resolve uses to scale stop / target.
    base_pending — list of {idx, entry_ts_ns, entry_price, sweep_extreme,
                            base_risk, direction} for every valid entry.
    """
    q1_min    = model_cfg['q1_min']
    min_range = model_cfg['min_range']
    sess_hrs  = model_cfg['session_hrs']
    label     = f"{model_key}_PREV_CISD"

    NS_PER_MIN = np.int64(60_000_000_000)
    q1_ns      = np.int64(q1_min) * NS_PER_MIN
    gap_limit  = np.int64(model_cfg['sweep_tf_min']) * NS_PER_MIN * 3

    s_ts    = s_arrs['ts_ns']
    s_high  = s_arrs['high']
    s_low   = s_arrs['low']
    s_n     = len(s_ts)

    m1_ts   = m1_arrs['ts_ns']
    m1_high = m1_arrs['high']
    m1_low  = m1_arrs['low']
    m1_close= m1_arrs['close']
    m1_hr   = m1_arrs['hr']
    m1_mn   = m1_arrs['mn']
    m1_dow  = m1_arrs['dow']
    m1_date = m1_arrs['trade_date']

    print(f"   [{label}] Scanning {s_n:,} sweep bars ...")

    base_rows    = []
    base_pending = []

    for i in range(1, s_n):
        curr_ts_ns = s_ts[i]

        if curr_ts_ns - s_ts[i - 1] > gap_limit:
            continue
        refs = {
            'SHORT': (s_high[i - 1], s_high[i - 1] - s_low[i - 1], 1),
            'LONG':  (s_low[i - 1],  s_high[i - 1] - s_low[i - 1], 1),
        }

        q1_start_ns = curr_ts_ns
        q1_end_ns   = curr_ts_ns + q1_ns - NS_PER_MIN
        q1_s = int(np.searchsorted(m1_ts, q1_start_ns, side='left'))
        q1_e = int(np.searchsorted(m1_ts, q1_end_ns,   side='right'))
        if q1_e - q1_s < 3:
            continue

        q1_h  = m1_high[q1_s:q1_e]
        q1_l  = m1_low[q1_s:q1_e]
        q1_c  = m1_close[q1_s:q1_e]
        q1_hr = m1_hr[q1_s:q1_e]
        q1_mn = m1_mn[q1_s:q1_e]
        q1_ts = m1_ts[q1_s:q1_e]

        if sess_hrs:
            hrf = q1_hr + q1_mn / 60.0
            if not np.any((hrf >= sess_hrs[0]) & (hrf < sess_hrs[1])):
                continue

        for direction in ('SHORT', 'LONG'):
            ref_level, ref_range, ref_lookback = refs[direction]

            if direction == 'SHORT':
                swept_mask = q1_h > ref_level
            else:
                swept_mask = q1_l < ref_level

            if not swept_mask.any():
                continue

            pos = int(swept_mask.argmax())

            if direction == 'SHORT':
                sweep_extreme = float(q1_h[swept_mask].max())
            else:
                sweep_extreme = float(q1_l[swept_mask].min())

            sweep_ext = abs(sweep_extreme - ref_level)

            rejected_by = ''
            if ref_range < min_range:
                rejected_by = 'F1_SMALL_RANGE'
            elif ref_range > 0 and (sweep_ext / ref_range) < SWEEP_MIN_PCT:
                rejected_by = 'F2_SWEEP_TOO_SMALL'
            elif ref_range > 0 and (sweep_ext / ref_range) > SWEEP_MAX_PCT:
                rejected_by = 'F3_SWEEP_TOO_LARGE'

            post_s = pos + 1
            if post_s >= len(q1_ts):
                continue

            if direction == 'SHORT':
                ret_mask = q1_l[post_s:] <= ref_level
            else:
                ret_mask = q1_h[post_s:] >= ref_level

            if not ret_mask.any():
                continue

            ret_rel   = int(ret_mask.argmax())
            ret_idx   = post_s + ret_rel
            ret_close = float(q1_c[ret_idx])
            ret_ts_ns = int(q1_ts[ret_idx])

            if not rejected_by:
                if direction == 'SHORT' and ret_close > ref_level:
                    rejected_by = 'F4_NO_CLOSE_BACK'
                elif direction == 'LONG' and ret_close < ref_level:
                    rejected_by = 'F4_NO_CLOSE_BACK'

            cisd_ts_ns, cisd_level = find_cisd(
                c_arrs, ret_ts_ns, direction, cisd_fast_bars, 'CISD'
            )

            base_row = dict(
                date          = str(s_arrs['trade_date'][i]),
                yr            = int(s_arrs['yr'][i]),
                dow           = int(s_arrs['dow'][i]),
                direction     = direction,
                ref_range     = round(float(ref_range), 2),
                sweep_ext     = round(float(sweep_ext), 2),
                sweep_pct     = round(sweep_ext / ref_range, 3) if ref_range > 0 else 0,
                sweep_extreme = round(float(sweep_extreme), 2),
                sweep_mode    = 'PREV',
                cisd_mode     = 'CISD',
                ref_lookback  = ref_lookback,
            )

            if cisd_ts_ns is None:
                base_row.update(
                    hr=int(s_arrs['hr'][i]), session='OTHER',
                    entry_price=None, base_risk=None, cisd_level=None,
                    stop_price=None, target_price=None, risk_pts=None,
                    outcome='SKIP', rejected_by=rejected_by or 'NO_CISD',
                    r=0.0, mae_pct=None, mfe_pct=None,
                )
                base_rows.append(base_row)
                continue

            # ── Entry bar: next CISD-TF candle open after CISD fires ──────────
            cisd_c_idx = int(np.searchsorted(c_arrs['ts_ns'], cisd_ts_ns, side='left'))
            next_c_idx = cisd_c_idx + 1
            if next_c_idx >= len(c_arrs['ts_ns']):
                continue

            entry_ts_ns = int(c_arrs['ts_ns'][next_c_idx])
            entry_price = float(c_arrs['open'][next_c_idx])

            entry_start = int(np.searchsorted(m1_ts, entry_ts_ns, side='left'))
            if entry_start >= len(m1_ts):
                continue

            # base_risk = |entry − sweep_extreme| = 1 full R-unit
            base_risk = abs(entry_price - sweep_extreme)

            # Entry must be on the correct side of the sweep extreme
            if direction == 'LONG'  and entry_price <= sweep_extreme:
                continue
            if direction == 'SHORT' and entry_price >= sweep_extreme:
                continue

            hr_val = int(m1_hr[entry_start])
            mn_val = int(m1_mn[entry_start])

            base_row.update(
                date         = str(m1_date[entry_start]),
                dow          = int(m1_dow[entry_start]),
                hr           = hr_val,
                mn           = mn_val,
                session      = get_session(hr_val + mn_val / 60.0),
                entry_price  = round(entry_price, 2),
                base_risk    = round(base_risk, 2),
                cisd_level   = round(cisd_level, 2) if cisd_level is not None else None,
                rejected_by  = rejected_by,
                # profile-dependent fields — filled by apply_profile_and_resolve
                stop_price   = None,
                target_price = None,
                risk_pts     = None,
                outcome      = '',
                r            = 0.0,
                mae_pct      = None,
                mfe_pct      = None,
            )
            base_rows.append(base_row)
            base_pending.append(dict(
                idx           = len(base_rows) - 1,
                entry_ts_ns   = entry_ts_ns,
                entry_price   = entry_price,
                sweep_extreme = float(sweep_extreme),
                base_risk     = base_risk,
                direction     = direction,
            ))

    print(f"      {len(base_pending):,} entries detected across all filters")
    return base_rows, base_pending


# ── PROFILE OUTCOME RESOLVER ──────────────────────────────────────────────────
def apply_profile_and_resolve(base_rows, base_pending, m1_arrs,
                               stop_val, target_val, profile_type='mult'):
    """
    Compute stop / target for each detected setup and resolve outcomes.

    profile_type='mult'  (sweep-relative):
        risk_pts   = stop_val  × base_risk   (base_risk = |entry − sweep_extreme|)
        LONG:  stop = entry − stop_val×base_risk,  target = entry + target_val×base_risk
        SHORT: stop = entry + stop_val×base_risk,  target = entry − target_val×base_risk

    profile_type='pct'  (fixed % of entry price):
        risk_pts   = entry_price × stop_val  / 100
        LONG:  stop = entry × (1 − stop_val/100),   target = entry × (1 + target_val/100)
        SHORT: stop = entry × (1 + stop_val/100),   target = entry × (1 − target_val/100)
    """
    rows = [dict(r) for r in base_rows]   # shallow copy per profile

    profile_pending = []
    for bp in base_pending:
        idx           = bp['idx']
        entry_price   = bp['entry_price']
        base_risk     = bp['base_risk']
        direction     = bp['direction']

        if profile_type == 'pct':
            stop_dist    = entry_price * stop_val   / 100.0
            target_dist  = entry_price * target_val / 100.0
        else:  # 'mult'
            stop_dist    = stop_val   * base_risk
            target_dist  = target_val * base_risk

        risk_pts = stop_dist

        if direction == 'LONG':
            stop_price   = entry_price - stop_dist
            target_price = entry_price + target_dist
        else:
            stop_price   = entry_price + stop_dist
            target_price = entry_price - target_dist

        rows[idx]['stop_price']   = round(stop_price,   2)
        rows[idx]['target_price'] = round(target_price, 2)
        rows[idx]['risk_pts']     = round(risk_pts,     2)

        if risk_pts < MIN_RISK_PTS:
            rows[idx]['outcome']     = 'INVALID'
            rows[idx]['rejected_by'] = rows[idx]['rejected_by'] or 'INVALID_RISK'
            continue

        profile_pending.append(dict(
            idx          = idx,
            entry_ts_ns  = bp['entry_ts_ns'],
            entry_price  = entry_price,
            stop_price   = stop_price,
            target_price = target_price,
            direction    = direction,
        ))

    if profile_type == 'structural':
        # Initialise structural columns so DataFrame always has them
        for r in rows:
            r.setdefault('tp1_hit', False)
            r.setdefault('runner_exit_r', 0.0)
        outcomes = resolve_outcomes_structural(m1_arrs, profile_pending)
        for po, (outcome, r_val, mae_pct, mfe_pct, tp1_hit, runner_exit_r) in zip(profile_pending, outcomes):
            idx = po['idx']
            rows[idx]['outcome']       = outcome
            rows[idx]['r']             = r_val
            rows[idx]['mae_pct']       = mae_pct
            rows[idx]['mfe_pct']       = mfe_pct
            rows[idx]['tp1_hit']       = tp1_hit
            rows[idx]['runner_exit_r'] = runner_exit_r
            if outcome == 'INVALID':
                rows[idx]['rejected_by'] = rows[idx]['rejected_by'] or 'INVALID_RISK'
    else:
        outcomes = resolve_outcomes_vectorised(m1_arrs, profile_pending)
        for po, (outcome, r_val, mae_pct, mfe_pct) in zip(profile_pending, outcomes):
            idx = po['idx']
            rows[idx]['outcome']  = outcome
            rows[idx]['r']        = r_val
            rows[idx]['mae_pct']  = mae_pct
            rows[idx]['mfe_pct']  = mfe_pct
            if outcome == 'INVALID':
                rows[idx]['rejected_by'] = rows[idx]['rejected_by'] or 'INVALID_RISK'

    df = pd.DataFrame(rows) if rows else pd.DataFrame()
    if not df.empty:
        passed = int((df['rejected_by'] == '').sum())
        wl_n   = int(df['outcome'].isin(['WIN','LOSS']).sum())
        print(f"        [{stop_val}:{target_val} {profile_type}]  {passed:,} filtered setups  "
              f"→  {wl_n:,} resolved (WIN/LOSS)")
    return df


# ── STATISTICS ────────────────────────────────────────────────────────────────
def build_model_stats(df_raw, trading_days, model_key, model_cfg,
                      stop_mult=1.0, target_mult=2.0, profile_key='1:2',
                      profile_type='mult'):
    df   = df_raw[df_raw['rejected_by'] == ''].copy()
    wl   = df[df['outcome'].isin(['WIN','LOSS'])].copy()
    wl['win'] = (wl['outcome'] == 'WIN').astype(int)

    # ── T-Spot variant classification ─────────────────────────────────────────
    # Classify each trade into one of 6 T-Spot types based on sweep_pct:
    #   ProTrend  (<0.30) — small sweep through midpoint region
    #   Normal    (0.30–0.80) — standard sweep of prior candle extreme
    #   Expansive (>0.80) — large / deep sweep
    # Combined with direction → 6 variants matching the T-Spot indicator study.
    def _classify_tspot(row):
        sp = row['sweep_pct']
        suffix = 'BULL' if row['direction'] == 'LONG' else 'BEAR'
        if sp < 0.30:   return f'ProTrend_{suffix}'
        if sp < 0.80:   return f'Normal_{suffix}'
        return f'Expansive_{suffix}'

    wl['tspot_type'] = wl.apply(_classify_tspot, axis=1)

    def agg(g):
        n = len(g)
        if n == 0:
            return dict(n=0, wins=0, wr=0, ev=0, pf=0, avg_risk_pts=0, avg_rr=0)
        wins    = int(g['win'].sum())
        wr      = round(wins / n, 4)
        # EV and PF from actual per-trade R (WIN r = actual_rr, LOSS r = -1.0)
        ev      = round(float(g['r'].sum()) / n, 3)
        win_r   = float(g.loc[g['win'] == 1, 'r'].sum())
        loss_r  = float(abs(g.loc[g['win'] == 0, 'r'].sum()))
        pf      = round(win_r / max(loss_r, 0.001), 3)
        avg_rr  = round(win_r / max(wins, 1), 2)
        ar      = round(float(g['risk_pts'].mean()), 1) \
                  if 'risk_pts' in g.columns and g['risk_pts'].notna().any() else 0
        return dict(n=n, wins=wins, wr=wr, ev=ev, pf=pf, avg_risk_pts=ar, avg_rr=avg_rr)

    by_hour = []
    for (hr, direction), g in wl.groupby(['hr', 'direction']):
        s = agg(g)
        if s['n'] >= 5:
            s.update(hr=int(hr), direction=direction,
                     hr_label=HR_LABELS.get(int(hr), f"{int(hr):02d}:00"))
            by_hour.append(s)

    by_session = []
    for (sess, direction), g in wl.groupby(['session', 'direction']):
        s = agg(g); s.update(session=sess, direction=direction)
        by_session.append(s)

    by_dow = []
    for (dow, direction), g in wl.groupby(['dow', 'direction']):
        s = agg(g)
        if s['n'] >= 5:
            s.update(dow=int(dow), dow_name=DOW_NAMES.get(int(dow), '?'),
                     direction=direction)
            by_dow.append(s)

    heatmap = []
    for (hr, dow), g in wl.groupby(['hr', 'dow']):
        s = agg(g)
        if s['n'] >= 3:
            s.update(hr=int(hr), dow=int(dow), dow_name=DOW_NAMES.get(int(dow), '?'))
            heatmap.append(s)

    combos = []
    for (hr, dow, direction), g in wl.groupby(['hr', 'dow', 'direction']):
        s = agg(g)
        if s['n'] >= 6:
            dn = DOW_NAMES.get(int(dow), '?')
            s.update(hr=int(hr), dow=int(dow), dow_name=dn, direction=direction,
                     label=f"{dn} {int(hr):02d}:00 {direction}")
            combos.append(s)
    combos.sort(key=lambda x: x['ev'], reverse=True)

    by_year = []
    for yr, g in wl.groupby('yr'):
        s = agg(g); s.update(yr=int(yr))
        by_year.append(s)

    target_r = target_mult / stop_mult if stop_mult > 0 else 2.0
    h  = target_r / 2
    r_buckets = [
        ('Loss',                  lambda r: r <  0),
        (f'0–{h:.2g}R',           lambda r: (0 <= r) & (r <  h)),
        (f'{h:.2g}–{target_r:.2g}R', lambda r: (h <= r) & (r <  target_r * 1.05)),
        (f'{target_r:.2g}R ✓',    lambda r: (target_r * 0.95 <= r) & (r <= target_r * 1.15)),
        (f'>{target_r:.2g}R',     lambda r: r > target_r * 1.05),
    ]
    df_r   = df[df['outcome'] != 'INVALID'].copy()
    r_hist = [{'bucket': lbl, 'n': int(fn(df_r['r']).sum())} for lbl, fn in r_buckets]

    dir_summary = []
    for direction, g in wl.groupby('direction'):
        s = agg(g); s.update(direction=direction)
        dir_summary.append(s)

    mae = wl['mae_pct'].dropna()
    mfe = wl['mfe_pct'].dropna()
    wins_wl2 = wl[wl['win'] == 1]
    loss_wl2 = wl[wl['win'] == 0]
    # Compact legacy fields (kept for meta tile compatibility)
    mae_dist_legacy = dict(
        median = round(float(mae.median()), 4) if len(mae) else 0,
        p80    = round(float(mae.quantile(.80)), 4) if len(mae) else 0,
        p90    = round(float(mae.quantile(.90)), 4) if len(mae) else 0,
        mean   = round(float(mae.mean()), 4) if len(mae) else 0,
        **{f'ext_{k}': v for k, v in _dist_stats(mae).items()},
    )
    mfe_dist_legacy = dict(
        median = round(float(mfe.median()), 4) if len(mfe) else 0,
        p80    = round(float(mfe.quantile(.80)), 4) if len(mfe) else 0,
        p90    = round(float(mfe.quantile(.90)), 4) if len(mfe) else 0,
        mean   = round(float(mfe.mean()), 4) if len(mfe) else 0,
        **{f'ext_{k}': v for k, v in _dist_stats(mfe).items()},
    )
    # Win-only and loss-only MAE/MFE
    mae_wins_dist = _dist_stats(wins_wl2['mae_pct'].dropna())
    mfe_wins_dist = _dist_stats(wins_wl2['mfe_pct'].dropna())
    mae_loss_dist = _dist_stats(loss_wl2['mae_pct'].dropna())
    mfe_loss_dist = _dist_stats(loss_wl2['mfe_pct'].dropna())

    rp = wl['risk_pts'].dropna()
    risk_dist = dict(
        mean   = round(float(rp.mean()),1)          if len(rp) else 0,
        median = round(float(rp.median()),1)        if len(rp) else 0,
        p25    = round(float(rp.quantile(.25)),1)   if len(rp) else 0,
        p75    = round(float(rp.quantile(.75)),1)   if len(rp) else 0,
        p90    = round(float(rp.quantile(.90)),1)   if len(rp) else 0,
        max    = round(float(rp.max()),1)            if len(rp) else 0,
    )

    overall  = agg(wl)
    ny_setup = df[df['session'].isin(['NY1', 'NY2'])]
    spd      = round(len(ny_setup) / max(trading_days, 1), 2)

    filter_impact = compute_filter_impact(df_raw)

    # ── T-Spot breakdown: stats per variant × DOW × Hour ─────────────────────
    TSPOT_KEYS = ['Normal_BULL','Normal_BEAR','Expansive_BULL',
                  'Expansive_BEAR','ProTrend_BULL','ProTrend_BEAR']
    tspot_breakdown = {}
    for tk in TSPOT_KEYS:
        grp = wl[wl['tspot_type'] == tk]
        if len(grp) == 0:
            tspot_breakdown[tk] = dict(
                overall=dict(n=0, wins=0, wr=0, ev=0, pf=0, avg_risk_pts=0),
                heatmap=[], by_hour=[], by_dow=[], top_combos=[])
            continue

        hm = []
        for (hr, dow), g in grp.groupby(['hr', 'dow']):
            s = agg(g)
            if s['n'] >= 2:
                s.update(hr=int(hr), dow=int(dow), dow_name=DOW_NAMES.get(int(dow), '?'))
                hm.append(s)

        bh = []
        for hr, g in grp.groupby('hr'):
            s = agg(g)
            if s['n'] >= 3:
                s.update(hr=int(hr), hr_label=HR_LABELS.get(int(hr), f'{int(hr):02d}:00'))
                bh.append(s)

        bd = []
        for dow, g in grp.groupby('dow'):
            s = agg(g)
            if s['n'] >= 3:
                s.update(dow=int(dow), dow_name=DOW_NAMES.get(int(dow), '?'))
                bd.append(s)

        tc = []
        for (hr, dow, direction), g in grp.groupby(['hr', 'dow', 'direction']):
            s = agg(g)
            if s['n'] >= 4:
                dn = DOW_NAMES.get(int(dow), '?')
                s.update(hr=int(hr), dow=int(dow), dow_name=dn, direction=direction,
                         label=f"{dn} {int(hr):02d}:00 {direction}")
                tc.append(s)
        tc.sort(key=lambda x: x['ev'], reverse=True)

        tspot_breakdown[tk] = dict(
            overall=agg(grp), heatmap=hm, by_hour=bh, by_dow=bd, top_combos=tc[:10])

    # Most recent 25 resolved trades
    recent_cols = ['date','direction','hr','mn','session','dow','entry_price',
                   'sweep_extreme','target_price','risk_pts','r','outcome']
    recent_rows = (wl[recent_cols]
                   .sort_values('date', ascending=False)
                   .head(40)
                   .copy())
    recent_rows['dow_name'] = recent_rows['dow'].map(lambda d: DOW_NAMES.get(int(d), '?'))
    recent_trades = recent_rows.to_dict('records')
    for t in recent_trades:
        t['date'] = str(t['date'])[:10]
        t['dow']  = int(t['dow'])
        t['hr']   = int(t['hr'])
        t['mn']   = int(t['mn'])
        for k in ('entry_price','sweep_extreme','target_price','risk_pts','r'):
            if t[k] is not None:
                t[k] = round(float(t[k]), 2)

    # ── Risk stats for hero tiles ─────────────────────────────────────────────
    wl_sorted   = wl.sort_values('date')
    outcomes_seq = wl_sorted['win'].tolist()
    def _max_consec(seq, val):
        mx = cur = 0
        for v in seq:
            cur = cur + 1 if v == val else 0
            mx  = max(mx, cur)
        return mx
    rs_max_cw = _max_consec(outcomes_seq, 1)
    rs_max_cl = _max_consec(outcomes_seq, 0)

    rr_actual   = round(target_mult / stop_mult, 4) if stop_mult > 0 else 2.0
    wins_df     = wl_sorted[wl_sorted['win'] == 1]
    losses_df   = wl_sorted[wl_sorted['win'] == 0]
    avg_win_r   = round(float(wins_df['r'].mean()), 4)   if len(wins_df)   else rr_actual
    avg_loss_r  = round(float(losses_df['r'].mean()), 4) if len(losses_df) else -1.0
    avg_win_usd  = round(avg_win_r  * RISK_PER_TRADE, 2)
    avg_loss_usd = round(avg_loss_r * RISK_PER_TRADE, 2)

    # sl_pct = avg (risk_pts / entry_price * 100); tp_pct = sl_pct * rr_actual
    entry_col = wl_sorted['entry_price'].replace(0, np.nan).dropna()
    rp_col    = wl_sorted.loc[entry_col.index, 'risk_pts']
    if len(entry_col):
        sl_pct_val = round(float((rp_col / entry_col * 100).mean()), 4)
        tp_pct_val = round(sl_pct_val * rr_actual, 4)
    else:
        sl_pct_val = None
        tp_pct_val = None

    eq = float(ACCOUNT_SIZE)
    peak_eq = eq
    min_eq = eq
    max_dd_abs = 0.0
    # Daily P&L for Sharpe: accumulate by date
    daily_pnl: dict = {}
    for _, row in wl_sorted.iterrows():
        r_val = row['r']
        if r_val is not None and not np.isnan(r_val):
            trade_pnl = float(r_val) * RISK_PER_TRADE
            eq += trade_pnl
            if eq < min_eq:
                min_eq = eq
            if eq > peak_eq:
                peak_eq = eq
            dd = (peak_eq - eq) / peak_eq if peak_eq > 0 else 0.0
            if dd > max_dd_abs:
                max_dd_abs = dd
            # date bucket for Sharpe
            try:
                trade_date = str(row['date'])[:10]
                if trade_date:
                    daily_pnl[trade_date] = daily_pnl.get(trade_date, 0.0) + trade_pnl
            except Exception:
                pass
    min_eq = round(min_eq, 2)
    max_dd_pct = round(max_dd_abs * 100, 2)
    total_pnl_usd = round(eq - ACCOUNT_SIZE, 2)
    blown  = min_eq <= 0.0

    # Annualised Sharpe (daily returns, 252 trading days)
    if len(daily_pnl) > 1:
        dpnl_arr = np.array(list(daily_pnl.values()))
        sharpe_val = round(float(dpnl_arr.mean() / dpnl_arr.std(ddof=1) * np.sqrt(252)), 2) \
                     if dpnl_arr.std(ddof=1) > 0 else None
    else:
        sharpe_val = None

    risk_stats = {
        'account_size':     ACCOUNT_SIZE,
        'risk_per_trade':   RISK_PER_TRADE,
        'trades':           overall['n'],
        'wins':             overall['wins'],
        'losses':           overall['n'] - overall['wins'],
        'be_count':         0,
        'avg_win_usd':      avg_win_usd,
        'avg_loss_usd':     avg_loss_usd,
        'blown':            blown,
        'min_equity_usd':   min_eq,
        'max_consec_wins':  rs_max_cw,
        'max_consec_losses': rs_max_cl,
        'sl_pct':           sl_pct_val,
        'tp_pct':           tp_pct_val,
        'max_dd_pct':       max_dd_pct,
        'total_pnl_usd':    total_pnl_usd,
        'sharpe':           sharpe_val,
    }

    # CE — Combined Edge: avg(MFE / MAE) for WIN trades
    wins_wl = wl_sorted[wl_sorted['win'] == 1]
    ce_mask = (wins_wl['mae_pct'] > 0) & (wins_wl['mfe_pct'] > 0)
    if ce_mask.sum() > 0:
        ce = round(float((wins_wl.loc[ce_mask, 'mfe_pct'] /
                           wins_wl.loc[ce_mask, 'mae_pct']).mean()), 3)
    else:
        ce = None
    risk_stats['ce'] = ce

    # Rich MAE / MFE distribution studies (FPFVG-style)
    rich_mae = _full_mae_stats(wl, ce=ce)
    rich_mfe = _full_mfe_stats(wl)

    # Bell curve of actual MAE distribution
    mae_vals = wl_sorted['mae_pct'].dropna()
    mae_vals = mae_vals[mae_vals > 0]
    if len(mae_vals) > 1 and mae_vals.std() > 0:
        mae_mu = float(mae_vals.mean())
        mae_sd = float(mae_vals.std(ddof=1))
        mae_np = mae_vals.values
        risk_stats['mae_bell'] = {
            'mean':      round(mae_mu, 4),
            'std':       round(mae_sd, 4),
            'plus_0_5s': round(mae_mu + 0.5 * mae_sd, 4),
            'plus_1s':   round(mae_mu + mae_sd, 4),
            'plus_1_5s': round(mae_mu + 1.5 * mae_sd, 4),
            'plus_2s':   round(mae_mu + 2.0 * mae_sd, 4),
            'cov_mean':  round(float(np.mean(mae_np <= mae_mu)) * 100, 1),
            'cov_0_5s':  round(float(np.mean(mae_np <= mae_mu + 0.5*mae_sd)) * 100, 1),
            'cov_1s':    round(float(np.mean(mae_np <= mae_mu + mae_sd)) * 100, 1),
            'cov_1_5s':  round(float(np.mean(mae_np <= mae_mu + 1.5*mae_sd)) * 100, 1),
            'cov_2s':    round(float(np.mean(mae_np <= mae_mu + 2.0*mae_sd)) * 100, 1),
        }
    else:
        risk_stats['mae_bell'] = None

    full_key = f"{model_key}_PREV_CISD"
    # For structural_dynamic, breakeven WR assumes runner always stops at BE (worst case)
    # BE: wr × 0.5R = (1-wr) × 1R  →  wr = 2/3
    if profile_type == 'structural':
        be_wr = round(2.0 / 3.0, 4)
    else:
        be_wr = round(stop_mult / (stop_mult + target_mult), 4)

    # ── Structural-dynamic extra stats ────────────────────────────────────────
    structural_stats = None
    if profile_type == 'structural' and 'tp1_hit' in wl.columns:
        tp1_rate   = round(float(wl['tp1_hit'].mean()), 4)
        tp1_trades = wl[wl['tp1_hit'] == True]
        runner_col = tp1_trades['runner_exit_r'].dropna()
        runner_ran = runner_col[runner_col > 0]
        structural_stats = {
            'tp1_hit_rate':        tp1_rate,
            'runner_ran_further_rate': round(float(len(runner_ran) / max(len(runner_col), 1)), 4),
            'runner_stats':        _dist_stats(runner_col),
            'runner_ran_stats':    _dist_stats(runner_ran),
        }

    return {
        'meta': {
            'model_key':          model_key,
            'full_key':           full_key,
            'profile_key':        profile_key,
            'profile_type':       profile_type,
            'model_label':        model_cfg['label'],
            'sweep_mode':         'PREV',
            'cisd_mode':          'CISD',
            'cisd_mode_label':    'Body CISD — close through open of last opposing run',
            'instrument':         TABLE.split('_')[0].upper(),
            'date_range':         f"{df['date'].min()} – {df['date'].max()}"
                                  if len(df) else '—',
            'trading_days':       trading_days,
            'total_raw':          len(df_raw),
            'total_wl':           overall['n'],
            'total_expired':      int((df['outcome'] == 'EXPIRED').sum()),
            'win_rate':           overall['wr'],
            'ev_per_trade':       overall['ev'],
            'profit_factor':      overall['pf'],
            'avg_risk_pts':       overall['avg_risk_pts'],
            'setups_per_day_ny':  spd,
            'risk_breakeven_wr':  be_wr,
            'rr_target':          rr_actual,
            'stop_mult':          stop_mult,
            'target_mult':        target_mult,
            **{f'risk_{k}': v for k, v in risk_dist.items()},
            **{f'mae_{k}':  v for k, v in mae_dist_legacy.items()},
            **{f'mfe_{k}':  v for k, v in mfe_dist_legacy.items()},
        },
        'by_hour':       by_hour,
        'by_session':    by_session,
        'by_dow':        by_dow,
        'heatmap':       heatmap,
        'top_combos':    combos[:15],
        'worst_combos':  sorted(combos, key=lambda x: x['ev'])[:5],
        'by_year':       by_year,
        'r_hist':        r_hist,
        'dir_summary':   dir_summary,
        'risk_dist':     risk_dist,
        'filter_impact': filter_impact,
        'lookback_dist':    [],
        'mae_dist':         mae_dist_legacy,
        'mfe_dist':         mfe_dist_legacy,
        'rich_mae':         rich_mae,
        'rich_mfe':         rich_mfe,
        'mae_wins_dist':    mae_wins_dist,
        'mfe_wins_dist':    mfe_wins_dist,
        'mae_loss_dist':    mae_loss_dist,
        'mfe_loss_dist':    mfe_loss_dist,
        'tspot_breakdown':  tspot_breakdown,
        'recent_trades':    recent_trades,
        'risk_stats':       risk_stats,
        'structural_stats': structural_stats,
        'by_classification': _compute_by_classification(wl_sorted),
        'by_tf':            _compute_by_tf(wl, wl_sorted, stop_mult, target_mult,
                                           sl_pct_val, tp_pct_val, agg,
                                           HR_LABELS, DOW_NAMES),
    }


def _compute_by_classification(wl_sorted: 'pd.DataFrame') -> dict:
    """Compute WR / PF / PnL / Sharpe / MaxDD per day classification."""
    if not DATE_CLASSIFICATION or wl_sorted is None or len(wl_sorted) == 0:
        return {}
    cls_order = ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']
    result = {}
    for cls in cls_order:
        dates = {d for d, c in DATE_CLASSIFICATION.items() if c == cls}
        sub = wl_sorted[wl_sorted['date'].astype(str).str[:10].isin(dates)].copy()
        n = len(sub)
        if n == 0:
            result[cls] = None
            continue
        wins   = int((sub['win'] == 1).sum())
        losses = n - wins
        wr     = round(wins / n * 100, 2)
        gross_win  = float(sub.loc[sub['win'] == 1, 'r'].sum()) * RISK_PER_TRADE
        gross_loss = abs(float(sub.loc[sub['win'] == 0, 'r'].sum())) * RISK_PER_TRADE
        pf     = round(gross_win / gross_loss, 3) if gross_loss > 0 else 0.0
        # Equity curve
        eq = float(ACCOUNT_SIZE)
        peak_eq = eq
        max_dd = 0.0
        daily_pnl: dict = {}
        for _, row in sub.iterrows():
            r_val = row['r']
            if r_val is None or np.isnan(float(r_val)):
                continue
            trade_pnl = float(r_val) * RISK_PER_TRADE
            eq += trade_pnl
            if eq > peak_eq:
                peak_eq = eq
            dd = (peak_eq - eq) / peak_eq if peak_eq > 0 else 0.0
            if dd > max_dd:
                max_dd = dd
            try:
                td = str(row['date'])[:10]
                daily_pnl[td] = daily_pnl.get(td, 0.0) + trade_pnl
            except Exception:
                pass
        total_pnl = round(eq - ACCOUNT_SIZE, 0)
        max_dd_pct = round(max_dd * 100, 2)
        blown = eq <= 0
        if len(daily_pnl) > 1:
            dpnl = np.array(list(daily_pnl.values()))
            sharpe = round(float(dpnl.mean() / dpnl.std(ddof=1) * np.sqrt(252)), 2) \
                     if dpnl.std(ddof=1) > 0 else None
        else:
            sharpe = None
        result[cls] = {
            'trades':    n,
            'wins':      wins,
            'losses':    losses,
            'wr':        wr,
            'pf':        pf,
            'pnl':       total_pnl,
            'sharpe':    sharpe,
            'max_dd':    max_dd_pct,
            'blown':     blown,
        }
    return result


def _compute_by_tf(wl_full, wl_sorted_full, stop_mult, target_mult,
                   sl_pct_val, tp_pct_val, agg_fn, hr_labels, dow_names) -> dict:
    """Compute compact hero+chart stats for each sub-timeframe slice."""
    from datetime import date, timedelta
    import pandas as pd

    if wl_full is None or len(wl_full) == 0:
        return {}

    # Determine reference date from data (latest trade date)
    max_date_str = str(wl_sorted_full['date'].max())[:10]
    try:
        ref = date.fromisoformat(max_date_str)
    except Exception:
        return {}

    TF_CUTOFFS = {
        '2y': ref - timedelta(days=730),
        '1y': ref - timedelta(days=365),
        '6m': ref - timedelta(days=182),
        '3m': ref - timedelta(days=91),
        '1m': ref - timedelta(days=30),
    }

    def _slice_stats(wl_sub):
        if len(wl_sub) == 0:
            return None
        wl_sub = wl_sub.copy()
        n     = len(wl_sub)
        wins  = int((wl_sub['win'] == 1).sum())
        wr    = round(wins / n, 4)
        ev    = round(float(wl_sub['r'].sum()) / n, 3)
        win_r = float(wl_sub.loc[wl_sub['win'] == 1, 'r'].sum())
        los_r = float(abs(wl_sub.loc[wl_sub['win'] == 0, 'r'].sum()))
        pf    = round(win_r / max(los_r, 0.001), 3)
        date_range = f"{str(wl_sub['date'].min())[:10]} – {str(wl_sub['date'].max())[:10]}"

        # Equity / risk stats
        ws_sorted = wl_sub.sort_values('date')
        eq = float(ACCOUNT_SIZE); peak_eq = eq; max_dd = 0.0
        daily_pnl: dict = {}
        def _mc(seq, val):
            mx = cur = 0
            for v in seq:
                cur = cur + 1 if v == val else 0; mx = max(mx, cur)
            return mx
        for _, row in ws_sorted.iterrows():
            r_val = row['r']
            if r_val is None or np.isnan(float(r_val)): continue
            tp = float(r_val) * RISK_PER_TRADE
            eq += tp
            if eq > peak_eq: peak_eq = eq
            dd = (peak_eq - eq) / peak_eq if peak_eq > 0 else 0.0
            if dd > max_dd: max_dd = dd
            try:
                td = str(row['date'])[:10]
                daily_pnl[td] = daily_pnl.get(td, 0.0) + tp
            except Exception:
                pass
        outcomes_seq = ws_sorted['win'].tolist()
        mcw = _mc(outcomes_seq, 1); mcl = _mc(outcomes_seq, 0)
        total_pnl = round(eq - ACCOUNT_SIZE, 2)
        max_dd_pct = round(max_dd * 100, 2)
        blown = eq <= 0.0
        if len(daily_pnl) > 1:
            dpnl = np.array(list(daily_pnl.values()))
            sharpe = round(float(dpnl.mean() / dpnl.std(ddof=1) * np.sqrt(252)), 2) \
                     if dpnl.std(ddof=1) > 0 else None
        else:
            sharpe = None
        # CE
        wins_wl = wl_sub[wl_sub['win'] == 1]
        ce_mask = (wins_wl['mae_pct'] > 0) & (wins_wl['mfe_pct'] > 0)
        ce = round(float((wins_wl.loc[ce_mask, 'mfe_pct'] /
                          wins_wl.loc[ce_mask, 'mae_pct']).mean()), 3) \
             if ce_mask.sum() > 0 else None
        # by_hour
        bh = []
        for (hr, direction), g in wl_sub.groupby(['hr', 'direction']):
            s = agg_fn(g)
            if s['n'] >= 3:
                s.update(hr=int(hr), hr_label=hr_labels.get(int(hr), f'{int(hr):02d}:00'))
                bh.append(s)
        # by_session
        def get_sess(hr):
            h = float(hr)
            if 8.5 <= h < 11.5: return 'NY1'
            if 11.5 <= h < 16.0: return 'NY2'
            return 'Other'
        wl_sub2 = wl_sub.copy()
        wl_sub2['session2'] = wl_sub2['hr'].apply(get_sess)
        bs = []
        for (sess, direction), g in wl_sub2.groupby(['session2', 'direction']):
            if sess == 'Other': continue
            s = agg_fn(g); s.update(session=sess, direction=direction); bs.append(s)
        # by_dow
        bd = []
        for (dow, direction), g in wl_sub.groupby(['dow', 'direction']):
            s = agg_fn(g)
            if s['n'] >= 3:
                s.update(dow=int(dow), dow_name=dow_names.get(int(dow), '?'))
                bd.append(s)
        # dir_summary
        ds = []
        for direction, g in wl_sub.groupby('direction'):
            s = agg_fn(g); s['direction'] = direction; ds.append(s)
        # by_year
        wl_sub['year'] = wl_sub['date'].astype(str).str[:4].astype(int)
        by_yr = []
        for yr, g in wl_sub.groupby('year'):
            s = agg_fn(g); s['yr'] = int(yr); by_yr.append(s)
        # r_hist (simplified)
        rr_actual = round(target_mult / stop_mult, 4) if stop_mult > 0 else 2.0
        r_hist = [
            {'bucket': f'-1R (loss)', 'n': n - wins, 'fill': 'loss'},
            {'bucket': f'{rr_actual}R (target)', 'n': wins, 'fill': 'win'},
        ]
        # recent_trades (last 40 in slice)
        recent_cols = ['date','direction','hr','mn','session','dow','entry_price',
                       'sweep_extreme','target_price','risk_pts','r','outcome']
        available = [c for c in recent_cols if c in wl_sub.columns]
        rt = wl_sub[available].sort_values('date', ascending=False).head(40).copy()
        rt['dow_name'] = rt['dow'].map(lambda d: dow_names.get(int(d), '?'))
        recent_trades = rt.to_dict('records')
        for t in recent_trades:
            t['date'] = str(t['date'])[:10]
            for k in ['dow','hr','mn']:
                if k in t: t[k] = int(t[k])
            for k in ['entry_price','sweep_extreme','target_price','risk_pts','r']:
                if k in t and t[k] is not None:
                    t[k] = round(float(t[k]), 2)

        return {
            'meta': {
                'total_wl': n, 'win_rate': wr, 'ev_per_trade': ev,
                'profit_factor': pf, 'date_range': date_range,
                'rr_target': round(target_mult / stop_mult, 4) if stop_mult > 0 else 2.0,
            },
            'risk_stats': {
                'account_size': ACCOUNT_SIZE, 'risk_per_trade': RISK_PER_TRADE,
                'trades': n, 'wins': wins, 'losses': n - wins, 'be_count': 0,
                'blown': blown, 'min_equity_usd': round(eq, 2),
                'max_consec_wins': mcw, 'max_consec_losses': mcl,
                'sl_pct': sl_pct_val, 'tp_pct': tp_pct_val,
                'max_dd_pct': max_dd_pct, 'total_pnl_usd': total_pnl,
                'sharpe': sharpe, 'ce': ce,
            },
            'by_hour':         bh,
            'by_session':      bs,
            'by_dow':          bd,
            'dir_summary':     ds,
            'by_year':         by_yr,
            'r_hist':          r_hist,
            'by_classification': _compute_by_classification(ws_sorted),
            'recent_trades':   recent_trades,
        }

    result = {}
    dates_str = wl_full['date'].astype(str).str[:10]
    for tf_key, cutoff in TF_CUTOFFS.items():
        mask = dates_str >= cutoff.isoformat()
        sub  = wl_full[mask]
        result[tf_key] = _slice_stats(sub)

    return result


def compute_filter_impact(df_all):
    FILTER_ORDER  = ['F1_SMALL_RANGE','F2_SWEEP_TOO_SMALL','F3_SWEEP_TOO_LARGE',
                     'F4_NO_CLOSE_BACK','NO_CISD','INVALID_RISK']
    FILTER_LABELS_MAP = {
        'F1_SMALL_RANGE':    'F1: Prior Range Floor',
        'F2_SWEEP_TOO_SMALL':'F2: Sweep Min Size',
        'F3_SWEEP_TOO_LARGE':'F3: Sweep Max Cap',
        'F4_NO_CLOSE_BACK':  'F4: Close-Back Required',
        'NO_CISD':           'No CISD Formed',
        'INVALID_RISK':      'Invalid Risk',
    }
    def ev_of(df):
        wl = df[df['outcome'].isin(['WIN','LOSS'])].copy()
        if len(wl) == 0: return 0, 0, 0, 0
        wl['win'] = (wl['outcome'] == 'WIN').astype(int)
        wins  = int(wl['win'].sum())
        wr    = wins / len(wl)
        ev    = float(wl['r'].sum()) / len(wl)
        win_r = float(wl.loc[wl['win'] == 1, 'r'].sum())
        pf    = win_r / max(float(abs(wl.loc[wl['win'] == 0, 'r'].sum())), 0.001)
        return round(wr,4), round(ev,3), round(pf,3), len(wl)

    base = df_all[~df_all['outcome'].isin(['SKIP','INVALID'])].copy()
    wr0, ev0, pf0, n0 = ev_of(base)
    results = [dict(label='Baseline (unfiltered)', n=n0, wr=wr0, ev=ev0,
                    pf=pf0, removed=0)]
    remaining = base.copy()
    for fcode in FILTER_ORDER:
        removed = remaining[remaining['rejected_by'] == fcode]
        if len(removed) == 0: continue
        remaining = remaining[remaining['rejected_by'] != fcode]
        wr, ev_v, pf, n = ev_of(remaining)
        results.append(dict(label=FILTER_LABELS_MAP.get(fcode, fcode),
                            filter_code=fcode, n=n, wr=wr, ev=ev_v, pf=pf,
                            removed=len(removed)))
    return results


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    global TABLE, CISD_FAST_BARS
    parser = argparse.ArgumentParser()
    parser.add_argument('--db',             default=str(DB_PATH))
    parser.add_argument('--table',          default=TABLE)
    parser.add_argument('--output',         default=str(OUT_PATH))
    parser.add_argument('--models',         nargs='+', default=list(MODELS.keys()),
                                            choices=list(MODELS.keys()))
    parser.add_argument('--cisd-fast-bars', type=int,  default=CISD_FAST_BARS,
                                            dest='cisd_fast_bars')
    args = parser.parse_args()
    TABLE          = args.table
    CISD_FAST_BARS = args.cisd_fast_bars

    print("\n" + "═"*65)
    print(f"  SWEEP MODEL ENGINE v6.0  ·  {TABLE.upper()}")
    print(f"  Profiles: {', '.join(pk for _, __, pk, ___ in RR_PROFILES)}")
    print(f"  Models: {', '.join(args.models)}")
    print(f"  Sweep mode: PREV  ·  CISD fast bars: {CISD_FAST_BARS}")
    print("═"*65)

    con = connect(args.db)
    df_1m_full, df_1m_rth = load_1m(con, args.table)
    trading_days = df_1m_rth['trade_date'].nunique()

    print("\n[2] Pre-building timeframes ...")
    needed_sweep_tfs = {MODELS[mk]['sweep_tf_min'] for mk in args.models}
    needed_cisd_tfs  = {MODELS[mk]['cisd_tf_min']  for mk in args.models}
    sweep_dfs = {
        tf: resample(df_1m_full if tf >= 1440 else df_1m_rth, tf,
                     "1D" if tf >= 1440 else f"{tf}min")
        for tf in sorted(needed_sweep_tfs)
    }
    cisd_dfs = {
        tf: resample(df_1m_rth, tf, f"{tf}min")
        for tf in sorted(needed_cisd_tfs)
    }

    print("\n[3] Converting to numpy arrays (built once, reused across all runs) ...")
    sweep_arrs   = {tf: df_to_arrays(df)    for tf, df in sweep_dfs.items()}
    cisd_arrs    = {tf: df_to_arrays(df)    for tf, df in cisd_dfs.items()}
    m1_full_arrs = df_1m_to_arrays(df_1m_full)
    m1_rth_arrs  = df_1m_to_arrays(df_1m_rth)
    print("   Done.")

    all_stats    = {}
    summary_rows = []

    for mk in args.models:
        cfg    = MODELS[mk]
        s_arrs = sweep_arrs[cfg['sweep_tf_min']]
        c_arrs = cisd_arrs[cfg['cisd_tf_min']]
        m1     = m1_full_arrs if cfg['sweep_tf_min'] >= 1440 else m1_rth_arrs

        full_key = f"{mk}_PREV_CISD"
        print(f"\n{'─'*65}")
        print(f"  {full_key}  —  {cfg['label']}")
        print(f"{'─'*65}")

        base_rows, base_pending = detect_setups_base(
            m1, s_arrs, c_arrs, mk, cfg, cisd_fast_bars=CISD_FAST_BARS)
        if not base_rows:
            print(f"   ⚠  No setups found")
            continue

        print(f"      Resolving outcomes across {len(RR_PROFILES)} profiles ...")
        model_profiles = {}
        for stop_val, target_val, pk, ptype in RR_PROFILES:
            df_p = apply_profile_and_resolve(
                base_rows, base_pending, m1, stop_val, target_val, ptype)
            if df_p.empty:
                continue
            stats = build_model_stats(
                df_p, trading_days, mk, cfg, stop_val, target_val, pk, ptype)
            model_profiles[pk] = stats

        if not model_profiles:
            continue

        all_stats[full_key] = {'profiles': model_profiles}

        # Summary uses the default profile (1:2)
        def_stats = model_profiles.get(DEFAULT_PROFILE, next(iter(model_profiles.values())))
        m_  = def_stats['meta']
        fi  = def_stats['filter_impact']
        summary_rows.append((
            full_key,
            fi[0].get('ev', 0) if fi else 0,
            fi[0].get('pf', 0) if fi else 0,
            m_['win_rate'], m_['ev_per_trade'], m_['profit_factor'],
            m_['total_wl'], m_['setups_per_day_ny'],
        ))

    # ── Write JSON ─────────────────────────────────────────────────────────────
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, 'w') as f:
        json.dump(all_stats, f, indent=2, default=str)

    # ── Summary table ──────────────────────────────────────────────────────────
    print(f"\n{'═'*78}")
    print(f"  SUMMARY  (profile = {DEFAULT_PROFILE}  ·  Base = unfiltered, Refined = all filters)")
    print(f"{'═'*78}")
    print(f"  {'Key':<24}  {'Base EV':>8}  {'Base PF':>7}  "
          f"{'→ WR':>7}  {'EV':>8}  {'PF':>7}  {'N':>6}  {'SPD':>5}")
    print(f"  {'─'*24}  {'─'*8}  {'─'*7}  {'─'*7}  {'─'*8}  {'─'*7}  {'─'*6}  {'─'*5}")

    prev_group = None
    for row in summary_rows:
        fk, bev, bpf, wr, ev, pf, n, spd = row
        group = '_'.join(fk.split('_')[:-1])
        if prev_group and group != prev_group:
            print()
        prev_group = group

        print(f"  {fk:<24}  {bev:>+8.3f}R  {bpf:>7.3f}  {wr:>6.1%}  "
              f"{ev:>+8.3f}R  {pf:>7.3f}  {n:>6,}  {spd:>5.2f}")

    print(f"{'═'*78}")
    print(f"\n  ✓  Written → {out}")
    print(f"     Open http://localhost:8000/model_dashboard.html\n")


if __name__ == '__main__':
    main()
