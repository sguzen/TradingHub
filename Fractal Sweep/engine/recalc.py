"""
Excel Formula Recalculation Script — Windows/Mac/Linux compatible
Recalculates all formulas in an Excel file using LibreOffice.
No external dependencies beyond LibreOffice being installed.
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

MACRO_FILENAME = "Module1.xba"
RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def get_soffice_cmd():
    """Find the soffice executable on any platform."""
    system = platform.system()

    if system == "Windows":
        candidates = [
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        # Also check PATH
        for path in candidates:
            if os.path.exists(path):
                return path
        # Try PATH
        return "soffice.exe"

    elif system == "Darwin":
        candidates = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
        for path in candidates:
            if os.path.exists(path):
                return path
        return "soffice"

    else:  # Linux
        return "soffice"


def get_macro_dir():
    """Get the LibreOffice macro directory for this platform."""
    system = platform.system()
    if system == "Windows":
        appdata = os.environ.get("APPDATA", "")
        return os.path.join(appdata, "LibreOffice", "4", "user", "basic", "Standard")
    elif system == "Darwin":
        return os.path.expanduser("~/Library/Application Support/LibreOffice/4/user/basic/Standard")
    else:
        return os.path.expanduser("~/.config/libreoffice/4/user/basic/Standard")


def setup_macro():
    """Install the RecalculateAndSave macro into LibreOffice."""
    macro_dir = get_macro_dir()
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    # Already installed
    if os.path.exists(macro_file) and "RecalculateAndSave" in Path(macro_file).read_text():
        return True

    soffice = get_soffice_cmd()

    # Run LibreOffice once headless to create its config dirs
    if not os.path.exists(macro_dir):
        try:
            subprocess.run(
                [soffice, "--headless", "--terminate_after_init"],
                capture_output=True, timeout=30
            )
        except Exception:
            pass
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO, encoding="utf-8")
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """Recalculate all formulas in an Excel file using LibreOffice."""
    if not Path(filename).exists():
        return {"error": f"File not found: {filename}"}

    abs_path = str(Path(filename).absolute())
    soffice = get_soffice_cmd()

    if not setup_macro():
        return {"error": "Failed to set up LibreOffice macro"}

    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    # Add timeout wrapper on Linux only
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd

    try:
        result = subprocess.run(
            cmd, capture_output=True, text=True,
            timeout=timeout + 30  # subprocess timeout slightly longer than LibreOffice timeout
        )
    except FileNotFoundError:
        return {
            "error": (
                f"LibreOffice not found. Please install it from:\n"
                f"  https://www.libreoffice.org/download/libreoffice-fresh/\n"
                f"Looked for: {soffice}"
            )
        }
    except subprocess.TimeoutExpired:
        return {"error": f"Timed out after {timeout} seconds"}

    if result.returncode not in (0, 124):
        stderr = result.stderr.strip() if result.stderr else "No error details available"
        return {"error": f"LibreOffice error:\n{stderr}"}

    # Scan for Excel errors in recalculated file
    try:
        from openpyxl import load_workbook

        wb = load_workbook(filename, data_only=True)
        excel_errors = ["#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                error_details[err].append(f"{sheet_name}!{cell.coordinate}")
                                total_errors += 1
                                break
        wb.close()

        # Count formulas
        wb2 = load_workbook(filename, data_only=False)
        formula_count = sum(
            1 for sn in wb2.sheetnames
            for row in wb2[sn].iter_rows()
            for cell in row
            if cell.value and isinstance(cell.value, str) and cell.value.startswith("=")
        )
        wb2.close()

        out = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "total_formulas": formula_count,
            "error_summary": {},
        }
        for err_type, locations in error_details.items():
            if locations:
                out["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }
        return out

    except Exception as e:
        return {"error": f"Failed to read recalculated file: {str(e)}"}


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
