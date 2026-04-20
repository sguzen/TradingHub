#!/usr/bin/env python3
"""
FVG Phase 2 — Fixed SL/TP Analyzer with Full Trade Log & Charts
================================================================
Same simulation engine as fvg_phase2_grid_search_v3.py.
Instead of sweeping a grid, you supply one or more fixed SL/TP pairs
interactively at the console.

For each pair AND each classification bucket, the script produces:
  • Full trade-by-trade account history (like a broker statement)
  • Equity Curve chart
  • Drawdown chart
  • Win/Loss distribution bar chart
  • Rolling Win Rate chart
  • Cumulative P&L chart
  • Per-trade P&L bar chart (green/red)
  • Summary metrics block (same formulas as v3)

Output: one Excel file per run, one sheet per combo × classification.
Color theme: identical to Phase 1/2 dark palette.
"""

import os, sys, glob, math, logging, random, string
import numpy as np
import pandas as pd
import duckdb
from pathlib import Path
from datetime import datetime, time as dtime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.series import SeriesLabel

logging.basicConfig(level=logging.INFO, format='%(message)s')
log = logging.getLogger(__name__)

# ── PALETTE (identical to Phase 1/2) ─────────────────────────────────────────
DARK_BG    = '0D0F14'
GOLD       = 'F5C842'
TEAL       = '3DD9B3'
RED_CLR    = 'F5504A'
WHITE      = 'E8EAF0'
MUTED      = '7A82A0'
CARD_BG    = '13161E'
RAISED_BG  = '1A1E28'
BORDER_CLR = '252A38'
ORANGE     = 'F5A623'

CLS_COLORS = {
    'DWP':          TEAL,
    'DNP':          GOLD,
    'R1':           '5B9CF6',
    'R2':           ORANGE,
    'Unclassified': MUTED,
    'Combined':     GOLD,
}

CLASSIFICATIONS = ['DWP', 'DNP', 'R1', 'R2', 'Unclassified']
EOD_HOUR, EOD_MINUTE = 16, 59
DEFAULT_ACCOUNT  = 4500
DEFAULT_RISK     = 225
DEFAULT_RETEST_CUTOFF = dtime(10, 0)
DB_PATH = Path(__file__).parent.parent / 'candle_science.duckdb'

# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def fill(c):    return PatternFill('solid', fgColor=c)
def center():   return Alignment(horizontal='center', vertical='center')
def left():     return Alignment(horizontal='left', vertical='center')
def tborder(c=BORDER_CLR):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def wcol(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def hf(sz=11, bold=True,  color=WHITE): return Font(name='Arial', size=sz, bold=bold, color=color)
def cf(sz=10, bold=False, color=WHITE): return Font(name='Arial', size=sz, bold=bold, color=color)
def wc(ws, r, c, v, font=None, fill_=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font          = font
    if fill_:  cell.fill          = fill_
    if align:  cell.alignment     = align
    if border: cell.border        = border
    if fmt:    cell.number_format = fmt
    return cell

def abort(msg):
    print(f"\n❌  ERROR\n{'─'*60}\n{msg}\n{'─'*60}")
    sys.exit(1)

def warn(msg):
    print(f"⚠️  WARNING: {msg}")

# ── FILE DETECTION ────────────────────────────────────────────────────────────
def pick_file(prompt, candidates, label):
    if not candidates:
        abort(f"No {label} files found in the script folder.")
    if len(candidates) == 1:
        print(f"  ✅  Auto-selected {label}: {os.path.basename(candidates[0])}")
        return candidates[0]
    print(f"\n📂  {label} files found:\n")
    for i, f in enumerate(candidates, 1):
        size_kb = os.path.getsize(f) // 1024
        marker  = " ← Phase 1 detected" if 'FVG_Phase1_' in os.path.basename(f) else ""
        print(f"    [{i}]  {os.path.basename(f)}  ({size_kb:,} KB){marker}")
    print()
    while True:
        try:
            idx = int(input(f"  👉  {prompt}: ").strip()) - 1
            if 0 <= idx < len(candidates): return candidates[idx]
            print(f"  ❌  Enter 1–{len(candidates)}")
        except (ValueError, KeyboardInterrupt):
            print("  ❌  Invalid"); sys.exit(0)

def find_files():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print("\n" + "="*60)
    print("  FVG PHASE 2 — Fixed SL/TP Analyzer + Charts")
    print("="*60)
    xlsx_files = sorted([
        f for f in glob.glob(os.path.join(script_dir, '*.xlsx'))
        if not f.endswith('_PHASE2_LEADERBOARD.xlsx')
        and not f.endswith('_BACKTESTER_OUTPUT.xlsx')
        and '_FIXED_SLTP_' not in f
        and not os.path.basename(f).startswith('FVG2_')
        and not os.path.basename(f).startswith('~$')
    ])
    phase1_file = pick_file("Select your PHASE 1 output file", xlsx_files, "Phase 1 xlsx")
    return phase1_file, script_dir

# ── OHLC LOADER (DuckDB) ─────────────────────────────────────────────────────
def load_ohlc(table='nq_1m'):
    if not DB_PATH.exists():
        abort(f"Database not found: {DB_PATH}")
    con = duckdb.connect(str(DB_PATH), read_only=True)
    df = con.execute(f"""
        SELECT
            timezone('America/New_York', timestamp) AS datetime,
            open, high, low, close
        FROM {table}
        WHERE date_part('hour', timezone('America/New_York', timestamp)) BETWEEN 9 AND 16
        ORDER BY timestamp
    """).df()
    con.close()
    df['datetime'] = pd.to_datetime(df['datetime']).dt.tz_localize(None)
    df.sort_values('datetime', inplace=True)
    df.reset_index(drop=True, inplace=True)
    log.info(f"  OHLC loaded: {len(df):,} bars | {df['datetime'].min().date()} to {df['datetime'].max().date()}")
    return df

# ── PHASE 1 TRADE LOADER (identical to v3) ───────────────────────────────────
NEEDED_COLS = {
    'date':        ['date'],
    'trade':       ['trade?', 'trade'],
    'direction':   ['trade dir', 'trade direction', 'direction'],
    'entry_price': ['entry price', 'entry'],
    'fill_time':   ['fill time', 'fill', 'retest time', 'retest'],
    'mae':         ['mae %', 'mae%', 'mae'],
    'mfe':         ['ext mfe %', 'mfe %', 'mfe%', 'mfe'],
}

def _parse_sheet_trades(ws, classification):
    hdr_row, col_map = None, {}
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = {str(ws.cell(r, c).value or '').strip().lower(): c
                    for c in range(1, ws.max_column + 1)}
        found = {}
        for key, aliases in NEEDED_COLS.items():
            for alias in aliases:
                if alias in row_vals: found[key] = row_vals[alias]; break
        if len(found) >= 5: hdr_row = r; col_map = found; break
    if not hdr_row: return [], 0
    trades, skipped = [], 0
    for r in range(hdr_row + 1, ws.max_row + 1):
        if str(ws.cell(r, col_map['trade']).value or '').strip().upper() != 'YES': continue
        try:
            date_val    = ws.cell(r, col_map['date']).value
            direction   = str(ws.cell(r, col_map['direction']).value or '').strip().upper()
            entry_price = float(ws.cell(r, col_map['entry_price']).value)
            retest_raw  = ws.cell(r, col_map['fill_time']).value
            mae         = float(ws.cell(r, col_map['mae']).value)
            mfe         = float(ws.cell(r, col_map['mfe']).value)
            if direction not in ('LONG', 'SHORT'): skipped += 1; continue
            trade_date = date_val.date() if isinstance(date_val, datetime) else pd.to_datetime(str(date_val)).date()
            if isinstance(retest_raw, dtime):
                t = retest_raw
            else:
                parts = [int(x) for x in str(retest_raw).split(':')]
                t = dtime(parts[0], parts[1], parts[2] if len(parts) > 2 else 0)
            trades.append({
                'date': trade_date, 'direction': direction,
                'entry_price': entry_price, 'entry_dt': datetime.combine(trade_date, t),
                'mae': mae, 'mfe': mfe, 'classification': classification,
            })
        except Exception: skipped += 1
    return trades, skipped

def load_phase1_trades(filepath):
    try: wb = load_workbook(filepath, data_only=True)
    except Exception as e: abort(f"Could not open Phase 1 file: {e}")
    all_trades, by_class, total_skipped = [], {cls: [] for cls in CLASSIFICATIONS}, 0
    for cls in CLASSIFICATIONS:
        sheet_name = cls + ' Data'
        if sheet_name not in wb.sheetnames: warn(f"Sheet '{sheet_name}' not found — skipping."); continue
        trades, skipped = _parse_sheet_trades(wb[sheet_name], cls)
        total_skipped  += skipped
        by_class[cls]   = trades
        all_trades.extend(trades)
        log.info(f"  {cls:>14} Data: {len(trades):>5} trades")
    if not all_trades: abort("No valid trade rows found in any classification sheet.")
    if total_skipped: warn(f"{total_skipped} trade rows skipped.")
    log.info(f"  {'TOTAL':>14}      : {len(all_trades):>5} trades")
    return all_trades, by_class

# ── BAR ARRAYS (identical to v3) ─────────────────────────────────────────────
def build_trade_bar_arrays(trades, df_ohlc):
    df_ohlc = df_ohlc.copy()
    df_ohlc['date']   = df_ohlc['datetime'].dt.date
    df_ohlc['time_s'] = df_ohlc['datetime'].dt.time
    eod_time = dtime(EOD_HOUR, EOD_MINUTE)
    trade_arrays, missing = [], 0
    for t in trades:
        day_bars = df_ohlc[
            (df_ohlc['date'] == t['date']) &
            (df_ohlc['datetime'] >= t['entry_dt']) &
            (df_ohlc['time_s'] <= eod_time)
        ].reset_index(drop=True)
        if len(day_bars) == 0: missing += 1; trade_arrays.append(None); continue
        trade_arrays.append({
            'direction':   t['direction'],
            'entry_price': t['entry_price'],
            'highs':       day_bars['high'].values.astype(np.float64),
            'lows':        day_bars['low'].values.astype(np.float64),
            'date':        t['date'],
        })
    if missing: warn(f"{missing} trades had no matching OHLC bars.")
    return trade_arrays

# ── SIMULATION (vectorized, identical to v3) ──────────────────────────────────
def simulate_trade_vec(bars, sl_price, tp_price):
    highs, lows = bars['highs'], bars['lows']
    if bars['direction'] == 'SHORT':
        stop_mask, target_mask = highs >= sl_price, lows <= tp_price
    else:
        stop_mask, target_mask = lows <= sl_price, highs >= tp_price
    stop_hit, target_hit = np.any(stop_mask), np.any(target_mask)
    if not stop_hit and not target_hit: return 'L'
    if stop_hit and not target_hit:     return 'L'
    if target_hit and not stop_hit:     return 'W'
    first_stop, first_target = int(np.argmax(stop_mask)), int(np.argmax(target_mask))
    return 'L' if first_stop <= first_target else 'W'

def simulate_fixed(trade_arrays, sl_pct, tp_pct, account_size, risk_per_trade, sharpe_n=1):
    """
    Runs ONE fixed SL/TP pair across all trade arrays.
    Returns full metrics dict + per-trade log list.
    """
    outcomes, pnl_list, trade_log = [], [], []
    rr = tp_pct / sl_pct

    for bars in trade_arrays:
        if bars is None: continue
        entry = bars['entry_price']
        dir_  = bars['direction']
        if dir_ == 'SHORT':
            sl_price = entry * (1 + sl_pct / 100)
            tp_price = entry * (1 - tp_pct / 100)
        else:
            sl_price = entry * (1 - sl_pct / 100)
            tp_price = entry * (1 + tp_pct / 100)

        result = simulate_trade_vec(bars, sl_price, tp_price)
        outcomes.append(result)
        pnl = risk_per_trade * rr if result == 'W' else -risk_per_trade
        pnl_list.append(pnl)
        trade_log.append({
            'date':      bars['date'],
            'direction': dir_,
            'entry':     round(entry, 2),
            'sl':        round(sl_price, 2),
            'tp':        round(tp_price, 2),
            'result':    result,
            'pnl':       round(pnl, 2),
        })

    n_total = len(outcomes)
    if n_total == 0: return None, []

    n_wins   = outcomes.count('W')
    n_losses = outcomes.count('L')
    wr       = n_wins  / n_total
    lr       = n_losses / n_total

    win_pnls  = [p for p in pnl_list if p > 0]
    loss_pnls = [abs(p) for p in pnl_list if p < 0]

    avg_win_d  = sum(win_pnls)  / len(win_pnls)  if win_pnls  else 0.0
    avg_loss_d = sum(loss_pnls) / len(loss_pnls) if loss_pnls else risk_per_trade
    gross_w    = sum(win_pnls)
    gross_l    = sum(loss_pnls)
    total_pl   = gross_w - gross_l

    avg_win_r  = avg_win_d / avg_loss_d if avg_loss_d > 0 else 0.0
    total_r    = (n_wins * avg_win_r) - (n_losses * 1.0)
    ev_dollar  = (wr * avg_win_d) - (lr * avg_loss_d)
    ev_r       = ev_dollar / avg_loss_d if avg_loss_d > 0 else 0.0
    pf         = gross_w / gross_l if gross_l > 0 else 999.0
    ce         = ev_r * pf
    risk_pct   = risk_per_trade / account_size
    n_bankroll = int(1.0 / risk_pct) if risk_pct > 0 else 20
    ror_pct    = 100.0 if ce <= 0 else min(100.0, ((1-ce)/(1+ce))**n_bankroll * 100)

    try:
        mcl_calc = math.ceil(math.log(n_total) / math.log(1.0/(1.0-wr))) if 0 < wr < 1 else n_losses
    except (ValueError, ZeroDivisionError):
        mcl_calc = n_losses

    r_arr    = np.array([p / risk_per_trade for p in pnl_list])
    std_r    = float(np.std(r_arr, ddof=1)) if len(r_arr) > 1 else 0.0
    sqn      = round((ev_r / std_r) * math.sqrt(n_total), 4) if std_r > 0 else 0.0

    pnl_arr  = np.array(pnl_list, dtype=np.float64)
    mean_pnl = float(np.mean(pnl_arr))
    std_pnl  = float(np.std(pnl_arr, ddof=1)) if len(pnl_arr) > 1 else 0.0
    sharpe   = round((mean_pnl / std_pnl) * math.sqrt(sharpe_n), 4) if std_pnl > 0 else 0.0

    def max_run(lst, target):
        max_s = cur_s = 0
        for o in lst:
            if o == target: cur_s += 1; max_s = max(max_s, cur_s)
            else: cur_s = 0
        return max_s

    max_w_run = max_run(outcomes, 'W')
    max_l_run = max_run(outcomes, 'L')

    equity = peak = low_eq = float(account_size)
    max_dd_d = 0.0
    blown = False
    breaches = 0

    # Build equity curve into trade_log
    for i, p in enumerate(pnl_list):
        equity += p
        if equity > peak: peak = equity
        dd = peak - equity
        if dd > max_dd_d: max_dd_d = dd
        if equity < low_eq: low_eq = equity
        if equity <= 0: blown = True; trade_log[i]['equity'] = round(equity, 2); break
        if equity < account_size: breaches += 1
        trade_log[i]['equity'] = round(equity, 2)

    # Fill equity for any remaining rows if blown early
    for i, t in enumerate(trade_log):
        if 'equity' not in t: t['equity'] = round(equity, 2)

    # Build drawdown series
    eq_vals = [t['equity'] for t in trade_log]
    running_peak = account_size
    for i, eq in enumerate(eq_vals):
        if eq > running_peak: running_peak = eq
        trade_log[i]['drawdown'] = round(running_peak - eq, 2)
        trade_log[i]['drawdown_pct'] = round((running_peak - eq) / running_peak * 100, 4) if running_peak > 0 else 0.0

    # Rolling win rate (20-trade window)
    window = 20
    for i in range(len(trade_log)):
        start = max(0, i - window + 1)
        window_outcomes = outcomes[start:i+1]
        trade_log[i]['rolling_wr'] = round(window_outcomes.count('W') / len(window_outcomes) * 100, 2)

    # Cumulative P&L
    cum_pnl = 0.0
    for i, t in enumerate(trade_log):
        cum_pnl += t['pnl']
        trade_log[i]['cum_pnl'] = round(cum_pnl, 2)

    dd_pct = (max_dd_d / peak * 100) if peak > 0 else 0.0
    drr    = max_dd_d / risk_per_trade if risk_per_trade > 0 else 0.0

    metrics = {
        'sl': round(sl_pct, 4), 'tp': round(tp_pct, 4), 'rr': round(rr, 4),
        'trades': n_total, 'wins': n_wins, 'losses': n_losses,
        'win_pct': round(wr * 100, 4),
        'avg_win': round(avg_win_d, 4), 'avg_loss': round(avg_loss_d, 4),
        'ratio_wl': round(avg_win_d / avg_loss_d, 4) if avg_loss_d > 0 else 0.0,
        'gross_w': round(gross_w, 4), 'gross_l': round(gross_l, 4),
        'total_pl': round(total_pl, 4),
        'largest_win':  round(max(win_pnls)  if win_pnls  else 0.0, 4),
        'largest_loss': round(-max(loss_pnls) if loss_pnls else 0.0, 4),
        'ev_dollar': round(ev_dollar, 4), 'pf': round(pf, 4),
        'ce': round(ce, 6), 'sqn': round(sqn, 4), 'sharpe': sharpe,
        'max_streak': max(max_w_run, max_l_run),
        'max_w_run': max_w_run, 'max_l_run': max_l_run,
        'dd_pct': round(dd_pct, 4), 'drr': round(drr, 4),
        'ror_pct': round(ror_pct, 4), 'ev_r': round(ev_r, 6),
        'total_r': round(total_r, 4), 'mcl_calc': mcl_calc,
        'max_dd_d': round(max_dd_d, 4), 'low_eq': round(low_eq, 4),
        'blown': 'YES' if blown else 'NO', 'breaches': breaches,
    }
    return metrics, trade_log

# ── CHART BUILDER ─────────────────────────────────────────────────────────────
def _style_line(series, hex_color, width=25000, smooth=True):
    """Apply color and width to a line chart series."""
    from openpyxl.drawing.spreadsheet_drawing import SpreadsheetDrawing
    from openpyxl.chart.data_source import NumDataSource
    series.graphicalProperties.line.solidFill = hex_color
    series.graphicalProperties.line.width      = width
    series.smooth = smooth

def _add_charts(ws, trade_log, metrics, account_size, tab_color, label, sl_pct, tp_pct):
    """
    Writes hidden data columns for charts, then adds 6 charts to the sheet.
    Charts are placed to the right of the trade log table.
    Data columns start at column 15 (hidden area).
    """
    n = len(trade_log)
    if n == 0: return

    DATA_START_COL = 15   # hidden chart data area
    CHART_START_COL = 'S' # where charts anchor

    # ── Write chart data (hidden cols) ───────────────────────────────────────
    # Col 15: Trade #  16: Equity  17: Drawdown$  18: Cum P&L
    # Col 19: Win(1)/Loss(0)  20: Rolling WR  21: Per-trade PNL
    ws.cell(1, DATA_START_COL,   'Trade#')
    ws.cell(1, DATA_START_COL+1, 'Equity $')
    ws.cell(1, DATA_START_COL+2, 'Drawdown $')
    ws.cell(1, DATA_START_COL+3, 'Cum P&L $')
    ws.cell(1, DATA_START_COL+4, 'W=1 L=0')
    ws.cell(1, DATA_START_COL+5, 'Rolling WR %')
    ws.cell(1, DATA_START_COL+6, 'Trade P&L $')

    for i, t in enumerate(trade_log):
        row = i + 2
        ws.cell(row, DATA_START_COL,   i + 1)
        ws.cell(row, DATA_START_COL+1, t['equity'])
        ws.cell(row, DATA_START_COL+2, t['drawdown'])
        ws.cell(row, DATA_START_COL+3, t['cum_pnl'])
        ws.cell(row, DATA_START_COL+4, 1 if t['result'] == 'W' else 0)
        ws.cell(row, DATA_START_COL+5, t['rolling_wr'])
        ws.cell(row, DATA_START_COL+6, t['pnl'])

    # Hide the data columns
    for col_idx in range(DATA_START_COL, DATA_START_COL + 7):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True

    max_row = n + 2

    # ── Helper to create a styled LineChart ───────────────────────────────────
    def make_line(title, y_col, color, y_title, width_cm=18, height_cm=10):
        chart = LineChart()
        chart.title  = title
        chart.style  = 2
        chart.y_axis.title = y_title
        chart.x_axis.title = 'Trade #'
        chart.legend = None
        chart.width  = width_cm
        chart.height = height_cm
        chart.y_axis.numFmt = '#,##0'
        data = Reference(ws, min_col=y_col, max_col=y_col, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        s = chart.series[0]
        s.graphicalProperties.line.solidFill = color
        s.graphicalProperties.line.width      = 20000
        s.smooth = True
        # Dark plot area styling via shape properties
        chart.plot_area.graphicalProperties = None
        return chart

    # ── Helper for bar chart ──────────────────────────────────────────────────
    def make_bar(title, y_col, y_title, width_cm=18, height_cm=10):
        chart = BarChart()
        chart.type   = 'col'
        chart.title  = title
        chart.style  = 2
        chart.y_axis.title = y_title
        chart.x_axis.title = 'Trade #'
        chart.legend = None
        chart.width  = width_cm
        chart.height = height_cm
        data = Reference(ws, min_col=y_col, max_col=y_col, min_row=1, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        return chart

    # Chart anchor positions (row, col letter)
    anchors = ['S2', 'S22', 'S42', 'AM2', 'AM22', 'AM42']

    # 1. Equity Curve
    c1 = make_line(
        f'Equity Curve — {label} | SL {sl_pct:.2f}% TP {tp_pct:.2f}%',
        DATA_START_COL + 1, TEAL, 'Account Value $')
    ws.add_chart(c1, anchors[0])

    # 2. Drawdown Curve (underwater — negative visual via negative values not needed, raw dd$)
    c2 = make_line(
        f'Drawdown $ — {label}',
        DATA_START_COL + 2, RED_CLR, 'Drawdown $')
    ws.add_chart(c2, anchors[1])

    # 3. Cumulative P&L
    c3 = make_line(
        f'Cumulative P&L — {label}',
        DATA_START_COL + 3, GOLD, 'Cumulative P&L $')
    ws.add_chart(c3, anchors[2])

    # 4. Rolling Win Rate
    c4 = make_line(
        f'Rolling Win Rate (20-trade window) — {label}',
        DATA_START_COL + 5, ORANGE, 'Win Rate %')
    ws.add_chart(c4, anchors[3])

    # 5. Per-trade P&L bar (green/red differentiation via conditional is not native in openpyxl
    #    so we use a bar chart — Excel will auto-color based on positive/negative)
    c5 = make_bar(
        f'Per-Trade P&L — {label}',
        DATA_START_COL + 6, 'P&L $')
    ws.add_chart(c5, anchors[4])

    # 6. Win/Loss Distribution
    # Compute distribution data inline in a separate small area
    dist_col = DATA_START_COL + 8
    ws.cell(1, dist_col,   'Outcome')
    ws.cell(2, dist_col,   'Wins')
    ws.cell(3, dist_col,   'Losses')
    ws.cell(1, dist_col+1, 'Count')
    ws.cell(2, dist_col+1, metrics['wins'])
    ws.cell(3, dist_col+1, metrics['losses'])
    ws.column_dimensions[get_column_letter(dist_col)].hidden   = True
    ws.column_dimensions[get_column_letter(dist_col+1)].hidden = True

    c6 = BarChart()
    c6.type   = 'col'
    c6.title  = f'Win/Loss Distribution — {label}'
    c6.style  = 2
    c6.y_axis.title = 'Count'
    c6.legend = None
    c6.width  = 18
    c6.height = 10
    cats = Reference(ws, min_col=dist_col,   min_row=2, max_row=3)
    data = Reference(ws, min_col=dist_col+1, min_row=1, max_row=3)
    c6.add_data(data, titles_from_data=True)
    c6.set_categories(cats)
    ws.add_chart(c6, anchors[5])


# ── WRITE ONE CLASSIFICATION SHEET ───────────────────────────────────────────
def write_analysis_sheet(wb, sheet_title, tab_color, metrics, trade_log,
                          account_size, risk_per_trade, sl_pct, tp_pct,
                          label, is_first=False):
    ws = wb.active if is_first else wb.create_sheet(sheet_title)
    ws.title = sheet_title
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color

    _bdr = tborder()
    _aln = center()
    _aln_l = left()

    # ── Banner ────────────────────────────────────────────────────────────────
    ws.merge_cells('A1:M1')
    wc(ws, 1, 1,
       f'{label}  |  SL {sl_pct:.2f}%  TP {tp_pct:.2f}%  RR {metrics["rr"]:.2f}  '
       f'|  {metrics["trades"]} Trades  |  {datetime.now().strftime("%Y-%m-%d %H:%M")}',
       font=hf(13, True, tab_color), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 30

    # ── Summary metrics block (2 columns × multiple rows) ────────────────────
    SUMMARY = [
        ('── PERFORMANCE ──', None,        True),
        ('Trades',      metrics['trades'],  False),
        ('Wins',        metrics['wins'],    False),
        ('Losses',      metrics['losses'],  False),
        ('Win %',       f"{metrics['win_pct']:.2f}%", False),
        ('Total P&L $', f"${metrics['total_pl']:,.2f}", False),
        ('Avg Win $',   f"${metrics['avg_win']:,.2f}",  False),
        ('Avg Loss $',  f"${metrics['avg_loss']:,.2f}", False),
        ('Largest Win', f"${metrics['largest_win']:,.2f}", False),
        ('Largest Loss',f"${metrics['largest_loss']:,.2f}", False),
        ('── EDGE ──',  None,               True),
        ('EV R',        f"{metrics['ev_r']:.4f}", False),
        ('EV $',        f"${metrics['ev_dollar']:,.2f}", False),
        ('Profit Factor', f"{metrics['pf']:.3f}", False),
        ('Comb Edge',   f"{metrics['ce']:.6f}", False),
        ('SQN',         f"{metrics['sqn']:.4f}", False),
        ('Sharpe',      f"{metrics['sharpe']:.4f}", False),
        ('── RISK ──',  None,               True),
        ('RoR %',       f"{metrics['ror_pct']:.4f}%", False),
        ('Max DD $',    f"${metrics['max_dd_d']:,.2f}", False),
        ('DD %',        f"{metrics['dd_pct']:.2f}%", False),
        ('DRR',         f"{metrics['drr']:.2f}", False),
        ('Low Eq $',    f"${metrics['low_eq']:,.2f}", False),
        ('Blown',       metrics['blown'],   False),
        ('Breaches',    metrics['breaches'],False),
        ('Max W Run',   metrics['max_w_run'], False),
        ('Max L Run',   metrics['max_l_run'], False),
        ('MCL',         metrics['mcl_calc'], False),
        ('── COMBO ──', None,               True),
        ('SL %',        f"{metrics['sl']:.4f}%", False),
        ('TP %',        f"{metrics['tp']:.4f}%", False),
        ('RR',          f"{metrics['rr']:.2f}", False),
        ('Account $',   f"${account_size:,.0f}", False),
        ('Risk/Trade $',f"${risk_per_trade:,.0f}", False),
    ]

    wcol(ws, 1, 18); wcol(ws, 2, 14)
    for i, (lbl, val, is_section) in enumerate(SUMMARY):
        row = i + 3
        bg  = '1C2333' if is_section else (RAISED_BG if i % 2 == 0 else CARD_BG)
        lfc = GOLD if is_section else MUTED
        vfc = GOLD if is_section else WHITE

        c1 = ws.cell(row=row, column=1, value=lbl)
        c1.font = Font(name='Arial', size=9, bold=is_section, color=lfc)
        c1.fill = PatternFill('solid', fgColor=bg)
        c1.alignment = _aln_l; c1.border = _bdr

        c2 = ws.cell(row=row, column=2, value=val if val is not None else '')
        c2.font = Font(name='Arial', size=9, bold=is_section, color=vfc)
        c2.fill = PatternFill('solid', fgColor=bg)
        c2.alignment = _aln; c2.border = _bdr
        ws.row_dimensions[row].height = 15

    # ── Trade Log Header ──────────────────────────────────────────────────────
    LOG_HEADERS = [
        '#', 'Date', 'Dir', 'Entry $', 'SL $', 'TP $',
        'Result', 'P&L $', 'Equity $', 'DD $', 'DD %', 'Cum P&L $', 'Rolling WR %'
    ]
    LOG_WIDTHS  = [5, 12, 6, 11, 11, 11, 8, 11, 12, 11, 9, 13, 14]

    header_row = 3
    for c, (h, w) in enumerate(zip(LOG_HEADERS, LOG_WIDTHS), 4):
        wc(ws, header_row, c, h,
           font=hf(9, True, DARK_BG), fill_=fill(tab_color),
           align=center(), border=tborder(DARK_BG))
        wcol(ws, c, w)
    ws.row_dimensions[header_row].height = 20

    # ── Trade Log Rows ────────────────────────────────────────────────────────
    for i, t in enumerate(trade_log):
        row = i + 4
        bg  = RAISED_BG if i % 2 == 0 else CARD_BG
        is_win = t['result'] == 'W'
        res_fc = TEAL if is_win else RED_CLR
        pnl_fc = TEAL if t['pnl'] >= 0 else RED_CLR
        eq_fc  = TEAL if t['equity'] >= account_size else RED_CLR

        def cell(col, val, fc=WHITE, bold=False, fmt=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(name='Arial', size=9, color=fc, bold=bold)
            c.fill      = PatternFill('solid', fgColor=bg)
            c.alignment = _aln
            c.border    = _bdr
            if fmt: c.number_format = fmt

        cell(4,  i + 1,            fc=GOLD if i < 1 else MUTED)
        cell(5,  str(t['date']),   fc=MUTED)
        cell(6,  t['direction'],   fc=TEAL if t['direction']=='LONG' else ORANGE)
        cell(7,  t['entry'],       fmt='#,##0.00')
        cell(8,  t['sl'],          fmt='#,##0.00', fc=RED_CLR)
        cell(9,  t['tp'],          fmt='#,##0.00', fc=TEAL)
        cell(10, t['result'],      fc=res_fc, bold=True)
        cell(11, t['pnl'],         fc=pnl_fc, bold=True, fmt='#,##0.00')
        cell(12, t['equity'],      fc=eq_fc,  fmt='#,##0.00')
        cell(13, t['drawdown'],    fc=RED_CLR if t['drawdown']>0 else TEAL, fmt='#,##0.00')
        cell(14, t['drawdown_pct'],fc=RED_CLR if t['drawdown_pct']>0 else TEAL, fmt='0.00')
        cell(15, t['cum_pnl'],     fc=TEAL if t['cum_pnl']>=0 else RED_CLR, fmt='#,##0.00')
        cell(16, t['rolling_wr'],  fc=TEAL if t['rolling_wr']>=50 else RED_CLR, fmt='0.00')
        ws.row_dimensions[row].height = 13

    ws.freeze_panes = 'E4'
    ws.auto_filter.ref = f'D3:{get_column_letter(16)}3'

    # ── Add Charts ────────────────────────────────────────────────────────────
    _add_charts(ws, trade_log, metrics, account_size, tab_color, label, sl_pct, tp_pct)

    return ws


# ── CONSOLE INPUT ─────────────────────────────────────────────────────────────
def get_sl_tp_pairs():
    print()
    print('  ── SL / TP Input ─────────────────────────────────────────')
    print('  Enter each SL/TP pair as two numbers separated by a space.')
    print('  Values are in % (e.g.  0.25 0.50 means SL=0.25% TP=0.50%)')
    print('  Type DONE when finished.')
    print()
    pairs = []
    while True:
        try:
            raw = input(f'  Pair {len(pairs)+1}: ').strip()
            if raw.lower() in ('done', 'd', ''):
                if not pairs:
                    print('  ❌  Enter at least one pair.')
                    continue
                break
            parts = raw.replace(',', ' ').split()
            if len(parts) != 2: print('  ❌  Enter exactly two numbers.'); continue
            sl, tp = float(parts[0]), float(parts[1])
            if sl <= 0 or tp <= 0: print('  ❌  Both must be > 0.'); continue
            pairs.append((round(sl, 4), round(tp, 4)))
            print(f'  ✅  Added SL={sl:.4f}%  TP={tp:.4f}%  RR={tp/sl:.2f}')
        except ValueError:
            print('  ❌  Invalid numbers.')
        except KeyboardInterrupt:
            print('\n  Cancelled.'); sys.exit(0)
    return pairs


def get_account_settings():
    print()
    print('  ── Account Settings ──────────────────────────────────────')
    while True:
        try:
            val = input(f'  Account size $ (Enter for ${DEFAULT_ACCOUNT:,}): ').strip()
            account_size = float(val.replace(',','').replace('$','')) if val else DEFAULT_ACCOUNT
            if account_size > 0: break
            print('  ❌  Must be > 0')
        except ValueError: print('  ❌  Invalid')

    while True:
        try:
            val = input(f'  Risk per trade $ (Enter for ${DEFAULT_RISK:,}): ').strip()
            risk_per_trade = float(val.replace(',','').replace('$','')) if val else DEFAULT_RISK
            if 0 < risk_per_trade < account_size: break
            print(f'  ❌  Must be between 0 and ${account_size:,}')
        except ValueError: print('  ❌  Invalid')

    return account_size, risk_per_trade


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    phase1_file, script_dir = find_files()

    sl_tp_pairs    = get_sl_tp_pairs()
    account_size, risk_per_trade = get_account_settings()

    print(f'\n  Pairs    : {sl_tp_pairs}')
    print(f'  Account  : ${account_size:,}  Risk: ${risk_per_trade:,}')
    confirm = input('\n  Ready to run? (Enter to start, Q to quit): ').strip().lower()
    if confirm == 'q': print('\n  Cancelled.'); sys.exit(0)

    print('\nLoading Phase 1 trades...')
    all_trades, by_class = load_phase1_trades(phase1_file)

    print('\nLoading OHLC data...')
    df_ohlc = load_ohlc('nq_1m')

    print('\nBuilding trade bar arrays...')
    combined_arrays = build_trade_bar_arrays(all_trades, df_ohlc)
    class_arrays    = {cls: build_trade_bar_arrays(by_class[cls], df_ohlc)
                       for cls in CLASSIFICATIONS}

    wb    = Workbook()
    first = True

    for sl_pct, tp_pct in sl_tp_pairs:
        print(f'\nRunning SL={sl_pct:.4f}%  TP={tp_pct:.4f}%...')

        buckets = [('Combined', all_trades, combined_arrays)] + [
            (cls, by_class[cls], class_arrays[cls]) for cls in CLASSIFICATIONS
        ]

        for label, trades, arrays in buckets:
            n_unique = max(len(set(t['date'] for t in trades)), 1)
            metrics, trade_log = simulate_fixed(
                arrays, sl_pct, tp_pct, account_size, risk_per_trade, sharpe_n=n_unique)

            if metrics is None:
                log.info(f'  [{label}] No valid trades — skipping.')
                continue

            sheet_title = f'{label} SL{sl_pct:.2f} TP{tp_pct:.2f}'
            tab_color   = CLS_COLORS.get(label, GOLD)

            log.info(f'  [{label}] {metrics["trades"]} trades | '
                     f'WR={metrics["win_pct"]:.1f}% | '
                     f'PL=${metrics["total_pl"]:,.0f} | '
                     f'Sharpe={metrics["sharpe"]} | '
                     f'Blown={metrics["blown"]}')

            write_analysis_sheet(
                wb, sheet_title, tab_color, metrics, trade_log,
                account_size, risk_per_trade, sl_pct, tp_pct,
                label, is_first=first)
            first = False

    ts_short = datetime.now().strftime('%m%d_%H%M')
    uid      = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    outname  = f'FVG2_FIXED_{ts_short}_{uid}.xlsx'
    outpath  = os.path.join(script_dir, outname)
    wb.save(outpath)

    print(f'\n{"="*60}')
    print(f'  DONE')
    print(f'  Pairs    : {len(sl_tp_pairs)}')
    print(f'  Output   : {outname}')
    print(f'{"="*60}\n')


if __name__ == '__main__':
    main()
