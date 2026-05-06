"""
================================================================================
  MASTER BACKTESTER SCRIPT
  -------------------------
  Takes any data collection Excel file and produces a fully calculated,
  ranked backtester output using your trusted TV Backtester formulas.

  HOW TO USE:
    1. Run:  python master_backtester.py
    2. Script scans the folder and asks you which file to use
    3. Enter your account size and risk per trade when prompted
    4. Open the output file — BEST PROFILES sheet has your ranked results

  OUTPUT SHEETS:
    1. BEST PROFILES        — Ranked survivors, color coded, ready to use
    2. [Strategy] Combos    — All combos as columns, all formulas intact
    3. Metadata             — Grading thresholds (copied from template)
    4. Raw Data             — Your original data collection, untouched

  REQUIREMENTS:
    pip install openpyxl
    LibreOffice must be installed (for formula recalculation)
================================================================================
"""

import os
import re
import sys
import copy
import shutil
import subprocess
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── ERROR HANDLING — defined first so everything below can use them ───────────
def abort(msg):
    print(f"\n\u274c  ERROR\n{chr(8212)*60}\n{msg}\n{chr(8212)*60}")
    sys.exit(1)

def warn(msg):
    print(f"\u26a0\ufe0f  WARNING: {msg}")


TEMPLATE_FILE  = 'Copy_of_TV_Backtester__1_.xlsx'
TEMPLATE_SHEET = 'Indicator TimeCandle TimeHorizo'
TRADING_DAYS   = 365

# ── INTERACTIVE SETUP ─────────────────────────────────────────────────────────
def interactive_setup():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    print("\n" + "="*60)
    print("  MASTER BACKTESTER")
    print("="*60)

    # Find all xlsx files in same folder, exclude template and output files
    all_xlsx = [
        f for f in os.listdir(script_dir)
        if f.endswith('.xlsx')
        and f != TEMPLATE_FILE
        and not f.endswith('_BACKTESTER_OUTPUT.xlsx')
    ]

    if not all_xlsx:
        abort(
            f"No Excel files found in the script folder: {script_dir}\n"
            f"Please place your data collection file in the same folder as this script."
        )

    # Ask user to pick data file
    print(f"\n📂  Excel files found in this folder:\n")
    for i, f in enumerate(all_xlsx, 1):
        size_kb = os.path.getsize(os.path.join(script_dir, f)) // 1024
        print(f"    [{i}]  {f}  ({size_kb:,} KB)")

    print()
    while True:
        try:
            choice = input("  👉  Enter the number of your data collection file: ").strip()
            idx = int(choice) - 1
            if 0 <= idx < len(all_xlsx):
                data_file = os.path.join(script_dir, all_xlsx[idx])
                print(f"\n  ✅  Selected: {all_xlsx[idx]}")
                break
            else:
                print(f"  ❌  Please enter a number between 1 and {len(all_xlsx)}")
        except ValueError:
            print(f"  ❌  Please enter a valid number")

    # Auto-detect or ask for template
    template_path = os.path.join(script_dir, TEMPLATE_FILE)
    if os.path.exists(template_path):
        print(f"  ✅  Template : {TEMPLATE_FILE} (auto-detected)")
    else:
        remaining = [f for f in all_xlsx if os.path.join(script_dir, f) != data_file]
        if not remaining:
            abort(
                f"Only one Excel file found and you selected it as the data file.\n"
                f"Please also place your TV Backtester template in the same folder.\n"
                f"Expected template name: \"{TEMPLATE_FILE}\""
            )
        print(f"\n  ⚠️  Template \"{TEMPLATE_FILE}\" not found.")
        print(f"  Please select your TV Backtester template from the list below:\n")
        for i, f in enumerate(remaining, 1):
            size_kb = os.path.getsize(os.path.join(script_dir, f)) // 1024
            print(f"    [{i}]  {f}  ({size_kb:,} KB)")
        print()
        while True:
            try:
                choice = input("  👉  Enter the number of your TV BACKTESTER TEMPLATE file: ").strip()
                idx = int(choice) - 1
                if 0 <= idx < len(remaining):
                    template_path = os.path.join(script_dir, remaining[idx])
                    print(f"\n  ✅  Template: {remaining[idx]}")
                    break
                else:
                    print(f"  ❌  Please enter a number between 1 and {len(remaining)}")
            except ValueError:
                print(f"  ❌  Please enter a valid number")

    # Account size
    print()
    while True:
        try:
            val = input("  👉  Account size in $ (press Enter for default $4,500): ").strip()
            if val == '':
                account_size = 4500
            else:
                account_size = float(val.replace(',', '').replace('$', ''))
            if account_size <= 0:
                print("  ❌  Account size must be greater than 0")
                continue
            print(f"  ✅  Account size: ${account_size:,.2f}")
            break
        except ValueError:
            print("  ❌  Please enter a valid number (e.g. 4500 or 10000)")

    # Risk per trade
    print()
    while True:
        try:
            val = input("  👉  Risk per trade in $ (press Enter for default $225): ").strip()
            if val == '':
                risk_per_trade = 225
            else:
                risk_per_trade = float(val.replace(',', '').replace('$', ''))
            if risk_per_trade <= 0:
                print("  ❌  Risk per trade must be greater than 0")
                continue
            if risk_per_trade >= account_size:
                print(f"  ❌  Risk per trade must be less than account size (${account_size:,.2f})")
                continue
            pct = risk_per_trade / account_size * 100
            print(f"  ✅  Risk per trade: ${risk_per_trade:,.2f} ({pct:.1f}% of account)")
            break
        except ValueError:
            print("  ❌  Please enter a valid number (e.g. 225)")

    # Confirm before running
    output_file = data_file.replace('.xlsx', '_BACKTESTER_OUTPUT.xlsx')
    print(f"\n{'─'*60}")
    print(f"  Data file    : {os.path.basename(data_file)}")
    print(f"  Template     : {os.path.basename(template_path)}")
    print(f"  Account      : ${account_size:,.2f}")
    print(f"  Risk/trade   : ${risk_per_trade:,.2f}")
    print(f"  Output file  : {os.path.basename(output_file)}")
    print(f"{'─'*60}\n")

    confirm = input("  👉  Ready to run? (press Enter to start, or Q to quit): ").strip().lower()
    if confirm == 'q':
        print("\n  Cancelled.\n")
        sys.exit(0)

    return data_file, template_path, output_file, account_size, risk_per_trade

if __name__ == '__main__':
    DATA_FILE, TEMPLATE_FILE, OUTPUT_FILE, ACCOUNT_SIZE, RISK_PER_TRADE = interactive_setup()
else:
    DATA_FILE = TEMPLATE_FILE = OUTPUT_FILE = None
    ACCOUNT_SIZE = RISK_PER_TRADE = 0

# ── COLUMN NAME ALIASES — script finds these regardless of order/case ─────────
COLUMN_ALIASES = {
    'sl':       ['sl %', 'sl', 'stop loss', 'stop loss %', 'sl%', 'stoploss', 'stop_loss'],
    'tp':       ['tp %', 'tp', 'take profit', 'take profit %', 'tp%', 'takeprofit', 'take_profit'],
    'trades':   ['trades', 'total trades', 'trade count', '# trades', 'num trades', 'total_trades'],
    'wins':     ['wins', 'win', 'total wins', 'w', 'total_wins'],
    'losses':   ['losses', 'loss', 'total losses', 'l', 'total_losses'],
    'be':       ['be', 'breakeven', 'breakevens', 'break even', 'break_even', 'b/e'],
    'avg_win':  ['avg win $', 'avg win', 'average win', 'avg w $', 'avgwin', 'avg_win', 'avg win$'],
    'avg_loss': ['avg loss $', 'avg loss', 'average loss', 'avg l $', 'avgloss', 'avg_loss', 'avg loss$'],
    'blown':    ['blown', 'blown?', 'account blown', 'bust', 'blownup', 'blown up'],
    'low_eq':   ['low eq $', 'low eq', 'low equity', 'lowest equity', 'min equity', 'loweq', 'low_eq', 'low eq$'],
    'mcw':      ['max w run', 'mcw', 'max win run', 'max consec wins', 'max win streak', 'max_w_run', 'maxwrun'],
    'mcl':      ['max l run', 'mcl', 'max loss run', 'max consec losses', 'max loss streak', 'max_l_run', 'maxlrun'],
}

FRIENDLY_NAMES = {
    'sl':       'SL % (Stop Loss)',
    'tp':       'TP % (Take Profit)',
    'trades':   'Trades (Total trades)',
    'wins':     'Wins (Total winning trades)',
    'losses':   'Losses (Total losing trades)',
    'be':       'BE (Breakeven trades)',
    'avg_win':  'Avg Win $ (Average winning trade $)',
    'avg_loss': 'Avg Loss $ (Average losing trade $)',
    'blown':    'Blown (YES/NO)',
    'low_eq':   'Low Eq $ (Lowest equity point)',
    'mcw':      'Max W Run (Max consecutive wins)',
    'mcl':      'Max L Run (Max consecutive losses)',
}

# ── STYLING HELPERS ───────────────────────────────────────────────────────────
GRADE_COLORS = {
    'A': ('1E7E34', 'FFFFFF'),
    'B': ('28A745', 'FFFFFF'),
    'C': ('FFC107', '000000'),
    'D': ('FD7E14', 'FFFFFF'),
    'F': ('DC3545', 'FFFFFF'),
}
GRADE_RANK = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'F': 1}

def hfill(h):        return PatternFill('solid', fgColor=h)
def tborder():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def gr(g):           return GRADE_RANK.get(str(g).strip(), 0) if g else 0
def safe(v, digits=4): return round(v, digits) if isinstance(v, float) else v

HDR_HEX = '1B2631'

# ── ERROR HANDLING ────────────────────────────────────────────────────────────
print("\n🔍  Checking files...")

if not os.path.exists(DATA_FILE):
    abort(
        f"Could not find your data collection file: '{DATA_FILE}'\n"
        f"Please check the filename is correct and the file is in the same folder as this script."
    )

if not os.path.exists(TEMPLATE_FILE):
    abort(
        f"Could not find the backtester template file: '{TEMPLATE_FILE}'\n"
        f"Please make sure 'Copy_of_TV_Backtester__1_.xlsx' is in the same folder as this script."
    )

try:
    wb_data = load_workbook(DATA_FILE, data_only=True)
except Exception:
    abort(
        f"Could not open your data file: '{DATA_FILE}'\n"
        f"Please make sure the file is not currently open in Excel and is a valid .xlsx file."
    )

try:
    wb_tmpl = load_workbook(TEMPLATE_FILE)
    wb_tmpl_data = load_workbook(TEMPLATE_FILE, data_only=True)
except Exception:
    abort(
        f"Could not open the backtester template: '{TEMPLATE_FILE}'\n"
        f"Please make sure it is not open in Excel and is a valid .xlsx file."
    )

if TEMPLATE_SHEET not in wb_tmpl.sheetnames:
    abort(
        f"The backtester template is missing the required sheet: '{TEMPLATE_SHEET}'\n"
        f"Please use the original TV Backtester template file."
    )

if 'Metadata' not in wb_tmpl.sheetnames:
    abort(
        f"The backtester template is missing the 'Metadata' sheet.\n"
        f"Please use the original TV Backtester template file."
    )

print("  ✅  Both files found and opened successfully.")

# ── STEP 2: FIND DATA SHEET AND HEADERS ──────────────────────────────────────
print("\n🔍  Scanning your data collection file...")

# Find the sheet — use first sheet that has data
ws_data = None
for sname in wb_data.sheetnames:
    ws = wb_data[sname]
    if ws.max_row > 1:
        ws_data = ws
        print(f"  ✅  Using sheet: '{sname}'")
        break

if ws_data is None:
    abort(
        f"Your data file '{DATA_FILE}' appears to be empty.\n"
        f"Please make sure it contains your trade combo data."
    )

# Find header row — scan first 10 rows for the one with most matches
best_header_row = None
best_match_count = 0
all_aliases_flat = [alias for aliases in COLUMN_ALIASES.values() for alias in aliases]

for r in range(1, min(11, ws_data.max_row + 1)):
    row_vals = [str(ws_data.cell(r, c).value or '').strip().lower() for c in range(1, ws_data.max_column + 1)]
    match_count = sum(1 for v in row_vals if v in all_aliases_flat)
    if match_count > best_match_count:
        best_match_count = match_count
        best_header_row = r

if best_header_row is None or best_match_count < 3:
    abort(
        f"Could not find a valid header row in your data file.\n"
        f"Please make sure your sheet has column headers and that they match the required names.\n"
        f"Required columns: SL %, TP %, Trades, Wins, Losses, BE, Avg Win $, Avg Loss $, Blown, Low Eq $, Max W Run, Max L Run"
    )

print(f"  ✅  Header row found at row {best_header_row} ({best_match_count} columns matched)")

# Map each required key to its column index
header_map = {}
raw_headers = {}
for c in range(1, ws_data.max_column + 1):
    val = str(ws_data.cell(best_header_row, c).value or '').strip().lower()
    raw_headers[c] = val

for key, aliases in COLUMN_ALIASES.items():
    for c, val in raw_headers.items():
        if val in aliases:
            header_map[key] = c
            break

# Check all 12 required columns found
missing = [FRIENDLY_NAMES[k] for k in COLUMN_ALIASES if k not in header_map]
if missing:
    abort(
        f"The following required columns could not be found in your data file:\n\n"
        + "\n".join(f"  • {m}" for m in missing) +
        f"\n\nPlease check your column headers match the required names.\n"
        f"Column names are case-insensitive — 'avg win $', 'Avg Win $', 'AVG WIN $' all work."
    )

print(f"  ✅  All 12 required columns found.")
for key, col in header_map.items():
    col_name = list(raw_headers.items())
    print(f"      {FRIENDLY_NAMES[key]:<40} → column {get_column_letter(col)}")

# ── STEP 3: READ ALL COMBOS ───────────────────────────────────────────────────
print(f"\n📥  Reading combo data...")

combos = []
skipped = 0
data_start_row = best_header_row + 1

for r in range(data_start_row, ws_data.max_row + 1):
    def get(key):
        return ws_data.cell(r, header_map[key]).value

    sl = get('sl')
    tp = get('tp')

    if sl is None and tp is None:
        continue  # empty row

    # Validate required numeric fields
    bad_fields = []
    for key in ['sl','tp','trades','wins','losses','avg_win','avg_loss']:
        v = get(key)
        if v is None or (isinstance(v, str) and v.strip() == ''):
            bad_fields.append(FRIENDLY_NAMES[key])

    if bad_fields:
        warn(f"Row {r} skipped — missing values in: {', '.join(bad_fields)}")
        skipped += 1
        continue

    try:
        combo = {
            'sl':       float(get('sl')),
            'tp':       float(get('tp')),
            'trades':   int(float(get('trades'))),
            'wins':     int(float(get('wins'))),
            'losses':   int(float(get('losses'))),
            'be':       int(float(get('be') or 0)),
            'avg_win':  float(get('avg_win')),
            'avg_loss': float(get('avg_loss')),
            'blown':    str(get('blown') or 'NO').strip().upper(),
            'low_eq':   float(get('low_eq') if get('low_eq') is not None else 0),
            'mcw':      int(float(get('mcw') or 0)),
            'mcl':      int(float(get('mcl') or 0)),
        }
        combos.append(combo)
    except (ValueError, TypeError) as e:
        warn(f"Row {r} skipped — could not read numeric values. Check the data in that row.")
        skipped += 1
        continue

if not combos:
    abort(
        f"No valid combos could be read from your data file.\n"
        f"Please check that your data rows contain numeric values in all required columns."
    )

print(f"  ✅  {len(combos)} combos loaded successfully.")
if skipped:
    warn(f"{skipped} rows were skipped due to missing or invalid values.")

# Derive strategy name from filename
strategy_name = os.path.splitext(os.path.basename(DATA_FILE))[0]
calc_sheet_name = (strategy_name[:24] + " Combos")[:31]  # Excel max 31 chars

# ── STEP 4: BUILD BACKTESTER CALCULATION SHEET ───────────────────────────────
print(f"\n🔨  Building backtester sheet with {len(combos)} combos...")

shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
wb_out = load_workbook(OUTPUT_FILE)

# Remove all original data sheets except Metadata
sheets_to_keep = {'Metadata'}
for sname in list(wb_out.sheetnames):
    if sname not in sheets_to_keep:
        del wb_out[sname]

ws_tmpl = wb_tmpl[TEMPLATE_SHEET]

# Create calc sheet
if calc_sheet_name in wb_out.sheetnames:
    del wb_out[calc_sheet_name]
ws_calc = wb_out.create_sheet(calc_sheet_name)

# Read template col B values and col A labels
tmpl_b, tmpl_a = {}, {}
for row in ws_tmpl.iter_rows(min_row=1, max_row=ws_tmpl.max_row):
    for cell in row:
        if cell.column == 2: tmpl_b[cell.row] = cell.value
        if cell.column == 1: tmpl_a[cell.row] = cell.value

# Write col A labels
for rn, val in tmpl_a.items():
    ws_calc.cell(row=rn, column=1, value=val)
ws_calc.column_dimensions['A'].width = 36

# Row heights
for rn, rd in ws_tmpl.row_dimensions.items():
    ws_calc.row_dimensions[rn].height = rd.height

INPUT_ROWS = {2, 3, 4, 5, 6, 7, 8, 9, 17, 18}
RISK_PCT = RISK_PER_TRADE / ACCOUNT_SIZE

def translate(formula, col_idx):
    if not formula or not isinstance(formula, str) or not formula.startswith('='):
        return formula
    cl = get_column_letter(col_idx)
    f = re.sub(r'(?<![A-Z])B(\d+)',  lambda m: f'{cl}{m.group(1)}', formula)
    f = re.sub(r'\$B\$(\d+)', lambda m: f'${cl}${m.group(1)}', f)
    f = re.sub(r'\$B(\d+)',   lambda m: f'${cl}{m.group(1)}', f)
    f = re.sub(r'B\$(\d+)',   lambda m: f'{cl}${m.group(1)}', f)
    return f

def get_input(r, combo):
    if r == 2:  return f"SL {combo['sl']}% TP {combo['tp']}%"
    if r == 3:  return TRADING_DAYS
    if r == 4:  return combo['trades']
    if r == 5:  return combo['wins']
    if r == 6:  return combo['losses']
    if r == 7:  return combo['be']
    if r == 8:  return combo['avg_win']
    if r == 9:  return combo['avg_loss']
    if r == 17: return ACCOUNT_SIZE
    if r == 18: return RISK_PCT

for i, combo in enumerate(combos):
    col_idx = i + 2
    for rn, val in tmpl_b.items():
        if rn in INPUT_ROWS:
            ws_calc.cell(row=rn, column=col_idx, value=get_input(rn, combo))
        elif val and isinstance(val, str) and val.startswith('='):
            ws_calc.cell(row=rn, column=col_idx, value=translate(val, col_idx))
        else:
            ws_calc.cell(row=rn, column=col_idx, value=val)
    ws_calc.column_dimensions[get_column_letter(col_idx)].width = 13
    if (i + 1) % 500 == 0:
        print(f"  {i+1}/{len(combos)} columns written...")

ws_calc.freeze_panes = 'B3'
print(f"  ✅  All {len(combos)} columns written.")

# ── STEP 5: ADD RAW DATA SHEET ────────────────────────────────────────────────
print(f"\n📋  Copying raw data sheet...")

ws_raw = wb_out.create_sheet('Raw Data')
for r in ws_data.iter_rows():
    for cell in r:
        ws_raw.cell(row=cell.row, column=cell.column, value=cell.value)

for col in ws_data.column_dimensions:
    ws_raw.column_dimensions[col].width = ws_data.column_dimensions[col].width

print("  ✅  Raw data copied.")

# Save before recalc
wb_out.save(OUTPUT_FILE)
print(f"\n💾  File saved: {OUTPUT_FILE}")

# ── STEP 6: RECALCULATE FORMULAS ─────────────────────────────────────────────
print(f"\n⚙️   Recalculating {len(combos)} combos × formulas (this may take a few minutes)...")

recalc_script = None
script_dir_recalc = os.path.dirname(os.path.abspath(__file__))
for candidate in [
    os.path.join(script_dir_recalc, 'recalc.py'),
    'recalc.py',
]:
    if os.path.exists(candidate):
        recalc_script = candidate
        break

if recalc_script is None:
    abort(
        f"Could not find 'recalc.py' in your script folder.\n"
        f"Please download recalc.py and place it in the same folder as master_backtester.py:\n"
        f"  {script_dir_recalc}"
    )

# Use sys.executable — same Python running this script — works on Windows, Mac, Linux
python_exe = sys.executable

try:
    result = subprocess.run(
        [python_exe, recalc_script, OUTPUT_FILE, '300'],
        capture_output=True, text=True, timeout=360,
        cwd=script_dir_recalc
    )

    if result.returncode != 0 and result.stderr:
        abort(
            f"Formula recalculation failed with the following error:\n\n"
            f"{result.stderr.strip()}\n\n"
            f"Most common cause: LibreOffice is not installed.\n"
            f"Download it free from: https://www.libreoffice.org/download/libreoffice-fresh/"
        )

    if not result.stdout.strip():
        abort(
            f"Formula recalculation produced no output.\n"
            f"LibreOffice is likely not installed or could not start.\n"
            f"Download it free from: https://www.libreoffice.org/download/libreoffice-fresh/\n"
            f"After installing, restart your terminal and run the script again."
        )

    recalc_result = json.loads(result.stdout)

except subprocess.TimeoutExpired:
    abort(
        f"Formula recalculation timed out after 6 minutes.\n"
        f"This can happen with very large datasets ({len(combos):,} combos).\n"
        f"Try splitting your data into smaller batches and running separately."
    )
except json.JSONDecodeError:
    abort(
        f"Formula recalculation returned unexpected output:\n\n"
        f"{result.stdout[:500]}\n\n"
        f"Please make sure LibreOffice is properly installed and try again."
    )
except FileNotFoundError:
    abort(
        f"Could not launch Python to run recalc.py.\n"
        f"Python executable: {python_exe}\n"
        f"Please make sure Python is correctly installed."
    )
except Exception as e:
    abort(
        f"Formula recalculation failed unexpectedly.\n"
        f"Error detail: {str(e)}\n\n"
        f"Most common cause: LibreOffice is not installed.\n"
        f"Download it free from: https://www.libreoffice.org/download/libreoffice-fresh/"
    )

if 'error' in recalc_result:
    abort(
        f"Formula recalculation failed:\n\n"
        f"{recalc_result['error']}\n\n"
        f"Most common cause: LibreOffice is not installed.\n"
        f"Download it free from: https://www.libreoffice.org/download/libreoffice-fresh/"
    )

if recalc_result.get('status') != 'success':
    errors = recalc_result.get('error_summary', {})
    if not errors:
        abort(
            f"Formula recalculation completed but reported errors with no detail.\n"
            f"Full recalc response: {recalc_result}\n\n"
            f"Please check the backtester template is the original version."
        )
    error_detail = '\n'.join([
        f"  • {etype}: {info['count']} cells — locations: {info['locations'][:5]}"
        for etype, info in errors.items()
    ])
    abort(
        f"Formula errors found after recalculation:\n\n{error_detail}\n\n"
        f"This usually means a formula references a cell or sheet that doesn't exist.\n"
        f"Please check the backtester template is the correct original version."
    )

print(f"  ✅  {recalc_result.get('total_formulas', 0):,} formulas calculated. Zero errors.")

# ── STEP 7: READ CALCULATED VALUES AND APPLY FILTERS ─────────────────────────
print(f"\n🔍  Reading grades and applying filters...")

wb_calc = load_workbook(OUTPUT_FILE, data_only=True)
ws_c    = wb_calc[calc_sheet_name]

ranked_combos = []
for i, combo in enumerate(combos):
    col = i + 2
    label = ws_c.cell(2, col).value

    c = {
        **combo,
        'label':    label,
        'trades_c': ws_c.cell(4,  col).value,
        'wr':       ws_c.cell(14, col).value,
        'ev_r':     ws_c.cell(30, col).value,
        'pf':       ws_c.cell(32, col).value,
        'ce':       ws_c.cell(33, col).value,
        'ror':      ws_c.cell(50, col).value,
        'mcl_calc': ws_c.cell(51, col).value,
        'total_pl': ws_c.cell(25, col).value,
        'gross_w':  ws_c.cell(26, col).value,
        'gross_l':  ws_c.cell(27, col).value,
        'g_ev':     ws_c.cell(57, col).value,
        'g_pf':     ws_c.cell(58, col).value,
        'g_ce':     ws_c.cell(59, col).value,
        'g_ror':    ws_c.cell(60, col).value,
        'g_str':    ws_c.cell(61, col).value,
    }
    ranked_combos.append(c)

total = len(ranked_combos)

# Apply 4 hard filters with counts
f1 = [c for c in ranked_combos if c['blown'] != 'YES']
f2 = [c for c in f1 if c['low_eq'] is not None and c['low_eq'] > 0]
f3 = [c for c in f2 if c['g_ror'] != 'F']
f4 = [c for c in f3 if c['g_str'] != 'F']
survivors = f4

print(f"  Filter 1 — Blown = YES:       removed {total - len(f1):>5}  →  {len(f1):>5} remaining")
print(f"  Filter 2 — Low Eq $ ≤ 0:      removed {len(f1)-len(f2):>5}  →  {len(f2):>5} remaining")
print(f"  Filter 3 — RoR Grade = F:      removed {len(f2)-len(f3):>5}  →  {len(f3):>5} remaining")
print(f"  Filter 4 — Streak Grade = F:   removed {len(f3)-len(f4):>5}  →  {len(f4):>5} remaining")
print(f"\n  ✅  {len(survivors)} survivors from {total} combos.")

if not survivors:
    warn(
        "No combos survived all 4 filters. This means every combo either "
        "blew the account, had an unacceptable risk of ruin, or had too long a losing streak.\n"
        "The BEST PROFILES sheet will be empty. Consider reviewing your strategy parameters."
    )

# Sort: CE grade → CE value → EV(R)
survivors.sort(key=lambda x: (
    gr(x['g_ce']),
    x['ce']   if x['ce']   else -999,
    x['ev_r'] if x['ev_r'] else -999,
), reverse=True)

# ── STEP 8: BUILD BEST PROFILES SHEET ────────────────────────────────────────
print(f"\n🏆  Building BEST PROFILES sheet...")

# Define border here so Step 8 is self-contained
s_bp = Side(style='thin', color='CCCCCC')
bdr  = Border(left=s_bp, right=s_bp, top=s_bp, bottom=s_bp)

wb_final  = load_workbook(OUTPUT_FILE)
wb_calc2  = load_workbook(OUTPUT_FILE, data_only=True)
ws_bt2    = wb_calc2[calc_sheet_name]

# Read Blown, Low Eq $, MCW from Raw Data using the SAME header_map we already built
# header_map keys: sl, tp, trades, wins, losses, be, avg_win, avg_loss, blown, low_eq, mcw, mcl
# These map to column indices in the ORIGINAL data file
# We re-read from ws_data (already open from Step 2)
ws_raw2 = wb_calc2['Raw Data']

# Find header row in Raw Data (copy of original sheet)
all_aliases_flat2 = [a for aliases in COLUMN_ALIASES.values() for a in aliases]
best_hrow2, best_hcount2 = 1, 0
for r in range(1, min(10, ws_raw2.max_row+1)):
    row_vals = [str(ws_raw2.cell(r,c).value or '').strip().lower() for c in range(1, ws_raw2.max_column+1)]
    cnt = sum(1 for v in row_vals if v in all_aliases_flat2)
    if cnt > best_hcount2:
        best_hcount2, best_hrow2 = cnt, r

# Build column map for Raw Data
raw_hmap2 = {}
for c in range(1, ws_raw2.max_column+1):
    val = str(ws_raw2.cell(best_hrow2, c).value or '').strip().lower()
    for key, aliases in COLUMN_ALIASES.items():
        if val in aliases and key not in raw_hmap2:
            raw_hmap2[key] = c

print(f"  Raw Data header row: {best_hrow2} | Columns found: {list(raw_hmap2.keys())}")

# Build leaderboard lookup
leaderboard2 = {}
sl_c  = raw_hmap2.get('sl')
tp_c  = raw_hmap2.get('tp')
bl_c  = raw_hmap2.get('blown')
le_c  = raw_hmap2.get('low_eq')
mcw_c = raw_hmap2.get('mcw')

if sl_c and tp_c:
    for r in range(best_hrow2+1, ws_raw2.max_row+1):
        sl = ws_raw2.cell(r, sl_c).value
        tp = ws_raw2.cell(r, tp_c).value
        if sl is None: continue
        label = f"SL {sl}% TP {tp}%"
        leaderboard2[label] = {
            'blown':  str(ws_raw2.cell(r, bl_c).value  if bl_c  else 'NO').strip().upper(),
            'low_eq': ws_raw2.cell(r, le_c).value      if le_c  else None,
            'mcw':    ws_raw2.cell(r, mcw_c).value     if mcw_c else None,
        }

print(f"  Leaderboard entries: {len(leaderboard2)}")

# Read all combos with computed grades
bp_combos = []
for col in range(2, ws_bt2.max_column+1):
    label = ws_bt2.cell(2, col).value
    if not label: continue
    lb = leaderboard2.get(label, {})
    bp_combos.append({
        'label':    label,
        'trades':   ws_bt2.cell(4,  col).value,
        'wr':       ws_bt2.cell(14, col).value,
        'ev_r':     ws_bt2.cell(30, col).value,
        'pf':       ws_bt2.cell(32, col).value,
        'ce':       ws_bt2.cell(33, col).value,
        'ror':      ws_bt2.cell(50, col).value,
        'mcl':      ws_bt2.cell(51, col).value,
        'total_pl': ws_bt2.cell(25, col).value,
        'gross_w':  ws_bt2.cell(26, col).value,
        'gross_l':  ws_bt2.cell(27, col).value,
        'g_ev':     ws_bt2.cell(57, col).value,
        'g_pf':     ws_bt2.cell(58, col).value,
        'g_ce':     ws_bt2.cell(59, col).value,
        'g_ror':    ws_bt2.cell(60, col).value,
        'g_str':    ws_bt2.cell(61, col).value,
        'blown':    lb.get('blown',  'NO'),
        'low_eq':   lb.get('low_eq', None),
        'mcw':      lb.get('mcw',    None),
    })

total_bp = len(bp_combos)

# 4 Hard filters
s = [c for c in bp_combos if c['blown'] != 'YES']
s = [c for c in s if c['low_eq'] is not None and c['low_eq'] > 0]
s = [c for c in s if c['g_ror'] != 'F']
s = [c for c in s if c['g_str'] != 'F']
survivors = s
eliminated_bp = total_bp - len(survivors)

print(f"  Filter 1 — Blown = YES:       removed {total_bp - len([c for c in bp_combos if c['blown'] != 'YES']):>5}")
print(f"  Filter 2 — Low Eq $ ≤ 0:      removed {len([c for c in bp_combos if c['blown'] != 'YES']) - len([c for c in bp_combos if c['blown'] != 'YES' and c['low_eq'] is not None and c['low_eq'] > 0]):>5}")
print(f"  Filter 3 — RoR Grade = F:      removed {len([c for c in bp_combos if c['blown'] != 'YES' and c['low_eq'] is not None and c['low_eq'] > 0]) - len([c for c in bp_combos if c['blown'] != 'YES' and c['low_eq'] is not None and c['low_eq'] > 0 and c['g_ror'] != 'F']):>5}")
print(f"  Filter 4 — Streak Grade = F:   removed {eliminated_bp - (total_bp - len([c for c in bp_combos if c['blown'] != 'YES' and c['low_eq'] is not None and c['low_eq'] > 0 and c['g_ror'] != 'F'])):>5}")
print(f"\n  ✅  {len(survivors)} survivors from {total_bp} combos.")

if not survivors:
    warn("No combos survived all 4 filters. BEST PROFILES sheet will be empty.")

# Sort: CE grade → CE value → EV(R)
survivors.sort(key=lambda x: (
    gr(x['g_ce']),
    x['ce']   if x['ce']   else -999,
    x['ev_r'] if x['ev_r'] else -999,
), reverse=True)

# Build sheet
BEST_SHEET = 'BEST PROFILES'
if BEST_SHEET in wb_final.sheetnames:
    del wb_final[BEST_SHEET]
ws_best = wb_final.create_sheet(BEST_SHEET, 0)

BP_HEADERS = [
    ('Rank',7),('Label',22),('Trades',8),('Win %',8),
    ('EV (R)',9),('PF',8),('CE',9),('RoR',9),('MCL',7),('MCW',7),
    ('Low Eq $',10),('EV\nGrade',8),('PF\nGrade',8),('CE\nGrade',8),
    ('RoR\nGrade',9),('Streak\nGrade',10),
    ('Gross\nProfit $',12),('Gross\nLoss $',12),('Total\nP&L $',12),
]
BP_NCOLS = len(BP_HEADERS)
BP_LC    = get_column_letter(BP_NCOLS)

ws_best.merge_cells(f'A1:{BP_LC}1')
tc = ws_best.cell(1, 1, f'BEST PROFILES — {strategy_name.upper()}  |  {len(survivors)} survivors from {total_bp}  |  4 Hard Filters Applied')
tc.font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
tc.fill = hfill(HDR_HEX); tc.alignment = Alignment(horizontal='center', vertical='center')
ws_best.row_dimensions[1].height = 22

ws_best.merge_cells(f'A2:{BP_LC}2')
info2 = ws_best.cell(2, 1,
    f'Filters: Blown + Low Eq≤$0 + RoR=F + Streak=F  |  '
    f'Rank: CE grade → CE value → EV(R)  |  '
    f'Account: ${ACCOUNT_SIZE:,}  Risk/trade: ${RISK_PER_TRADE:,}  |  '
    f'MCL=Max Consecutive Losses  MCW=Max Consecutive Wins'
)
info2.font = Font(name='Calibri', size=9, italic=True, color='555555')
info2.alignment = Alignment(horizontal='left', vertical='center')
ws_best.row_dimensions[2].height = 14

ws_best.row_dimensions[3].height = 30
for ci, (hdr, width) in enumerate(BP_HEADERS, 1):
    c = ws_best.cell(3, ci, hdr)
    c.font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
    c.fill = hfill(HDR_HEX)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = bdr
    ws_best.column_dimensions[get_column_letter(ci)].width = width

for rank, combo in enumerate(survivors, 1):
    r = rank + 3
    ws_best.row_dimensions[r].height = 16
    alt = hfill('F2F3F4') if rank % 2 == 0 else None

    def wc(col, val, fmt=None, bold=False, bg=None, fg='000000'):
        c = ws_best.cell(r, col, val)
        c.font = Font(name='Calibri', size=10, bold=bold, color=fg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr
        if fmt: c.number_format = fmt
        if bg:  c.fill = PatternFill('solid', fgColor=bg)
        elif alt: c.fill = PatternFill('solid', fgColor='F2F3F4')

    def gc(col, grade):
        bg, fg = GRADE_COLORS.get(str(grade).strip(), ('EEEEEE','000000')) if grade else ('EEEEEE','000000')
        wc(col, grade, bold=True, bg=bg, fg=fg)

    wc(1,  rank,             '0',         bold=True, bg=HDR_HEX, fg='FFFFFF')
    wc(2,  combo['label'])
    wc(3,  combo['trades'],  '0')
    wc(4,  combo['wr'],      '0.00%')
    wc(5,  combo['ev_r'],    '0.0000')
    wc(6,  combo['pf'],      '0.00')
    wc(7,  combo['ce'],      '0.0000')
    wc(8,  combo['ror'],     '0.00%')
    wc(9,  combo['mcl'],     '0')
    wc(10, combo['mcw'],     '0')
    wc(11, combo['low_eq'],  '$#,##0.00')
    gc(12, combo['g_ev'])
    gc(13, combo['g_pf'])
    gc(14, combo['g_ce'])
    gc(15, combo['g_ror'])
    gc(16, combo['g_str'])
    wc(17, combo['gross_w'], '$#,##0.00')
    wc(18, combo['gross_l'], '$#,##0.00')
    wc(19, combo['total_pl'],'$#,##0.00')

ws_best.freeze_panes = 'A4'
ws_best.auto_filter.ref = f'A3:{BP_LC}3'

# ── STEP 9: REORDER SHEETS ────────────────────────────────────────────────────
# Order: BEST PROFILES, Combos, Metadata, Raw Data
desired_order = [BEST_SHEET, calc_sheet_name, 'Metadata', 'Raw Data']
for sheet in desired_order:
    if sheet in wb_final.sheetnames:
        wb_final.move_sheet(sheet, offset=-len(wb_final.sheetnames))

wb_final.save(OUTPUT_FILE)

# ── DONE ──────────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print(f"✅  COMPLETE")
print(f"{'='*60}")
print(f"   Output file  : {OUTPUT_FILE}")
print(f"   Total combos : {total:,}")
print(f"   Survivors    : {len(survivors):,}")
print(f"   Eliminated   : {eliminated_bp:,}")
if survivors:
    top = survivors[0]
    print(f"\n   🏆  #1 Profile  : {top['label']}")
    print(f"       CE          : {safe(top['ce'])}")
    print(f"       EV(R)       : {safe(top['ev_r'])}")
    print(f"       PF          : {safe(top['pf'], 3)}")
    print(f"       MCL / MCW   : {top['mcl']} / {top['mcw']}")
    print(f"       Total P&L   : ${top['total_pl']:,.2f}" if top['total_pl'] else "")
print(f"{'='*60}\n")
