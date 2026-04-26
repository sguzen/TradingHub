"""
Build the forward-test tracker as a ready-to-use .xlsx file.

Run from this folder:
    python3 build_tracker.py

Output: forward_test_tracker.xlsx with two pre-formatted sheets (Trades + Summary)
and an embedded equity-curve chart. All formulas pre-populated for the first
500 trade rows; just import to Google Sheets or open in Excel and start logging.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

OUT = Path(__file__).parent / "forward_test_tracker.xlsx"

# How many empty rows to pre-populate with row formulas.
# 500 covers ~3.5 years of trading at the 1H_5M model's frequency.
N_ROWS = 500

# Column layout for the Trades sheet
TRADES_COLS = [
    ("trade_no",         12, "Sequence number (1, 2, 3 …)"),
    ("date",             12, "Trade date (YYYY-MM-DD)"),
    ("time_et",          10, "Entry time in ET (HH:MM)"),
    ("combo",            10, "1H/5M or 30M/3M"),
    ("direction",        10, "LONG or SHORT"),
    ("smt",              8,  "TRUE / FALSE — was SMT divergence true at entry"),
    ("planned_entry",    14, "Entry price from indicator alert"),
    ("planned_sl",       14, "Stop-loss price (sweep extreme)"),
    ("planned_tp",       14, "Take-profit price (1R from entry)"),
    ("planned_risk_pts", 16, "Auto: |planned_entry - planned_sl|"),
    ("actual_entry",     14, "Price you actually filled at"),
    ("actual_exit",      14, "Price the trade exited at (TP, SL, BE, or manual)"),
    ("outcome",          14, "WIN / LOSS / BE / SKIPPED / MANUAL_EXIT"),
    ("r_realized",       12, "Actual R captured (e.g. +1.0, -1.0, 0)"),
    ("mae_R",            10, "Deepest adverse excursion in R-units (eyeball OK)"),
    ("mfe_R",            10, "Furthest favorable excursion in R-units"),
    ("slippage_pts",     14, "actual_entry - planned_entry (signed)"),
    ("contracts",        10, "Number of MNQ contracts traded"),
    ("pnl_usd",          12, "Auto: r_realized × contracts × risk_pts × $2"),
    ("notes",            40, "Free text — context, news, fill quality, regime"),
]

# Color palette — muted, prints well, distinct enough for at-a-glance reading
HDR_FILL  = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HDR_FONT  = Font(color="F9FAFB", bold=True, size=11)
ZEBRA     = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
AUTO_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # amber for auto-computed cells
GOOD_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")  # green for healthy stats
WARN_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
LABEL_FONT = Font(bold=True, size=11, color="111827")
HEADLINE   = Font(bold=True, size=14, color="111827")
SUBHEAD    = Font(bold=True, size=11, color="374151")
SUBTLE     = Font(size=10, color="6B7280", italic=True)

THIN = Side(border_style="thin", color="D1D5DB")
BOX  = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def col_letter(idx_1_based: int) -> str:
    return get_column_letter(idx_1_based)


def build_trades_sheet(wb):
    ws = wb.active
    ws.title = "Trades"

    # ── Header row ──────────────────────────────────────────────────────────
    for i, (name, width, _) in enumerate(TRADES_COLS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[col_letter(i)].width = width

    # Tooltip / comment on each header cell so users know what goes where
    from openpyxl.comments import Comment
    for i, (name, _, doc) in enumerate(TRADES_COLS, start=1):
        ws.cell(row=1, column=i).comment = Comment(doc, "FractalSweep")

    ws.row_dimensions[1].height = 22

    # ── Pre-populated rows ──────────────────────────────────────────────────
    # Column indices (1-based) for formula references
    COL = {name: idx for idx, (name, _, _) in enumerate(TRADES_COLS, start=1)}

    for r in range(2, 2 + N_ROWS):
        # trade_no = sequential
        ws.cell(row=r, column=COL["trade_no"], value=r - 1)

        # planned_risk_pts: auto from planned_entry - planned_sl (absolute value)
        # Empty if either is blank → the IFERROR keeps the cell clean.
        e_ref  = f"{col_letter(COL['planned_entry'])}{r}"
        sl_ref = f"{col_letter(COL['planned_sl'])}{r}"
        ws.cell(
            row=r, column=COL["planned_risk_pts"],
            value=f'=IFERROR(IF(OR({e_ref}="",{sl_ref}=""),"",ABS({e_ref}-{sl_ref})),"")',
        ).fill = AUTO_FILL

        # pnl_usd: auto from r_realized * contracts * risk_pts * $2/pt (MNQ)
        r_ref = f"{col_letter(COL['r_realized'])}{r}"
        c_ref = f"{col_letter(COL['contracts'])}{r}"
        rp_ref = f"{col_letter(COL['planned_risk_pts'])}{r}"
        ws.cell(
            row=r, column=COL["pnl_usd"],
            value=f'=IFERROR(IF(OR({r_ref}="",{c_ref}="",{rp_ref}=""),"",{r_ref}*{c_ref}*{rp_ref}*2),"")',
        ).fill = AUTO_FILL

    # Freeze the header row + the first column
    ws.freeze_panes = "B2"

    # Number formats on the columns where it helps readability
    price_fmt = "#,##0.00"
    r_fmt     = "+0.00;-0.00;0.00"
    usd_fmt   = '"$"#,##0.00'
    pct_fmt   = "0.0%"

    for r in range(2, 2 + N_ROWS):
        for col_name in ("planned_entry", "planned_sl", "planned_tp",
                          "actual_entry", "actual_exit"):
            ws.cell(row=r, column=COL[col_name]).number_format = price_fmt
        ws.cell(row=r, column=COL["planned_risk_pts"]).number_format = price_fmt
        ws.cell(row=r, column=COL["r_realized"]).number_format       = r_fmt
        ws.cell(row=r, column=COL["mae_R"]).number_format            = r_fmt
        ws.cell(row=r, column=COL["mfe_R"]).number_format            = r_fmt
        ws.cell(row=r, column=COL["slippage_pts"]).number_format     = price_fmt
        ws.cell(row=r, column=COL["pnl_usd"]).number_format          = usd_fmt

    return COL


def build_equity_sheet(wb, trades_cols):
    """Helper sheet: running_R and drawdown columns, plus the equity chart.
    Hidden by default — Summary references it for the Max DD value and the
    chart lives anchored here, so the Summary sheet stays clean.
    """
    ws = wb.create_sheet("_Equity")

    R_COL_LETTER = col_letter(trades_cols["r_realized"])
    OUTCOME_COL  = col_letter(trades_cols["outcome"])

    # Headers
    ws["A1"] = "trade_no"
    ws["B1"] = "running_R"
    ws["C1"] = "drawdown_R"
    for c in ("A1", "B1", "C1"):
        ws[c].font = HDR_FONT
        ws[c].fill = HDR_FILL

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    # Row 2 = first trade
    ws["A2"] = 1
    ws["B2"] = (
        f'=IF(OR(Trades!{OUTCOME_COL}2="WIN",Trades!{OUTCOME_COL}2="LOSS"),'
        f'Trades!{R_COL_LETTER}2,0)'
    )
    ws["C2"] = "=MIN(0,B2-MAX($B$2:B2))"
    ws["B2"].number_format = "+0.00;-0.00;0.00"
    ws["C2"].number_format = "0.00"

    # Rows 3..N+1 (one row per trade)
    for offset in range(1, N_ROWS):
        excel_row  = 2 + offset
        trades_row = 2 + offset
        prev = f"B{excel_row - 1}"
        ws[f"A{excel_row}"] = offset + 1
        ws[f"B{excel_row}"] = (
            f'=IF(OR(Trades!{OUTCOME_COL}{trades_row}="WIN",'
            f'Trades!{OUTCOME_COL}{trades_row}="LOSS"),'
            f'{prev}+Trades!{R_COL_LETTER}{trades_row},'
            f'{prev})'
        )
        ws[f"C{excel_row}"] = f"=MIN(0,B{excel_row}-MAX($B$2:B{excel_row}))"
        ws[f"B{excel_row}"].number_format = "+0.00;-0.00;0.00"
        ws[f"C{excel_row}"].number_format = "0.00"

    # Embed equity chart on this helper sheet (visible if user unhides _Equity).
    chart = LineChart()
    chart.title = "Equity curve (cumulative R)"
    chart.style = 12
    chart.y_axis.title = "R"
    chart.x_axis.title = "Trade #"
    chart.height = 10
    chart.width = 18
    chart.legend = None
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=1 + N_ROWS)
    chart.add_data(data, titles_from_data=True)
    ws.add_chart(chart, "E2")

    ws.freeze_panes = "A2"
    # Note: we leave the sheet visible by default so Sheets users can see the
    # chart. To hide: ws.sheet_state = "hidden". Keeping visible for now —
    # users can hide it themselves if they want a tidier tab list.


def build_summary_sheet(wb, trades_cols):
    """Single clean dashboard. No overlapping data ranges, no hidden helpers
    inside this sheet — all chart-data lives on _Equity."""
    ws = wb.create_sheet("Summary")

    # Wider columns. Layout:
    #   A: stat label    | B: stat value    | C: spacer
    #   D: rolling label | E: rolling value | F: spacer
    #   G: status label  | H: status value
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 3
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 38

    R_COL_LETTER = col_letter(trades_cols["r_realized"])
    OUTCOME_COL  = col_letter(trades_cols["outcome"])
    PNL_COL      = col_letter(trades_cols["pnl_usd"])

    # ── Title block (rows 1-2) ─────────────────────────────────────────────
    ws["A1"] = "Fractal Sweep — Forward-Test Summary"
    ws["A1"].font = HEADLINE
    ws.merge_cells("A1:H1")

    ws["A2"] = ("Auto-updates from the Trades tab. Only WIN/LOSS rows count "
                "toward stats; SKIPPED and BE excluded from WR/EV.")
    ws["A2"].font = SUBTLE
    ws.merge_cells("A2:H2")

    # ── Headline stats (col A-B, rows 4-15) ────────────────────────────────
    ws["A4"] = "Headline stats"
    ws["A4"].font = SUBHEAD

    headline = [
        # (label, formula, number_format)
        ("Total trades logged",
         f'=COUNTA(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1})',
         "0"),
        ("Wins",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"WIN")',
         "0"),
        ("Losses",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"LOSS")',
         "0"),
        ("Breakevens",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"BE")',
         "0"),
        ("Skipped",
         f'=COUNTIF(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1},"SKIPPED")',
         "0"),
        ("Win rate (excl BE)",
         "=IFERROR(B6/(B6+B7),0)",
         "0.0%"),
        ("Avg R per trade (EV)",
         f'=IFERROR(SUM(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1})/(B6+B7+B8),0)',
         "+0.000;-0.000;0.000"),
        ("Profit factor",
         f'=IFERROR(SUMIF(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1},">0")/'
         f'ABS(SUMIF(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1},"<0")),0)',
         "0.00"),
        ("Total R",
         f'=SUM(Trades!{R_COL_LETTER}2:{R_COL_LETTER}{N_ROWS+1})',
         "+0.00;-0.00;0.00"),
        ("Total P&L ($)",
         f'=SUM(Trades!{PNL_COL}2:{PNL_COL}{N_ROWS+1})',
         '"$"#,##0.00'),
        ("Max drawdown (R)",
         f"=IFERROR(MIN(_Equity!C2:C{N_ROWS+1}),0)",
         "0.00"),
    ]
    for i, (label, formula, fmt) in enumerate(headline, start=5):
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = LABEL_FONT
        ws[f"B{i}"] = formula
        ws[f"B{i}"].number_format = fmt
        ws[f"B{i}"].alignment = Alignment(horizontal="right")

    # ── Rolling stats (col D-E, rows 4-10) ─────────────────────────────────
    ws["D4"] = "Rolling windows"
    ws["D4"].font = SUBHEAD

    # COUNTA on the outcome column gives the count of populated rows
    # (assumes outcomes are filled top-down without gaps).
    populated = f"COUNTA(Trades!{OUTCOME_COL}2:{OUTCOME_COL}{N_ROWS+1})"

    rolling = []
    for window in (30, 50, 100):
        # Range of the last N outcomes (capped at populated count)
        out_range = (
            f'OFFSET(Trades!{OUTCOME_COL}2,MAX(0,{populated}-{window}),0,'
            f'MIN({window},{populated}),1)'
        )
        r_range = (
            f'OFFSET(Trades!{R_COL_LETTER}2,MAX(0,{populated}-{window}),0,'
            f'MIN({window},{populated}),1)'
        )
        # WR = wins / (wins + losses), excludes BE/SKIPPED
        wr_formula = (
            f'=IFERROR(COUNTIF({out_range},"WIN")/'
            f'(COUNTIF({out_range},"WIN")+COUNTIF({out_range},"LOSS")),0)'
        )
        # EV = sum(R) / (wins+losses+BE). SKIPPED have empty R so they
        # don't contribute to numerator; we explicitly exclude them from
        # the denominator too so the average matches the headline EV.
        denom = (
            f'(COUNTIF({out_range},"WIN")+'
            f'COUNTIF({out_range},"LOSS")+'
            f'COUNTIF({out_range},"BE"))'
        )
        ev_formula = (
            f'=IFERROR(SUM({r_range})/{denom},0)'
        )
        rolling.append((f"Last {window} — WR",  wr_formula, "0.0%"))
        rolling.append((f"Last {window} — EV",  ev_formula, "+0.000;-0.000;0.000"))

    for i, (label, formula, fmt) in enumerate(rolling, start=5):
        ws[f"D{i}"] = label
        ws[f"D{i}"].font = LABEL_FONT
        ws[f"E{i}"] = formula
        ws[f"E{i}"].number_format = fmt
        ws[f"E{i}"].alignment = Alignment(horizontal="right")

    # ── Status indicator (col G-H, rows 4-5) ───────────────────────────────
    ws["G4"] = "Forward-test status"
    ws["G4"].font = SUBHEAD

    # Status formula references B5 (total trades), B10 (WR), B11 (EV)
    ws["G5"] = "Status"
    ws["G5"].font = LABEL_FONT
    ws["H5"] = (
        '=IF(B5<30,"⏳ Noise — keep tracking",'
        'IF(B5<100,"📊 Trend — too early to call",'
        'IF(B10<0.45,"🛑 STOP — WR below 45%",'
        'IF(B11<0,"🛑 STOP — EV negative",'
        '"✓ On track"))))'
    )
    ws["H5"].font = Font(bold=True, size=12)
    ws["H5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H5"].fill = GOOD_FILL
    ws.row_dimensions[5].height = 28

    ws["G6"] = "Trades to next phase"
    ws["G6"].font = LABEL_FONT
    ws["H6"] = (
        '=IF(B5<30,"~"&(30-B5)&" more for trend signal",'
        'IF(B5<100,"~"&(100-B5)&" more for early call",'
        'IF(B5<200,"~"&(200-B5)&" more for robust validation",'
        '"Validated")))'
    )
    ws["H6"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Decision rules (rows 17-23) ────────────────────────────────────────
    ws["A17"] = "Decision rules"
    ws["A17"].font = SUBHEAD
    rules = [
        "🛑 Stop trading if cumulative drawdown > 15R = $3,375",
        "🛑 Stop if WR < 45% on Last-50 window",
        "🛑 Stop if EV < 0R on Last-100 window",
        "✓  Continue if WR 50-65% on rolling windows (within expected variance)",
        "✓  Continue through 5-trade losing streaks (math says they're common at 59% WR)",
    ]
    for i, rule in enumerate(rules, start=18):
        ws[f"A{i}"] = rule
        ws[f"A{i}"].font = Font(size=10, color="374151")
        ws.merge_cells(f"A{i}:H{i}")

    # ── Validation criteria (rows 24-29) ───────────────────────────────────
    ws["A24"] = "Validation criteria — what 'edge confirmed' looks like after ~200 trades:"
    ws["A24"].font = SUBHEAD
    ws.merge_cells("A24:H24")
    criteria = [
        "WR consistently 55–62% on rolling windows",
        "Avg R per trade between +0.10 and +0.18",
        "Max drawdown stayed within 15R",
        "Profit factor 1.4 or higher",
    ]
    for i, c in enumerate(criteria, start=25):
        ws[f"A{i}"] = "• " + c
        ws[f"A{i}"].font = Font(size=10, color="374151")
        ws.merge_cells(f"A{i}:H{i}")

    # ── Note pointing to equity chart (row 31) ─────────────────────────────
    ws["A31"] = "Equity curve chart: see _Equity tab."
    ws["A31"].font = SUBTLE
    ws.merge_cells("A31:H31")

    # No frozen pane — whole summary fits in one screen, no scroll needed.


def main():
    wb = Workbook()
    trades_cols = build_trades_sheet(wb)
    build_equity_sheet(wb, trades_cols)
    build_summary_sheet(wb, trades_cols)
    # Tab order: Trades → Summary → _Equity. Move Summary to second.
    wb.move_sheet("Summary", offset=-1)
    wb.save(OUT)
    print(f"✓ Wrote {OUT}")
    print(f"  {N_ROWS} trade rows pre-populated")
    print(f"  Open in Excel or upload to Google Sheets (File → Import) to start")


if __name__ == "__main__":
    main()
