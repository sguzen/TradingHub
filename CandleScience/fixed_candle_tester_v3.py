#!/usr/bin/env python3
"""
MNQ Fixed Candle Strategy Tester v1
=====================================
Four modes:
  BREAKOUT   — GREEN candle → LONG,  RED candle → SHORT  (entry = next candle open)
  REVERSAL   — GREEN candle → SHORT, RED candle → LONG   (entry = next candle open)
  LONG ONLY  — always LONG,  entry = fixed candle's own open
  SHORT ONLY — always SHORT, entry = fixed candle's own open

One fixed candle time per run (e.g. 10:06).
If fixed candle missing on a day → skip.
TIME exit: X minutes from signal candle open, or EOD.
Grid sweeps 10,000 SL/TP combos (0.01%–1.00% × 0.01%–1.00%).
Per-combo SL/TP distribution stats: MAE mode/median (wins), MFE mode/median (losses+TIME).
"""

import os, sys, glob, math, logging
import numpy as np
import pandas as pd
from itertools import product
from datetime import datetime, timedelta, time as dtime
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format='%(message)s')
log = logging.getLogger(__name__)

MNQ_MULTIPLIER = 2.0

# ── Color palette ──────────────────────────────────────────────────────────────
DARK_BG    = '0D0F14'
GOLD       = 'F5C842'
TEAL       = '3DD9B3'
RED_CLR    = 'F5504A'
BLUE_CLR   = '5B9CF6'
GREEN_CLR  = '2ECC71'
WHITE      = 'E8EAF0'
MUTED      = '7A82A0'
CARD_BG    = '13161E'
RAISED_BG  = '1A1E28'
BORDER_CLR = '252A38'

def fill(c):   return PatternFill('solid', start_color=c, fgColor=c)
def center():  return Alignment(horizontal='center', vertical='center')
def left():    return Alignment(horizontal='left',   vertical='center')
def tborder(c=BORDER_CLR):
    s = Side(style='thin', color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def wcol(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def wc(ws, r, c, v, font=None, fill_=None, align=None, border=None, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    if font:   cell.font      = font
    if fill_:  cell.fill      = fill_
    if align:  cell.alignment = align
    if border: cell.border    = border
    if fmt:    cell.number_format = fmt
    return cell
def hf(sz=11, bold=True,  color=WHITE): return Font(name='Arial', size=sz, bold=bold,  color=color)
def cf(sz=10, bold=False, color=WHITE): return Font(name='Arial', size=sz, bold=bold, color=color)


# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
def find_csv_files():
    patterns = ['*.csv', '*.txt', '*.tsv']
    found = []
    for p in patterns:
        found.extend(glob.glob(p))
        found.extend(glob.glob(f'/mnt/user-data/uploads/{p}'))
    return list(dict.fromkeys(found))

def load_ohlc(filepath):
    with open(filepath, 'r') as f: sample = f.read(2048)
    sep = '\t' if sample.count('\t') > sample.count(',') else ','
    df  = pd.read_csv(filepath, sep=sep, header=0, low_memory=False)
    df.columns = [c.strip().lower() for c in df.columns]

    col_map = {}
    for col in df.columns:
        c = col.strip().lower()
        if c in ('o', 'open'):               col_map[col] = 'open'
        elif c in ('h', 'high'):             col_map[col] = 'high'
        elif c in ('l', 'low'):              col_map[col] = 'low'
        elif c in ('c', 'close', 'last'):    col_map[col] = 'close'
    df = df.rename(columns=col_map)

    raw      = [c for c in df.columns if c not in ('open','high','low','close')]
    dt_col   = next((c for c in raw if c in ('datetime','timestamp','dt','ts_event')), None)
    date_col = next((c for c in raw if c == 'date' or ('date' in c and 'time' not in c)), None)
    time_col = next((c for c in raw if c == 'time' or ('time' in c and 'date' not in c)), None)

    if dt_col is not None:
        df['datetime'] = pd.to_datetime(df[dt_col].astype(str))
    elif date_col is not None and time_col is not None:
        df['datetime'] = pd.to_datetime(
            df[date_col].astype(str).str.strip() + ' ' +
            df[time_col].astype(str).str.strip())
    elif date_col is not None:
        df['datetime'] = pd.to_datetime(df[date_col].astype(str))
    else:
        raise ValueError(f"Cannot find datetime column. Columns: {list(df.columns)}")

    for r in ['open','high','low','close']:
        if r not in df.columns:
            raise ValueError(f"Missing column '{r}'. Columns: {list(df.columns)}")

    df = df[['datetime','open','high','low','close']].copy()
    df[['open','high','low','close']] = df[['open','high','low','close']].apply(pd.to_numeric, errors='coerce')
    df.dropna(inplace=True)
    df.sort_values('datetime', inplace=True)
    df.reset_index(drop=True, inplace=True)
    log.info(f"Loaded {len(df):,} bars | {df['datetime'].min().date()} → {df['datetime'].max().date()}")
    return df

def resample_candles(df, minutes):
    r = df.set_index('datetime').resample(f'{minutes}min').agg(
        {'open':'first','high':'max','low':'min','close':'last'}).dropna()
    r.reset_index(inplace=True)
    return r


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1: EXTRACT TRADE SKELETONS
# ══════════════════════════════════════════════════════════════════════════════
def extract_trade_skeletons(df_candles, fixed_time, mode, time_exit_type,
                             time_exit_minutes, eod_time, df_m1=None):
    """
    fixed_time         : datetime.time — the candle to anchor on (e.g. time(10,6))
    mode               : 'breakout' | 'reversal' | 'long' | 'short'
    time_exit_type     : 'minutes' | 'eod'
    time_exit_minutes  : int — minutes from signal candle open (used if type='minutes')
    eod_time           : datetime.time — end of day time (used if type='eod')
    """
    df_candles = df_candles.copy()
    df_candles['date'] = df_candles['datetime'].dt.date

    # Build M1 lookup by date
    m1_by_day = {}
    if df_m1 is not None:
        df_m1 = df_m1.copy()
        df_m1['date'] = df_m1['datetime'].dt.date
        for day, grp in df_m1.groupby('date'):
            g = grp.reset_index(drop=True)
            m1_by_day[day] = {
                'highs': g['high'].values,
                'lows':  g['low'].values,
                'dts':   g['datetime'].values,
            }

    day_skeletons = {}

    for day, grp in df_candles.groupby('date'):
        grp = grp.reset_index(drop=True)

        # Find the fixed candle
        fixed_mask = grp['datetime'].dt.time == fixed_time
        if not fixed_mask.any():
            continue
        fixed_idx = fixed_mask.idxmax()
        fixed_bar = grp.iloc[fixed_idx]
        fixed_dt  = fixed_bar['datetime']

        # Determine direction and entry
        is_green = fixed_bar['close'] >= fixed_bar['open']

        if mode == 'breakout':
            direction = 'LONG' if is_green else 'SHORT'
            signal_color = 'GREEN' if is_green else 'RED'
            # Entry = next candle open
            if fixed_idx + 1 >= len(grp):
                continue
            next_bar    = grp.iloc[fixed_idx + 1]
            entry_price = float(next_bar['open'])
            entry_dt    = next_bar['datetime']

        elif mode == 'reversal':
            direction = 'SHORT' if is_green else 'LONG'
            signal_color = 'GREEN' if is_green else 'RED'
            if fixed_idx + 1 >= len(grp):
                continue
            next_bar    = grp.iloc[fixed_idx + 1]
            entry_price = float(next_bar['open'])
            entry_dt    = next_bar['datetime']

        elif mode == 'long':
            direction    = 'LONG'
            signal_color = 'GREEN' if is_green else 'RED'
            entry_price  = float(fixed_bar['open'])
            entry_dt     = fixed_dt

        elif mode == 'short':
            direction    = 'SHORT'
            signal_color = 'GREEN' if is_green else 'RED'
            entry_price  = float(fixed_bar['open'])
            entry_dt     = fixed_dt

        if entry_price <= 0 or np.isnan(entry_price):
            continue

        # Compute TIME exit datetime
        if time_exit_type == 'minutes':
            time_exit_dt = fixed_dt + timedelta(minutes=time_exit_minutes)
        else:
            # EOD — same day at eod_time
            time_exit_dt = datetime.combine(day, eod_time)

        # Get M1 window: from entry_dt to time_exit_dt
        m1_day = m1_by_day.get(day)
        if m1_day is not None:
            m1_dts   = m1_day['dts']
            m1_highs = m1_day['highs']
            m1_lows  = m1_day['lows']
            entry_np    = np.datetime64(entry_dt)
            exit_np     = np.datetime64(time_exit_dt)
            mask = (m1_dts >= entry_np) & (m1_dts <= exit_np)
            m1_scan_highs = m1_highs[mask]
            m1_scan_lows  = m1_lows[mask]
            m1_scan_dts   = m1_dts[mask]
        else:
            # Fallback: use candle TF bars
            fallback = grp[grp['datetime'] >= entry_dt].copy()
            fallback = fallback[fallback['datetime'] <= time_exit_dt]
            m1_scan_highs = fallback['high'].values
            m1_scan_lows  = fallback['low'].values
            m1_scan_dts   = fallback['datetime'].values

        # TIME exit price = last available M1 close before or at time_exit_dt
        # Use last bar's close in fallback, or look up M1
        if m1_day is not None:
            te_mask = m1_day['dts'] <= np.datetime64(time_exit_dt)
            if te_mask.any():
                time_exit_price = float(df_m1[
                    (df_m1['date'] == day) &
                    (df_m1['datetime'] <= time_exit_dt)
                ]['close'].iloc[-1])
            else:
                time_exit_price = entry_price
        else:
            candidates = grp[grp['datetime'] <= time_exit_dt]
            time_exit_price = float(candidates['close'].iloc[-1]) if not candidates.empty else entry_price

        day_skeletons[day] = {
            'fixed_dt':        fixed_dt,
            'entry_dt':        entry_dt,
            'entry_price':     entry_price,
            'direction':       direction,
            'signal_color':    signal_color,
            'mode':            mode,
            'm1_scan_highs':   m1_scan_highs,
            'm1_scan_lows':    m1_scan_lows,
            'm1_scan_dts':     m1_scan_dts,
            'time_exit_price': time_exit_price,
            'time_exit_dt':    time_exit_dt,
        }

    log.info(f"  Fixed candle found on {len(day_skeletons)} trading days")
    return day_skeletons


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2: APPLY ONE SL/TP COMBO
# ══════════════════════════════════════════════════════════════════════════════
def apply_sltp_to_day(day, sk, sl_pct, tp_pct, risk_per_trade, budget_mode='flag'):
    entry_price      = sk['entry_price']
    direction        = sk['direction']
    m1_highs         = sk['m1_scan_highs']
    m1_lows          = sk['m1_scan_lows']
    m1_dts           = sk['m1_scan_dts']
    time_exit_price  = sk['time_exit_price']
    time_exit_dt     = sk['time_exit_dt']

    sl_dist  = entry_price * (sl_pct / 100.0)
    tp_dist  = entry_price * (tp_pct / 100.0)
    sl_level = (entry_price - sl_dist) if direction == 'LONG' else (entry_price + sl_dist)
    tp_level = (entry_price + tp_dist) if direction == 'LONG' else (entry_price - tp_dist)

    dollar_risk   = sl_dist * MNQ_MULTIPLIER
    contracts     = max(1, int(risk_per_trade / dollar_risk)) if dollar_risk > 0 else 1
    actual_risk   = sl_dist * contracts * MNQ_MULTIPLIER
    budget_breach = actual_risk > risk_per_trade

    if budget_breach and budget_mode == 'hard_skip':
        return None

    n_scan = len(m1_highs)

    if n_scan == 0:
        exit_price  = time_exit_price
        exit_reason = 'TIME'
        exit_dt     = time_exit_dt
        exit_bar    = 0
    else:
        if direction == 'LONG':
            sl_hits = np.where(m1_lows  <= sl_level)[0]
            tp_hits = np.where(m1_highs >= tp_level)[0]
        else:
            sl_hits = np.where(m1_highs >= sl_level)[0]
            tp_hits = np.where(m1_lows  <= tp_level)[0]

        sl_bar = int(sl_hits[0]) if len(sl_hits) else n_scan
        tp_bar = int(tp_hits[0]) if len(tp_hits) else n_scan

        if sl_bar <= tp_bar and sl_bar < n_scan:
            exit_bar    = sl_bar
            exit_reason = 'SL'
            exit_price  = float(sl_level)
            exit_dt     = m1_dts[sl_bar]
        elif tp_bar < sl_bar and tp_bar < n_scan:
            exit_bar    = tp_bar
            exit_reason = 'TP'
            exit_price  = float(tp_level)
            exit_dt     = m1_dts[tp_bar]
        else:
            exit_bar    = n_scan - 1
            exit_reason = 'TIME'
            exit_price  = float(time_exit_price)
            exit_dt     = time_exit_dt

    # MAE / MFE
    seg_h = m1_highs[:exit_bar + 1] if n_scan > 0 else np.array([entry_price])
    seg_l = m1_lows[:exit_bar + 1]  if n_scan > 0 else np.array([entry_price])
    if direction == 'LONG':
        mae = max(0.0, float(entry_price - seg_l.min()))
        mfe = max(0.0, float(seg_h.max() - entry_price))
    else:
        mae = max(0.0, float(seg_h.max() - entry_price))
        mfe = max(0.0, float(entry_price - seg_l.min()))

    pnl_pts = (exit_price - entry_price) if direction == 'LONG' else (entry_price - exit_price)
    pnl_usd = round(pnl_pts * contracts * MNQ_MULTIPLIER, 2)

    return {
        'date':          day,
        'fixed_dt':      sk['fixed_dt'],
        'entry_time':    sk['entry_dt'],
        'direction':     direction,
        'signal_color':  sk['signal_color'],
        'mode':          sk['mode'],
        'entry_price':   round(entry_price, 4),
        'exit_price':    round(exit_price,  4),
        'sl_level':      round(sl_level,    4),
        'tp_level':      round(tp_level,    4),
        'exit_reason':   exit_reason,
        'contracts':     contracts,
        'actual_risk':   round(actual_risk, 2),
        'budget_breach': budget_breach,
        'pnl_points':    round(pnl_pts, 4),
        'pnl_dollars':   pnl_usd,
        'mae_pct':       round((mae / entry_price) * 100, 4) if entry_price > 0 else 0.0,
        'mfe_pct':       round((mfe / entry_price) * 100, 4) if entry_price > 0 else 0.0,
        'bars_in_trade': exit_bar + 1,
    }

def run_backtest_fast(day_skeletons, sl_pct, tp_pct, risk_per_trade, budget_mode='flag'):
    trades = []
    for day, sk in day_skeletons.items():
        t = apply_sltp_to_day(day, sk, sl_pct, tp_pct, risk_per_trade, budget_mode)
        if t is not None:
            trades.append(t)
    return pd.DataFrame(trades) if trades else pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# ORIGINAL ENGINE — pure Python reference for integrity validation
# ══════════════════════════════════════════════════════════════════════════════
def run_backtest_original(day_skeletons, sl_pct, tp_pct, risk_per_trade, budget_mode='flag'):
    trades = []
    for day, sk in day_skeletons.items():
        entry_price     = sk['entry_price']
        direction       = sk['direction']
        m1_highs        = sk['m1_scan_highs']
        m1_lows         = sk['m1_scan_lows']
        m1_dts          = sk['m1_scan_dts']
        time_exit_price = sk['time_exit_price']
        time_exit_dt    = sk['time_exit_dt']

        sl_dist  = entry_price * (sl_pct / 100.0)
        tp_dist  = entry_price * (tp_pct / 100.0)
        sl_level = (entry_price - sl_dist) if direction == 'LONG' else (entry_price + sl_dist)
        tp_level = (entry_price + tp_dist) if direction == 'LONG' else (entry_price - tp_dist)

        dollar_risk   = sl_dist * MNQ_MULTIPLIER
        contracts     = max(1, int(risk_per_trade / dollar_risk)) if dollar_risk > 0 else 1
        actual_risk   = sl_dist * contracts * MNQ_MULTIPLIER
        budget_breach = actual_risk > risk_per_trade

        if budget_breach and budget_mode == 'hard_skip':
            continue

        exit_price  = None
        exit_reason = None
        exit_dt_val = None
        mae = 0.0; mfe = 0.0; bars = 0

        for i in range(len(m1_highs)):
            bars += 1
            h = m1_highs[i]; l = m1_lows[i]
            if direction == 'LONG':
                mae = max(mae, entry_price - l)
                mfe = max(mfe, h - entry_price)
                if l <= sl_level: exit_price=sl_level; exit_reason='SL'; exit_dt_val=m1_dts[i]; break
                if h >= tp_level: exit_price=tp_level; exit_reason='TP'; exit_dt_val=m1_dts[i]; break
            else:
                mae = max(mae, h - entry_price)
                mfe = max(mfe, entry_price - l)
                if h >= sl_level: exit_price=sl_level; exit_reason='SL'; exit_dt_val=m1_dts[i]; break
                if l <= tp_level: exit_price=tp_level; exit_reason='TP'; exit_dt_val=m1_dts[i]; break

        if exit_price is None:
            exit_price  = time_exit_price
            exit_reason = 'TIME'
            exit_dt_val = time_exit_dt

        mae = max(0.0, mae); mfe = max(0.0, mfe)
        pnl_pts = (exit_price - entry_price) if direction=='LONG' else (entry_price - exit_price)
        pnl_usd = round(pnl_pts * contracts * MNQ_MULTIPLIER, 2)

        trades.append({
            'date':          day,
            'fixed_dt':      sk['fixed_dt'],
            'entry_time':    sk['entry_dt'],
            'direction':     direction,
            'signal_color':  sk['signal_color'],
            'mode':          sk['mode'],
            'entry_price':   round(entry_price, 4),
            'exit_price':    round(exit_price,  4),
            'sl_level':      round(sl_level,    4),
            'tp_level':      round(tp_level,    4),
            'exit_reason':   exit_reason,
            'contracts':     contracts,
            'actual_risk':   round(actual_risk, 2),
            'budget_breach': budget_breach,
            'pnl_points':    round(pnl_pts, 4),
            'pnl_dollars':   pnl_usd,
            'mae_pct':       round((mae / entry_price) * 100, 4) if entry_price > 0 else 0.0,
            'mfe_pct':       round((mfe / entry_price) * 100, 4) if entry_price > 0 else 0.0,
            'bars_in_trade': bars,
        })
    return pd.DataFrame(trades) if trades else pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# SL/TP DISTRIBUTION STATS
# ══════════════════════════════════════════════════════════════════════════════
def _mode_val(series):
    """Most common value in a series (rounded to 4dp). Returns None if empty."""
    if series is None or len(series) == 0:
        return None
    rounded = [round(float(v), 4) for v in series]
    c = Counter(rounded)
    return c.most_common(1)[0][0]

def calc_distribution_stats(trades_df):
    """
    MAE stats  → winning trades only  (how much room did winners need?)
    MFE stats  → losing + TIME trades (how far did losers go your way before reversing?)
    Returns dict of mode/median for SL (MAE) and TP (MFE).
    """
    if trades_df is None or trades_df.empty:
        return {
            'mae_median': None, 'mae_mode': None,
            'mfe_median': None, 'mfe_mode': None,
            'mae_mode_freq': None, 'mfe_mode_freq': None,
        }

    wins     = trades_df[trades_df['pnl_dollars'] > 0]
    losers   = trades_df[trades_df['pnl_dollars'] <= 0]

    mae_vals = wins['mae_pct'].values   if not wins.empty   else np.array([])
    mfe_vals = losers['mfe_pct'].values if not losers.empty else np.array([])

    mae_median = round(float(np.median(mae_vals)), 4) if len(mae_vals) > 0 else None
    mae_mode   = _mode_val(mae_vals)
    mfe_median = round(float(np.median(mfe_vals)), 4) if len(mfe_vals) > 0 else None
    mfe_mode   = _mode_val(mfe_vals)

    # Mode frequency as % of winning trades
    mae_mode_count = sum(1 for v in mae_vals if round(float(v),4) == mae_mode) if mae_mode is not None else 0
    mae_mode_freq  = round(mae_mode_count / len(mae_vals) * 100, 1) if len(mae_vals) > 0 else None

    mfe_mode_count = sum(1 for v in mfe_vals if round(float(v),4) == mfe_mode) if mfe_mode is not None else 0
    mfe_mode_freq  = round(mfe_mode_count / len(mfe_vals) * 100, 1) if len(mfe_vals) > 0 else None

    return {
        'mae_median':   mae_median,
        'mae_mode':     mae_mode,
        'mae_mode_freq': mae_mode_freq,
        'mfe_median':   mfe_median,
        'mfe_mode':     mfe_mode,
        'mfe_mode_freq': mfe_mode_freq,
    }


# ══════════════════════════════════════════════════════════════════════════════
# METRICS
# ══════════════════════════════════════════════════════════════════════════════
def _calc_max_consec_losing_days(trades_df):
    if trades_df is None or trades_df.empty or 'date' not in trades_df.columns:
        return 0
    daily_pnl = trades_df.groupby('date')['pnl_dollars'].sum()
    streak = 0; best = 0
    for v in daily_pnl:
        if v < 0: streak += 1; best = max(best, streak)
        else:     streak = 0
    return best

def calc_metrics(trades_df, account_size, risk_per_trade):
    if trades_df is None or trades_df.empty: return None
    wins   = trades_df[trades_df['pnl_dollars'] > 0]
    losses = trades_df[trades_df['pnl_dollars'] < 0]
    n      = len(trades_df)
    if n == 0: return None

    n_wins = len(wins); n_loss = len(losses)
    win_pct  = n_wins / n
    loss_pct = n_loss / n
    avg_win  = float(wins['pnl_dollars'].mean())        if n_wins > 0 else 0.0
    avg_loss = float(abs(losses['pnl_dollars'].mean())) if n_loss > 0 else 0.0

    gross_profit = float(wins['pnl_dollars'].sum())        if n_wins > 0 else 0.0
    gross_loss   = float(abs(losses['pnl_dollars'].sum())) if n_loss > 0 else 0.0

    pf_num = win_pct * avg_win
    pf_den = loss_pct * avg_loss
    pf = pf_num / pf_den if pf_den > 0 else (999.0 if pf_num > 0 else 0.0)

    ev            = pf_num - pf_den
    ev_r          = ev / avg_loss if avg_loss > 0 else 0.0
    combined_edge = ev_r * pf

    max_streak = math.ceil(math.log(n) / math.log(1.0 / loss_pct)) if (n > 1 and 0 < loss_pct < 1) else n_loss
    risk_pct   = risk_per_trade / account_size if account_size > 0 else 0.05
    dd_pct     = max_streak * risk_pct * 100

    bankroll_n = math.floor(1 / risk_pct) if risk_pct > 0 else 20
    if combined_edge <= 0:
        ror = 100.0
    else:
        ror = min(1.0, ((1 - combined_edge) / (1 + combined_edge)) ** bankroll_n) * 100

    r_multiples = trades_df['pnl_dollars'] / risk_per_trade if risk_per_trade > 0 else trades_df['pnl_dollars']
    avg_r = float(r_multiples.mean())
    std_r = float(r_multiples.std(ddof=1)) if n > 1 else 0.0
    sqn   = round((avg_r / std_r) * math.sqrt(n), 4) if std_r > 0 else 0.0

    cumul          = trades_df['pnl_dollars'].cumsum()
    equity_curve   = account_size + cumul
    cumul_anch     = pd.concat([pd.Series([0.0]), cumul]).reset_index(drop=True)
    max_dd_usd     = float((cumul_anch.cummax() - cumul_anch).max())
    total_pnl      = float(trades_df['pnl_dollars'].sum())
    largest_win    = float(wins['pnl_dollars'].max())   if n_wins > 0 else 0.0
    largest_loss   = float(losses['pnl_dollars'].min()) if n_loss > 0 else 0.0
    lowest_equity  = round(min(float(account_size), float(equity_curve.min())), 2)
    account_blown  = bool(lowest_equity < 0)

    _cw = 0; _cl = 0; mcw = 0; mcl = 0
    for _p in trades_df['pnl_dollars']:
        if   _p > 0: _cw += 1; _cl = 0; mcw = max(mcw, _cw)
        elif _p < 0: _cl += 1; _cw = 0; mcl = max(mcl, _cl)
        else:        _cw  = 0; _cl = 0

    n_breaches = int(trades_df['budget_breach'].sum()) if 'budget_breach' in trades_df.columns else 0
    breach_pct = round(n_breaches / n * 100, 2)

    dist = calc_distribution_stats(trades_df)

    return {
        'total_trades':     n,
        'n_wins':           n_wins,
        'n_losses':         n_loss,
        'n_breakeven':      n - n_wins - n_loss,
        'win_pct':          round(win_pct  * 100, 2),
        'loss_pct':         round(loss_pct * 100, 2),
        'avg_win_usd':      round(avg_win,  2),
        'avg_loss_usd':     round(avg_loss, 2),
        'ratio_win_loss':   round(avg_win / avg_loss, 3) if avg_loss > 0 else 0.0,
        'gross_profit':     round(gross_profit, 2),
        'gross_loss':       round(gross_loss,   2),
        'total_pnl':        round(total_pnl,    2),
        'largest_win':      round(largest_win,  2),
        'largest_loss':     round(largest_loss, 2),
        'ev':               round(ev, 2),
        'profit_factor':    round(pf, 3),
        'combined_edge':    round(combined_edge, 4),
        'sqn':              sqn,
        'max_streak':       max_streak,
        'max_consec_wins':  mcw,
        'max_consec_losses':mcl,
        'dd_pct':           round(dd_pct, 2),
        'drr':              round(max_streak, 2),
        'ror_pct':          round(ror, 4),
        'avg_mae_pct':      round(float(trades_df['mae_pct'].mean()), 4),
        'avg_mfe_pct':      round(float(trades_df['mfe_pct'].mean()), 4),
        'max_dd_dollars':   round(max_dd_usd, 2),
        'n_long':           int((trades_df['direction'] == 'LONG').sum()),
        'n_short':          int((trades_df['direction'] == 'SHORT').sum()),
        'lowest_equity':    lowest_equity,
        'account_blown':    account_blown,
        'n_breaches':       n_breaches,
        'breach_pct':       breach_pct,
        'max_consec_losing_days': _calc_max_consec_losing_days(trades_df),
        # Distribution stats
        'mae_median':       dist['mae_median'],
        'mae_mode':         dist['mae_mode'],
        'mae_mode_freq':     dist['mae_mode_freq'],
        'mfe_median':       dist['mfe_median'],
        'mfe_mode':         dist['mfe_mode'],
        'mfe_mode_freq':     dist['mfe_mode_freq'],
    }


# ══════════════════════════════════════════════════════════════════════════════
# THREE-LAYER INTEGRITY VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════
def run_integrity_check(day_skeletons, risk_per_trade, account_size,
                        mode, n_samples=20, budget_mode='flag'):
    log.info(f"\n{'='*60}")
    log.info(f"  THREE-LAYER INTEGRITY CHECK")
    log.info(f"{'='*60}")
    all_passed = True
    fails = []

    steps  = [round(x * 0.01, 2) for x in range(1, 101)]
    combos = list(product(steps, steps))
    rng    = np.random.default_rng(42)
    sample_combos = [combos[i] for i in rng.choice(len(combos), n_samples, replace=False)]

    compare_fields = ['direction','signal_color','entry_price','exit_price',
                      'sl_level','tp_level','exit_reason','contracts',
                      'pnl_points','pnl_dollars','mae_pct','mfe_pct','bars_in_trade']

    # LAYER 1: Fast vs Original
    log.info(f"\n  LAYER 1 — Fast vs Original ({n_samples} random combos)")
    for sl_pct, tp_pct in sample_combos:
        label = f"SL={sl_pct}% TP={tp_pct}%"
        fast  = run_backtest_fast(day_skeletons, sl_pct, tp_pct, risk_per_trade, budget_mode)
        orig  = run_backtest_original(day_skeletons, sl_pct, tp_pct, risk_per_trade, budget_mode)

        if len(fast) != len(orig):
            msg = f"  ✗ {label} | count mismatch: fast={len(fast)} orig={len(orig)}"
            log.error(msg); fails.append(msg); all_passed = False; continue

        if len(fast) == 0:
            log.info(f"  ✓ {label} | 0 trades — skipped"); continue

        combo_ok = True
        for field in compare_fields:
            fv = fast[field].values; ov = orig[field].values
            if field in ('entry_price','exit_price','sl_level','tp_level',
                         'pnl_points','pnl_dollars','mae_pct','mfe_pct'):
                if not np.allclose(fv.astype(float), ov.astype(float), atol=1e-4):
                    diff_idx = np.where(~np.isclose(fv.astype(float), ov.astype(float), atol=1e-4))[0]
                    msg = f"  ✗ {label} | field '{field}' mismatch row {diff_idx[0]}: fast={fv[diff_idx[0]]} orig={ov[diff_idx[0]]}"
                    log.error(msg); fails.append(msg); all_passed = False; combo_ok = False
            else:
                bad = fv != ov
                if bad.any():
                    idx = np.where(bad)[0][0]
                    msg = f"  ✗ {label} | field '{field}' mismatch row {idx}: fast='{fv[idx]}' orig='{ov[idx]}'"
                    log.error(msg); fails.append(msg); all_passed = False; combo_ok = False

        if combo_ok:
            log.info(f"  ✓ {label} | {len(fast)} trades | all fields match")

    # LAYER 2: Manual spot check
    log.info(f"\n  LAYER 2 — Manual spot check (3 known trades)")
    test_sl, test_tp = 0.20, 0.40
    spot = run_backtest_fast(day_skeletons, test_sl, test_tp, risk_per_trade, budget_mode)
    if len(spot) >= 3:
        for i in range(3):
            t  = spot.iloc[i]
            ep = t['entry_price']
            sl_exp = round(ep*(1-test_sl/100) if t['direction']=='LONG' else ep*(1+test_sl/100), 4)
            tp_exp = round(ep*(1+test_tp/100) if t['direction']=='LONG' else ep*(1-test_tp/100), 4)
            sl_ok  = abs(t['sl_level'] - sl_exp) < 0.01
            tp_ok  = abs(t['tp_level'] - tp_exp) < 0.01
            pnl_rc = round(((t['exit_price']-ep) if t['direction']=='LONG' else (ep-t['exit_price']))
                           * t['contracts'] * MNQ_MULTIPLIER, 2)
            pnl_ok = abs(t['pnl_dollars'] - pnl_rc) < 0.01
            if sl_ok and tp_ok and pnl_ok:
                log.info(f"  ✓ Trade {i+1}: entry={ep} SL={t['sl_level']} TP={t['tp_level']} P&L=${t['pnl_dollars']} — all correct")
            else:
                msg = f"  ✗ Trade {i+1}: SL_ok={sl_ok} TP_ok={tp_ok} PNL_ok={pnl_ok}"
                log.error(msg); fails.append(msg); all_passed = False
    else:
        log.info(f"  ⚠ Fewer than 3 trades — spot check skipped")

    # LAYER 3: Edge case stress
    log.info(f"\n  LAYER 3 — Edge case stress tests")
    tight = run_backtest_fast(day_skeletons, 0.01, 1.00, risk_per_trade, budget_mode)
    if not tight.empty:
        sl_rate = (tight['exit_reason']=='SL').mean()
        ok = sl_rate > 0.50
        log.info(f"  {'✓' if ok else '✗'} Tight SL (0.01%): SL rate={sl_rate:.1%} — {'expected >50%' if ok else 'UNEXPECTED'}")
        if not ok: fails.append("Tight SL stress: SL rate low"); all_passed = False

    wide = run_backtest_fast(day_skeletons, 1.00, 0.01, risk_per_trade, budget_mode)
    if not wide.empty:
        tp_rate = (wide['exit_reason']=='TP').mean()
        ok = tp_rate > 0.50
        log.info(f"  {'✓' if ok else '✗'} Wide SL/tight TP: TP rate={tp_rate:.1%} — {'expected >50%' if ok else 'UNEXPECTED'}")
        if not ok: fails.append("Wide SL stress: TP rate low"); all_passed = False

    if not spot.empty:
        mae_ok = (spot['mae_pct'] >= 0).all()
        mfe_ok = (spot['mfe_pct'] >= 0).all()
        if mae_ok and mfe_ok:
            log.info(f"  ✓ MAE/MFE non-negative for all trades")
        else:
            msg = "  ✗ Negative MAE or MFE found"
            log.error(msg); fails.append(msg); all_passed = False

    # Direction consistency check by mode
    if not spot.empty:
        if mode == 'long':
            ok = (spot['direction'] == 'LONG').all()
            log.info(f"  {'✓' if ok else '✗'} LONG ONLY mode: all trades are LONG")
            if not ok: fails.append("Long only: non-LONG trade found"); all_passed = False
        elif mode == 'short':
            ok = (spot['direction'] == 'SHORT').all()
            log.info(f"  {'✓' if ok else '✗'} SHORT ONLY mode: all trades are SHORT")
            if not ok: fails.append("Short only: non-SHORT trade found"); all_passed = False
        elif mode == 'breakout':
            long_ok  = (spot[spot['direction']=='LONG']['signal_color']  == 'GREEN').all()
            short_ok = (spot[spot['direction']=='SHORT']['signal_color'] == 'RED').all()
            log.info(f"  {'✓' if (long_ok and short_ok) else '✗'} BREAKOUT direction consistency")
            if not (long_ok and short_ok): fails.append("Breakout direction mismatch"); all_passed = False
        elif mode == 'reversal':
            long_ok  = (spot[spot['direction']=='LONG']['signal_color']  == 'RED').all()
            short_ok = (spot[spot['direction']=='SHORT']['signal_color'] == 'GREEN').all()
            log.info(f"  {'✓' if (long_ok and short_ok) else '✗'} REVERSAL direction consistency")
            if not (long_ok and short_ok): fails.append("Reversal direction mismatch"); all_passed = False

    log.info(f"\n{'='*60}")
    if all_passed:
        log.info(f"  ✅  ALL THREE LAYERS PASSED — engine is valid")
    else:
        log.error(f"  ❌  INTEGRITY FAILED — {len(fails)} issue(s):")
        for f in fails: log.error(f"      {f}")
        log.error(f"  DO NOT TRUST RESULTS — fix before grid sweep")
    log.info(f"{'='*60}\n")
    return all_passed


# ══════════════════════════════════════════════════════════════════════════════
# GRID SWEEP
# ══════════════════════════════════════════════════════════════════════════════
def run_grid(day_skeletons, account_size, risk_per_trade,
             label='', budget_mode='flag'):
    steps  = [round(x * 0.01, 2) for x in range(1, 101)]
    combos = list(product(steps, steps))

    log.info(f"  [{label}] {len(day_skeletons)} days cached. Sweeping {len(combos):,} combos...")

    try:
        from tqdm import tqdm
        iterator = tqdm(combos, desc=f"  [{label}]", unit="combo",
                        bar_format='{l_bar}{bar:35}{r_bar}', colour='yellow')
    except ImportError:
        iterator = combos

    null_fields = ['total_trades','n_wins','n_losses','n_breakeven','win_pct','loss_pct',
                   'avg_win_usd','avg_loss_usd','ratio_win_loss','gross_profit','gross_loss',
                   'total_pnl','largest_win','largest_loss','ev','profit_factor',
                   'combined_edge','sqn','max_streak','max_consec_wins','max_consec_losses',
                   'dd_pct','drr','ror_pct','avg_mae_pct','avg_mfe_pct','max_dd_dollars',
                   'n_long','n_short','lowest_equity','account_blown','n_breaches','breach_pct',
                   'max_consec_losing_days',
                   'mae_median','mae_mode','mae_mode_freq','mfe_median','mfe_mode','mfe_mode_freq']

    results    = []
    all_trades = {}

    for sl_pct, tp_pct in iterator:
        trades  = run_backtest_fast(day_skeletons, sl_pct, tp_pct, risk_per_trade, budget_mode)
        metrics = calc_metrics(trades, account_size, risk_per_trade)
        rr      = round(tp_pct / sl_pct, 2) if sl_pct > 0 else 0
        key     = f"SL{sl_pct}_TP{tp_pct}"

        if metrics is None:
            results.append({'sl_pct':sl_pct,'tp_pct':tp_pct,'rr_ratio':rr,
                            'passed':False,'fail_reason':'No trades',
                            **{k:None for k in null_fields}})
            continue

        passed      = metrics['profit_factor'] >= 1.00
        fail_reason = '' if passed else f"PF={metrics['profit_factor']:.3f}<1.00"
        row = {'sl_pct':sl_pct,'tp_pct':tp_pct,'rr_ratio':rr,
               'passed':passed,'fail_reason':fail_reason}
        row.update(metrics)
        results.append(row)
        if passed:
            all_trades[key] = trades

    return pd.DataFrame(results), all_trades


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — LEADERBOARD
# ══════════════════════════════════════════════════════════════════════════════
def build_leaderboard_sheet(ws, results_df, label, mode_label):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD
    passed = results_df[results_df['passed']==True].copy()
    passed = passed.sort_values('combined_edge', ascending=False).reset_index(drop=True)

    ncols = 39
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    wc(ws,1,1,f'🏆  LEADERBOARD — {mode_label} | {label} | Sorted by Combined Edge',
       font=hf(13,True,GOLD), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
    wc(ws,2,1,f'Passing combos (PF ≥ 1.00): {len(passed)} of {len(results_df):,}',
       font=cf(10,False,MUTED), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[2].height = 18

    headers = ['Rank','SL %','TP %','RR','Trades','Wins','Losses','BE',
               'Win %','Avg Win $','Avg Loss $','Ratio W/L',
               'Gross Profit $','Gross Loss $','Total P&L $',
               'Largest Win $','Largest Loss $',
               'EV $','PF','Comb Edge','SQN',
               'Max Streak','Max W Run','Max L Run','DD %','DRR','RoR %',
               'Avg MAE %','Avg MFE %',
               'Max DD $','Low Eq $','Blown','Breaches',
               'MAE Median %','MAE Mode %','MAE Mode Freq',
               'MFE Median %','MFE Mode %','MFE Mode Freq']
    col_keys = ['_rank','sl_pct','tp_pct','rr_ratio','total_trades','n_wins','n_losses','n_breakeven',
                'win_pct','avg_win_usd','avg_loss_usd','ratio_win_loss',
                'gross_profit','gross_loss','total_pnl','largest_win','largest_loss',
                'ev','profit_factor','combined_edge','sqn',
                'max_streak','max_consec_wins','max_consec_losses','dd_pct','drr','ror_pct',
                'avg_mae_pct','avg_mfe_pct',
                'max_dd_dollars','lowest_equity','account_blown','n_breaches',
                'mae_median','mae_mode','mae_mode_freq',
                'mfe_median','mfe_mode','mfe_mode_freq']
    col_widths=[6,7,7,7,8,7,8,6,
                8,11,12,10,
                14,13,13,14,14,
                10,9,11,9,
                11,11,11,8,8,9,
                12,12,
                12,12,8,10,
                12,12,12,
                12,12,12]

    for c,(h,w) in enumerate(zip(headers,col_widths),1):
        wc(ws,3,c,h, font=hf(9,True,DARK_BG), fill_=fill(GOLD),
           align=center(), border=tborder(DARK_BG))
        wcol(ws,c,w)
    ws.row_dimensions[3].height = 20

    for r_idx, row in passed.iterrows():
        er  = r_idx + 4
        bg  = RAISED_BG if r_idx % 2 == 0 else CARD_BG
        rv  = dict(row)
        rv['_rank'] = r_idx + 1

        for c, key in enumerate(col_keys, 1):
            val  = rv.get(key)
            fc   = WHITE; bold = False; fmt = None
            if key == '_rank':         fc = GOLD; bold = True
            if key == 'combined_edge': fc = TEAL; bold = True
            if key == 'total_pnl':
                fc = TEAL if (val or 0) >= 0 else RED_CLR; bold = True
                fmt = '$#,##0.00'
            if key in ('avg_win_usd','avg_loss_usd','gross_profit','gross_loss',
                       'largest_win','largest_loss','ev','max_dd_dollars','lowest_equity'):
                fmt = '$#,##0.00'
            if key in ('sl_pct','tp_pct','win_pct','dd_pct','ror_pct','breach_pct'):
                fmt = '0.00"%"'
            if key == 'account_blown':
                val = '⚠ YES' if val else 'NO'
                fc  = RED_CLR if val == '⚠ YES' else TEAL
            if key == 'profit_factor':
                fc = TEAL if (val or 0) >= 1.5 else (GOLD if (val or 0) >= 1.0 else RED_CLR)
            # Distribution stat coloring
            if key in ('mae_median','mae_mode','mfe_median','mfe_mode'):
                fc = TEAL if val is not None else MUTED
            if key in ('mae_mode_freq','mfe_mode_freq') and val is not None:
                fc = TEAL if val >= 50 else (GOLD if val >= 25 else WHITE)
                fmt = '0.0"%"'
            wc(ws, er, c, val,
               font=Font(name='Arial', size=9, color=fc, bold=bold),
               fill_=fill(bg), align=center(), border=tborder(), fmt=fmt)
        ws.row_dimensions[er].height = 15

    ws.freeze_panes = 'A4'


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — TRADE LOG
# ══════════════════════════════════════════════════════════════════════════════
def build_trade_log_sheet(ws, trades_df, sl_pct, tp_pct, mode_label):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = BLUE_CLR

    headers   = ['#','Date','Fixed Candle','Entry Time','Mode','Direction','Signal',
                 'Entry $','Exit $','SL Level','TP Level','Exit',
                 'Contracts','Risk $','Budget','P&L pts','P&L $','Cumul $','MAE %','MFE %','Bars']
    col_widths= [5,12,16,14,10,11,8,12,12,12,12,7,10,11,9,10,12,13,9,9,6]
    col_keys  = ['_num','date','fixed_dt','entry_time','mode','direction','signal_color',
                 'entry_price','exit_price','sl_level','tp_level','exit_reason',
                 'contracts','actual_risk','budget_breach',
                 'pnl_points','pnl_dollars','_cumul','bars_in_trade']

    ncols = len(headers)
    ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
    wc(ws,1,1,f'📋  TRADE LOG — {mode_label} | SL: {sl_pct}% | TP: {tp_pct}%',
       font=hf(12,True,BLUE_CLR), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 28

    for c,(h,w) in enumerate(zip(headers,col_widths),1):
        wc(ws,2,c,h, font=hf(9,True,DARK_BG), fill_=fill(BLUE_CLR),
           align=center(), border=tborder(DARK_BG))
        wcol(ws,c,w)
    ws.row_dimensions[2].height = 20

    cumul = 0.0
    trades_df = trades_df.reset_index(drop=True)
    for r_idx, trade in trades_df.iterrows():
        er     = r_idx + 3
        cumul += trade['pnl_dollars']
        breach = bool(trade.get('budget_breach', False))
        is_win = trade['pnl_dollars'] > 0
        is_loss= trade['pnl_dollars'] < 0
        bg = ('3D1A00' if r_idx%2==0 else '2E1200') if breach else \
             (('1A2E1A' if r_idx%2==0 else RAISED_BG) if is_win else \
             (('2E1A1A' if r_idx%2==0 else RAISED_BG) if is_loss else CARD_BG))

        rv = {
            '_num':         r_idx + 1,
            'date':         str(trade['date']),
            'fixed_dt':     str(trade['fixed_dt']),
            'entry_time':   str(trade['entry_time']),
            'mode':         str(trade.get('mode','')).upper(),
            'direction':    trade['direction'],
            'signal_color': trade['signal_color'],
            'entry_price':  trade['entry_price'],
            'exit_price':   trade['exit_price'],
            'sl_level':     trade['sl_level'],
            'tp_level':     trade['tp_level'],
            'exit_reason':  trade['exit_reason'],
            'contracts':    int(trade.get('contracts',1)),
            'actual_risk':  trade.get('actual_risk',0),
            'budget_breach':trade.get('budget_breach',False),
            'pnl_points':   trade['pnl_points'],
            'pnl_dollars':  trade['pnl_dollars'],
            '_cumul':       round(cumul,2),
            'mae_pct':      trade.get('mae_pct',0),
            'mfe_pct':      trade.get('mfe_pct',0),
            'bars_in_trade':int(trade.get('bars_in_trade',0)),
        }

        for c, key in enumerate(col_keys, 1):
            val=rv[key]; fc=WHITE; bold=False; fmt=None
            if key=='_num':         fc=MUTED
            if key=='direction':    fc=TEAL if val=='LONG' else RED_CLR; bold=True
            if key=='mode':         fc=GOLD
            if key=='signal_color': fc=TEAL if val=='GREEN' else RED_CLR
            if key=='exit_reason':
                fc=TEAL if val=='TP' else (RED_CLR if val=='SL' else GOLD); bold=val in('TP','SL')
            if key=='pnl_dollars':
                fc=TEAL if(isinstance(val,(int,float))and val>0)else RED_CLR; bold=True; fmt='$#,##0.00'
            if key=='_cumul':
                fc=TEAL if(isinstance(val,(int,float))and val>=0)else RED_CLR; bold=True; fmt='$#,##0.00'
            if key=='actual_risk': fc=RED_CLR if breach else TEAL; fmt='$#,##0.00'
            if key=='budget_breach':
                val='⚠ BREACH' if val else 'OK'; fc=RED_CLR if val=='⚠ BREACH' else TEAL; bold=val=='⚠ BREACH'
            if key in('entry_price','exit_price','sl_level','tp_level'): fmt='#,##0.00'
            if key=='pnl_points': fmt='0.00'
            if key in('mae_pct','mfe_pct'): fmt='0.0000"%"'; fc=MUTED
            wc(ws,er,c,val,font=Font(name='Arial',size=9,color=fc,bold=bold),
               fill_=fill(bg),align=center(),border=tborder(),fmt=fmt)
        ws.row_dimensions[er].height = 15
    ws.freeze_panes='A3'


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — EQUITY CURVE
# ══════════════════════════════════════════════════════════════════════════════
def build_equity_sheet(ws, results_df, all_trades, account_size, mode_label):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GREEN_CLR
    passed = results_df[results_df['passed']==True].sort_values(
        'combined_edge', ascending=False).head(5).reset_index(drop=True)

    ws.merge_cells('A1:G1')
    wc(ws,1,1,f'📈  EQUITY CURVE — Top 5 Combos | {mode_label}',
       font=hf(13,True,GREEN_CLR), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 28
    wc(ws,2,1,'Trade #', font=hf(9,True,DARK_BG), fill_=fill(GREEN_CLR), align=center()); wcol(ws,1,10)

    for ri, row in passed.iterrows():
        key = f"SL{row['sl_pct']}_TP{row['tp_pct']}"
        col = ri + 2
        label = f"SL{row['sl_pct']}%/TP{row['tp_pct']}%"
        wc(ws,2,col,label, font=hf(9,True,DARK_BG), fill_=fill(GREEN_CLR), align=center())
        wcol(ws,col,16)
        if key not in all_trades: continue
        t = all_trades[key].reset_index(drop=True)
        cumul = 0.0
        for ti, trade in t.iterrows():
            cumul += trade['pnl_dollars']
            er = ti + 3
            wc(ws,er,1, ti+1, font=cf(9), fill_=fill(RAISED_BG if ti%2==0 else CARD_BG), align=center())
            equity = account_size + cumul
            fc = TEAL if equity >= account_size else RED_CLR
            wc(ws,er,col, round(equity,2),
               font=Font(name='Arial',size=9,color=fc,bold=True),
               fill_=fill(RAISED_BG if ti%2==0 else CARD_BG), align=center(), fmt='$#,##0.00')
    ws.freeze_panes='A3'


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — DAILY BREAKDOWN
# ══════════════════════════════════════════════════════════════════════════════
def build_daily_sheet(ws, trades_df, sl_pct, tp_pct, mode_label):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '9B59B6'

    ws.merge_cells('A1:H1')
    wc(ws,1,1,f'📅  DAILY BREAKDOWN | {mode_label} | SL:{sl_pct}% TP:{tp_pct}%',
       font=hf(12,True,'9B59B6'), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height = 28

    headers    = ['Date','Trades','Wins','Losses','Win %','P&L $','Cumul $','Result']
    col_widths = [13,9,8,9,9,13,14,10]
    for c,(h,w) in enumerate(zip(headers,col_widths),1):
        wc(ws,2,c,h, font=hf(9,True,DARK_BG), fill_=fill('9B59B6'),
           align=center(), border=tborder(DARK_BG))
        wcol(ws,c,w)
    ws.row_dimensions[2].height = 20

    daily = trades_df.groupby('date').agg(
        trades=('pnl_dollars','count'),
        wins  =('pnl_dollars', lambda x:(x>0).sum()),
        losses=('pnl_dollars', lambda x:(x<0).sum()),
        pnl   =('pnl_dollars','sum'),
    ).reset_index().sort_values('date').reset_index(drop=True)

    cumul=0.0
    for r_idx, row in daily.iterrows():
        er=r_idx+3; cumul+=row['pnl']
        wp = row['wins']/row['trades']*100 if row['trades']>0 else 0
        result='WIN' if row['pnl']>0 else('LOSS' if row['pnl']<0 else 'FLAT')
        bg=RAISED_BG if r_idx%2==0 else CARD_BG
        vals=[str(row['date']),int(row['trades']),int(row['wins']),int(row['losses']),
              f"{wp:.0f}%",row['pnl'],round(cumul,2),result]
        fmts=[None,None,None,None,None,'$#,##0.00','$#,##0.00',None]
        for c,(val,fmt) in enumerate(zip(vals,fmts),1):
            fc=WHITE
            if c==6: fc=TEAL if row['pnl']>=0 else RED_CLR
            if c==7: fc=TEAL if cumul>=0 else RED_CLR
            if c==8: fc=TEAL if row['pnl']>0 else(RED_CLR if row['pnl']<0 else GOLD)
            wc(ws,er,c,val,font=Font(name='Arial',size=9,color=fc,bold=(c in(6,7,8))),
               fill_=fill(bg),align=center(),border=tborder(),fmt=fmt)
        ws.row_dimensions[er].height=15
    ws.freeze_panes='A3'


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — DRAWDOWN TIMELINE
# ══════════════════════════════════════════════════════════════════════════════
def build_drawdown_sheet(ws, trades_df, sl_pct, tp_pct, account_size, mode_label):
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = RED_CLR

    ws.merge_cells('A1:I1')
    wc(ws,1,1,f'📉  DRAWDOWN TIMELINE | {mode_label} | SL:{sl_pct}% TP:{tp_pct}%',
       font=hf(12,True,RED_CLR), fill_=fill(DARK_BG), align=center())
    ws.row_dimensions[1].height=28

    headers   =['#','Date','Entry Time','Exit','P&L $','Equity $','Peak $','DD from Peak $','Blown?']
    col_widths=[5,12,14,7,12,13,13,16,8]
    for c,(h,w) in enumerate(zip(headers,col_widths),1):
        wc(ws,2,c,h,font=hf(9,True,DARK_BG),fill_=fill(RED_CLR),align=center(),border=tborder(DARK_BG))
        wcol(ws,c,w)
    ws.row_dimensions[2].height=20

    cumul=0.0; peak=float(account_size)
    trades_df=trades_df.reset_index(drop=True)
    for r_idx,trade in trades_df.iterrows():
        er=r_idx+3; cumul+=trade['pnl_dollars']
        equity=account_size+cumul
        peak=max(peak,equity)
        dd=peak-equity
        blown=equity<0
        bg=RAISED_BG if r_idx%2==0 else CARD_BG
        vals=[r_idx+1,str(trade['date']),str(trade['entry_time']),
              trade['exit_reason'],trade['pnl_dollars'],
              round(equity,2),round(peak,2),round(dd,2),'⚠ YES' if blown else 'NO']
        fmts=[None,None,None,None,'$#,##0.00','$#,##0.00','$#,##0.00','$#,##0.00',None]
        for c,(val,fmt) in enumerate(zip(vals,fmts),1):
            fc=WHITE
            if c==4: fc=TEAL if val=='TP' else(RED_CLR if val=='SL' else GOLD)
            if c==5: fc=TEAL if(isinstance(val,(int,float))and val>0)else RED_CLR
            if c==6: fc=TEAL if(isinstance(val,(int,float))and val>=account_size)else RED_CLR
            if c==8: fc=RED_CLR if(isinstance(val,(int,float))and val>0)else TEAL
            if c==9: fc=RED_CLR if blown else TEAL
            wc(ws,er,c,val,font=Font(name='Arial',size=9,color=fc,bold=(c in(5,6,8))),
               fill_=fill(bg),align=center(),border=tborder(),fmt=fmt)
        ws.row_dimensions[er].height=15
    ws.freeze_panes='A3'


# ══════════════════════════════════════════════════════════════════════════════
# BUILD FULL EXCEL WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
def build_excel(results_df, all_trades, label, account_size, risk_per_trade,
                mode_label, out_dir):
    wb = Workbook()

    passed = results_df[results_df['passed']==True].sort_values('combined_edge', ascending=False)
    best_sl = float(passed.iloc[0]['sl_pct']) if len(passed) > 0 else 0.20
    best_tp = float(passed.iloc[0]['tp_pct']) if len(passed) > 0 else 0.40
    best_key= f"SL{best_sl}_TP{best_tp}"
    best_trades = all_trades.get(best_key, pd.DataFrame())

    ws_lb  = wb.active;                ws_lb.title  = 'Leaderboard'
    ws_tl  = wb.create_sheet('Trade Log')
    ws_eq  = wb.create_sheet('Equity Curve')
    ws_day = wb.create_sheet('Daily Breakdown')
    ws_dd  = wb.create_sheet('Drawdown')

    build_leaderboard_sheet(ws_lb, results_df, label, mode_label)
    if not best_trades.empty:
        build_trade_log_sheet(ws_tl, best_trades, best_sl, best_tp, mode_label)
        build_equity_sheet(ws_eq, results_df, all_trades, account_size, mode_label)
        build_daily_sheet(ws_day, best_trades, best_sl, best_tp, mode_label)
        build_drawdown_sheet(ws_dd, best_trades, best_sl, best_tp, account_size, mode_label)

    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname= f"{mode_label}_{label}_{ts}.xlsx"
    path = os.path.join(out_dir, fname)
    wb.save(path)
    log.info(f"  Saved: {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# MASTER CONSENSUS
# ══════════════════════════════════════════════════════════════════════════════
def build_master_consensus(all_period_results, period_days, mode_label, out_dir):
    wb = Workbook()
    labels    = list(all_period_results.keys())
    n_periods = len(labels)

    if n_periods < 2:
        ws = wb.active; ws.title = 'CONSENSUS'
        wc(ws,1,1,'Need at least 2 periods for consensus.',
           font=hf(12,True,RED_CLR), fill_=fill(DARK_BG), align=center())
        path = os.path.join(out_dir, f'{mode_label}_MASTER.xlsx')
        wb.save(path); return path

    dfs = {}
    for lbl in labels:
        df = all_period_results[lbl]
        passed = df[df['passed']==True].copy()
        passed['_key'] = list(zip(passed['sl_pct'].round(2), passed['tp_pct'].round(2)))
        passed = passed.sort_values('combined_edge', ascending=False).drop_duplicates('_key')
        dfs[lbl] = {row['_key']: row for _, row in passed.iterrows()}

    common_keys = set(dfs[labels[0]].keys())
    for lbl in labels[1:]:
        common_keys &= set(dfs[lbl].keys())

    min_trades_per_day = 0.33
    min_trades = {}
    for lbl, days in period_days.items():
        trading_days = int(days * 5/7)
        min_trades[lbl] = max(10, int(trading_days * min_trades_per_day))

    filtered_keys = set()
    for key in common_keys:
        if all(int(dfs[lbl][key]['total_trades']) >= min_trades[lbl] for lbl in labels):
            filtered_keys.add(key)

    log.info(f"\nMaster Consensus: {len(common_keys)} combos pass PF≥1.00 across all {n_periods} periods")
    log.info(f"  After min-trade filter: {len(filtered_keys)} remain")

    if not filtered_keys:
        ws = wb.active; ws.title = 'CONSENSUS'
        wc(ws,1,1,'No combos passed PF≥1.00 across all periods.',
           font=hf(12,True,RED_CLR), fill_=fill(DARK_BG), align=center())
        path = os.path.join(out_dir, f'{mode_label}_MASTER.xlsx')
        wb.save(path); return path

    rows = []
    for key in filtered_keys:
        sl, tp = key
        rr = round(tp/sl, 2) if sl > 0 else 0
        pd_data = {lbl: dfs[lbl][key] for lbl in labels}
        ces  = [float(pd_data[lbl]['combined_edge']) for lbl in labels]
        sqns = [float(pd_data[lbl].get('sqn', 0))   for lbl in labels]
        pfs  = [float(pd_data[lbl]['profit_factor']) for lbl in labels]
        wps  = [float(pd_data[lbl]['win_pct'])       for lbl in labels]
        evs  = [float(pd_data[lbl]['ev'])            for lbl in labels]
        rors = [float(pd_data[lbl]['ror_pct'])       for lbl in labels]
        drrs = [float(pd_data[lbl]['drr'])           for lbl in labels]
        pnls = [float(pd_data[lbl]['total_pnl'])     for lbl in labels]
        avg_ce  = round(float(np.mean(ces)), 4)
        ce_std  = round(float(np.std(ces)),  4)
        consistency = round(avg_ce / (ce_std + 0.0001), 2)

        row = {
            'sl_pct': sl, 'tp_pct': tp, 'rr_ratio': rr,
            'avg_combined_edge':  avg_ce,
            'avg_sqn':            round(float(np.mean(sqns)),  4),
            'avg_profit_factor':  round(float(np.mean(pfs)),   3),
            'avg_win_pct':        round(float(np.mean(wps)),   2),
            'avg_ev_usd':         round(float(np.mean(evs)),   2),
            'avg_ror_pct':        round(float(np.mean(rors)),  4),
            'avg_drr':            round(float(np.mean(drrs)),  2),
            'avg_total_pnl':      round(float(np.mean(pnls)),  2),
            'consistency_score':  consistency,
            'ce_std':             ce_std,
        }
        profitable_periods = sum(1 for lbl in labels if float(pd_data[lbl]['total_pnl']) > 0)
        row['profitable_periods'] = f"{profitable_periods}/{len(labels)}"
        for lbl in labels:
            r = pd_data[lbl]
            row[f'ce_{lbl}']      = round(float(r['combined_edge']),  4)
            row[f'pf_{lbl}']      = round(float(r['profit_factor']),  3)
            row[f'pnl_{lbl}']     = round(float(r['total_pnl']),      2)
            row[f'trades_{lbl}']  = int(r['total_trades'])
            row[f'loweq_{lbl}']   = round(float(r.get('lowest_equity', 0)), 2)
            row[f'blown_{lbl}']   = bool(r.get('account_blown', False))
            row[f'mcldays_{lbl}'] = int(r.get('max_consec_losing_days', 0))
            row[f'mcltrds_{lbl}'] = int(r.get('max_consec_losses', 0))
        rows.append(row)

    cdf = pd.DataFrame(rows).sort_values('avg_combined_edge', ascending=False).reset_index(drop=True)

    per_period_headers = []
    per_period_keys    = []
    per_period_widths  = []
    for lbl in labels:
        per_period_headers += [f'CE {lbl}', f'PF {lbl}', f'P&L {lbl}', f'Trades {lbl}', f'Low Eq {lbl}', f'Blown {lbl}']
        per_period_keys    += [f'ce_{lbl}', f'pf_{lbl}', f'pnl_{lbl}', f'trades_{lbl}', f'loweq_{lbl}', f'blown_{lbl}']
        per_period_widths  += [10, 9, 12, 11, 12, 9]

    trail_headers = (['Prof Periods'] +
                     [f'Max Loss Days {lbl}' for lbl in labels] +
                     [f'Max Loss Trades {lbl}' for lbl in labels])
    trail_keys    = (['profitable_periods'] +
                     [f'mcldays_{lbl}' for lbl in labels] +
                     [f'mcltrds_{lbl}' for lbl in labels])
    trail_widths  = [13] + [15] * len(labels) + [16] * len(labels)

    mh = (['Rank','SL %','TP %','RR','Avg CE','Avg SQN','Avg PF','Avg Win %','Avg EV $',
           'Avg RoR %','Avg DRR','Avg P&L $','Consistency','CE Std'] + per_period_headers + trail_headers)
    mk = ([None,'sl_pct','tp_pct','rr_ratio','avg_combined_edge','avg_sqn','avg_profit_factor',
           'avg_win_pct','avg_ev_usd','avg_ror_pct','avg_drr','avg_total_pnl',
           'consistency_score','ce_std'] + per_period_keys + trail_keys)
    mw = ([6,7,7,7,11,10,10,10,11,10,9,12,13,10] + per_period_widths + trail_widths)
    ncols = len(mh)

    def _render_consensus(ws, sorted_df, title, subtitle, hdr_color):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = hdr_color
        ws.merge_cells(f'A1:{get_column_letter(ncols)}1')
        wc(ws,1,1,title, font=hf(13,True,hdr_color), fill_=fill(DARK_BG), align=center())
        ws.row_dimensions[1].height = 30
        ws.merge_cells(f'A2:{get_column_letter(ncols)}2')
        wc(ws,2,1,f'Qualifying combos: {len(sorted_df)}  |  {subtitle}  |  Mode: {mode_label}',
           font=cf(10,False,MUTED), fill_=fill(DARK_BG), align=center())
        ws.row_dimensions[2].height = 18
        for c,(h,w) in enumerate(zip(mh,mw),1):
            wc(ws,3,c,h, font=hf(9,True,DARK_BG), fill_=fill(hdr_color),
               align=center(), border=tborder(DARK_BG))
            wcol(ws,c,w)
        ws.row_dimensions[3].height = 22

        for r_idx, row in sorted_df.reset_index(drop=True).iterrows():
            er = r_idx + 4
            rank = r_idx + 1
            bg = '2E2A00' if rank==1 else ('1A2E1A' if rank<=10 else ('1F2433' if r_idx%2==0 else RAISED_BG))
            values = [rank] + [row.get(k) for k in mk[1:]]
            for c,(val,key) in enumerate(zip(values,mk),1):
                fc=WHITE; bold=False; fmt=None
                if c == 1: fc=GOLD; bold=(rank<=3)
                elif key=='avg_combined_edge':
                    fc=TEAL if (val and val>=0.40) else (GOLD if (val and val>=0.20) else WHITE); bold=True
                elif key=='avg_sqn' and val is not None:
                    fc=TEAL if val>=2.0 else (GOLD if val>=1.6 else (WHITE if val>=0 else RED_CLR)); bold=(val>=2.0)
                elif key=='avg_profit_factor' and val:
                    fc=TEAL if val>=1.5 else (GOLD if val>=1.0 else RED_CLR)
                elif key=='avg_ror_pct' and val and val>1: fc=RED_CLR; bold=True
                elif key=='avg_drr' and val and val>8:     fc=RED_CLR; bold=True
                elif key in ('avg_total_pnl',) + tuple(f'pnl_{l}' for l in labels):
                    fc=TEAL if (val and val>=0) else RED_CLR
                elif key=='consistency_score' and val:
                    fc=TEAL if val>5 else (GOLD if val>2 else WHITE)
                elif key in tuple(f'ce_{l}' for l in labels) and val:
                    fc=TEAL if val>=0.20 else (GOLD if val>=0 else RED_CLR)
                elif key in tuple(f'blown_{l}' for l in labels) and val is not None:
                    fc=RED_CLR if val else TEAL
                elif key in tuple(f'loweq_{l}' for l in labels) and val is not None:
                    fc=RED_CLR if val<0 else (GOLD if val<500 else TEAL); bold=(val<0)
                elif key == 'profitable_periods' and val is not None:
                    try:
                        num, den = (int(x) for x in str(val).split('/'))
                        fc = TEAL if num==den else (GOLD if num>0 else RED_CLR)
                        bold = (num==den)
                    except Exception: pass
                elif key in tuple(f'mcldays_{l}' for l in labels) and val is not None:
                    fc = TEAL if val<=2 else (GOLD if val<=4 else RED_CLR); bold=(val>=5)
                elif key in tuple(f'mcltrds_{l}' for l in labels) and val is not None:
                    fc = TEAL if val<=3 else (GOLD if val<=6 else RED_CLR); bold=(val>=7)

                if key in ('sl_pct','tp_pct','avg_win_pct'): fmt='0.00"%"'
                elif key in ('avg_sqn',): fmt='0.0000'
                elif key in ('avg_ev_usd','avg_total_pnl') + tuple(f'pnl_{l}' for l in labels): fmt='$#,##0.00'
                elif key in tuple(f'loweq_{l}' for l in labels): fmt='$#,##0.00'
                elif key in ('avg_combined_edge','ce_std','consistency_score') + tuple(f'ce_{l}' for l in labels): fmt='0.0000'
                elif key in ('avg_profit_factor','rr_ratio','avg_drr') + tuple(f'pf_{l}' for l in labels): fmt='0.000'
                elif key=='avg_ror_pct': fmt='0.0000"%"'

                if key in tuple(f'blown_{l}' for l in labels) and val is not None:
                    val = 'YES' if val else 'NO'

                wc(ws,er,c,val, font=Font(name='Arial',size=9,bold=bold,color=fc),
                   fill_=fill(bg), align=center(), border=tborder(), fmt=fmt)
            ws.row_dimensions[er].height = 16
        ws.freeze_panes = 'A4'

    ws1 = wb.active; ws1.title = 'CONSENSUS by Edge'
    _render_consensus(ws1, cdf.sort_values('avg_combined_edge', ascending=False),
        f'🏆  MASTER CONSENSUS — {mode_label} | Ranked by Avg Combined Edge',
        'Sorted by Avg Combined Edge', GOLD)

    _render_consensus(wb.create_sheet('CONSENSUS by SQN'),
        cdf.sort_values('avg_sqn', ascending=False),
        f'📐  MASTER CONSENSUS — {mode_label} | Ranked by Avg SQN',
        'Sorted by Avg SQN', TEAL)

    _render_consensus(wb.create_sheet('CONSENSUS by Survival'),
        cdf.sort_values('avg_ror_pct', ascending=True),
        f'🛡  MASTER CONSENSUS — {mode_label} | Ranked by Risk of Ruin',
        'Sorted by Avg RoR % ascending', '4A90D9')

    _render_consensus(wb.create_sheet('CONSENSUS by DRR'),
        cdf.sort_values('avg_drr', ascending=True),
        f'📉  MASTER CONSENSUS — {mode_label} | Ranked by DRR',
        'Sorted by Avg DRR ascending', 'D97A4A')

    # BEST PROFILE — ordinal rank aggregation
    ws_best = wb.create_sheet('BEST PROFILE')
    ws_best.sheet_view.showGridLines = False
    ws_best.sheet_properties.tabColor = GOLD
    if not cdf.empty:
        bp = cdf.copy()
        bp['rank_edge']     = bp['avg_combined_edge'].rank(ascending=False, method='min').astype(int)
        bp['rank_sqn']      = bp['avg_sqn'].rank(ascending=False, method='min').astype(int)
        bp['rank_survival'] = bp['avg_ror_pct'].rank(ascending=True,  method='min').astype(int)
        bp['rank_drr']      = bp['avg_drr'].rank(ascending=True,       method='min').astype(int)
        bp['total_rank']    = bp['rank_edge'] + bp['rank_sqn'] + bp['rank_survival'] + bp['rank_drr']
        bp = bp.sort_values('total_rank', ascending=True).reset_index(drop=True)

        bp_cols = 13
        ws_best.merge_cells(f'A1:{get_column_letter(bp_cols)}1')
        wc(ws_best,1,1,f'🥇  BEST PROFILE — {mode_label} | Ordinal Rank Aggregation',
           font=hf(13,True,GOLD), fill_=fill(DARK_BG), align=center())
        ws_best.row_dimensions[1].height = 32
        ws_best.merge_cells(f'A2:{get_column_letter(bp_cols)}2')
        wc(ws_best,2,1,
           'Total Rank = Rank(Edge) + Rank(SQN) + Rank(Survival) + Rank(DRR)  |  Lowest total = best all-round live combo',
           font=cf(10,False,MUTED), fill_=fill(DARK_BG), align=center())
        ws_best.row_dimensions[2].height = 18

        bph = ['Rank','SL %','TP %','RR','Total Rank','Rank Edge','Rank SQN','Rank Survival','Rank DRR',
               'Avg CE','Avg SQN','Avg RoR %','Avg DRR']
        bpk = [None,'sl_pct','tp_pct','rr_ratio','total_rank','rank_edge','rank_sqn','rank_survival','rank_drr',
               'avg_combined_edge','avg_sqn','avg_ror_pct','avg_drr']
        bpw = [6,7,7,7,11,11,11,14,11,12,12,12,11]
        for c,(h,w) in enumerate(zip(bph,bpw),1):
            wc(ws_best,3,c,h, font=hf(9,True,DARK_BG), fill_=fill(GOLD),
               align=center(), border=tborder(DARK_BG))
            wcol(ws_best,c,w)
        ws_best.row_dimensions[3].height = 22
        for r_idx, row in bp.iterrows():
            er=r_idx+4; rank=r_idx+1
            bg='3D2E00' if rank==1 else ('1A2E1A' if rank<=3 else ('1F2433' if r_idx%2==0 else RAISED_BG))
            values=[rank]+[row.get(k) for k in bpk[1:]]
            for c,(val,key) in enumerate(zip(values,bpk),1):
                fc=WHITE; bold=False; fmt=None
                if c==1: fc=GOLD if rank==1 else (TEAL if rank<=3 else WHITE); bold=rank<=3
                elif key=='total_rank': fc=GOLD if rank==1 else(TEAL if rank<=3 else WHITE); bold=rank<=3
                elif key in('rank_edge','rank_sqn','rank_survival','rank_drr') and val is not None:
                    fc=TEAL if val<=3 else(GOLD if val<=10 else WHITE)
                elif key=='avg_combined_edge':
                    fc=TEAL if(val and val>=0.40)else(GOLD if(val and val>=0.20)else WHITE); bold=True
                elif key=='avg_sqn' and val is not None:
                    fc=TEAL if val>=2.0 else(GOLD if val>=1.6 else(WHITE if val>=0 else RED_CLR))
                elif key=='avg_ror_pct' and val is not None:
                    fc=TEAL if val<1 else(GOLD if val<5 else RED_CLR)
                elif key=='avg_drr' and val is not None:
                    fc=TEAL if val<4 else(GOLD if val<8 else RED_CLR)
                if key in('sl_pct','tp_pct'): fmt='0.00"%"'
                elif key=='avg_combined_edge': fmt='0.0000'
                elif key=='avg_sqn':           fmt='0.0000'
                elif key=='avg_ror_pct':        fmt='0.0000"%"'
                elif key in('avg_drr','rr_ratio'): fmt='0.000'
                wc(ws_best,er,c,val, font=Font(name='Arial',size=9,bold=bold,color=fc),
                   fill_=fill(bg), align=center(), border=tborder(), fmt=fmt)
            ws_best.row_dimensions[er].height=16
        ws_best.freeze_panes='A4'

    ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(out_dir, f'{mode_label}_MASTER_{ts}.xlsx')
    wb.save(path)
    log.info(f"  Master saved: {path}")
    return path


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print()
    print("=" * 60)
    print("  MNQ FIXED CANDLE STRATEGY TESTER v1")
    print("  One candle. One trade per day. Four modes.")
    print("=" * 60)

    # ── Data file ──
    files = find_csv_files()
    if not files:
        filepath = input("\nM1 data CSV path: ").strip().strip('"').strip("'")
    elif len(files) == 1:
        filepath = files[0]; log.info(f"Found: {filepath}")
    else:
        print("\nFiles found:")
        for i,f in enumerate(files): print(f"  [{i+1}] {f}")
        while True:
            try:
                c=int(input("Select: "))-1
                if 0<=c<len(files): filepath=files[c]; break
            except (ValueError,KeyboardInterrupt): pass

    df_1m = load_ohlc(filepath)

    # ── Candle timeframe ──
    while True:
        try:
            candle_min = int(input("\nCandle size in minutes (e.g. 1, 5, 15): ").strip())
            if candle_min >= 1: break
        except ValueError: pass

    df_candles = resample_candles(df_1m, candle_min)

    # ── Fixed candle time ──
    available_times = sorted(df_candles['datetime'].dt.time.unique())
    print(f"\nAvailable candle times (first 20): {[str(t) for t in available_times[:20]]} ...")
    while True:
        raw = input("Fixed candle time (HH:MM or H:MM, e.g. 10:06): ").strip()
        try:
            parts = raw.split(':')
            h, m  = int(parts[0]), int(parts[1])
            fixed_time = dtime(h, m)
            days_with_candle = (df_candles['datetime'].dt.time == fixed_time).sum()
            if days_with_candle == 0:
                print(f"  No candles found at {fixed_time} — check your time and TF.")
                continue
            log.info(f"  Fixed candle {fixed_time} found on {days_with_candle} bars")
            break
        except Exception:
            print("  Invalid format. Use HH:MM (e.g. 10:06)")

    # ── Mode ──
    print("\nStrategy mode:")
    print("  1. BREAKOUT   — follow the candle (GREEN=LONG, RED=SHORT), entry = next candle open")
    print("  2. REVERSAL   — fade  the candle  (GREEN=SHORT, RED=LONG), entry = next candle open")
    print("  3. LONG ONLY  — always LONG,  entry = fixed candle own open")
    print("  4. SHORT ONLY — always SHORT, entry = fixed candle own open")
    while True:
        try:
            mc = int(input("Select (1-4): ").strip())
            if mc in (1,2,3,4): break
        except ValueError: pass
    mode       = {1:'breakout', 2:'reversal', 3:'long', 4:'short'}[mc]
    mode_label = {1:'BREAKOUT', 2:'REVERSAL', 3:'LONG_ONLY', 4:'SHORT_ONLY'}[mc]

    # ── TIME exit ──
    print("\nTIME exit:")
    print("  1. X minutes from signal candle open")
    print("  2. EOD (end of day fixed time)")
    while True:
        try:
            te = int(input("Select (1-2): ").strip())
            if te in (1,2): break
        except ValueError: pass

    time_exit_type    = 'minutes' if te == 1 else 'eod'
    time_exit_minutes = None
    eod_time          = None

    if te == 1:
        while True:
            try:
                time_exit_minutes = int(input("Minutes from signal candle open (e.g. 60): ").strip())
                if time_exit_minutes >= 1: break
            except ValueError: pass
    else:
        while True:
            raw = input("EOD time (HH:MM, e.g. 16:00): ").strip()
            try:
                parts    = raw.split(':')
                eod_time = dtime(int(parts[0]), int(parts[1]))
                break
            except Exception:
                print("  Invalid format. Use HH:MM")

    # ── Account params ──
    try:
        account_size   = float(input("\nAccount size     (default 4500): ").strip() or 4500)
        risk_per_trade = float(input("Risk per trade $ (default  225): ").strip() or 225)
    except ValueError:
        account_size, risk_per_trade = 4500.0, 225.0

    # ── Budget mode ──
    print("\nBudget enforcement:")
    print("  1. Hard Skip — skip if 1 contract > budget")
    print("  2. Flag Only — take trade, mark violations")
    while True:
        try:
            bm = int(input("Select (1-2): ").strip())
            if bm in (1,2): break
        except ValueError: pass
    budget_mode = 'hard_skip' if bm==1 else 'flag'

    # ── Lookback periods ──
    print("\nLookback periods:")
    print("  1. Standard (90d / 180d / 365d)")
    print("  2. Custom")
    while True:
        try:
            pc = int(input("Select (1-2): ").strip())
            if pc in (1,2): break
        except ValueError: pass

    if pc == 1:
        period_days = {'90d':90, '180d':180, '365d':365}
    else:
        period_days = {}
        while True:
            try:
                n_p = int(input("How many periods?: ").strip())
                if n_p >= 1: break
            except ValueError: pass
        for i in range(n_p):
            while True:
                try:
                    d   = int(input(f"  Period {i+1} days: ").strip())
                    if d >= 1:
                        lbl = f"{d}d" if f"{d}d" not in period_days else f"{d}d_{i+1}"
                        period_days[lbl] = d; break
                except ValueError: pass

    # ── Integrity check ──
    print("\nIntegrity check (strongly recommended):")
    run_ic = input("Run? (Y/n): ").strip().lower()
    if run_ic != 'n':
        try: n_samp = int(input("Combos to validate (default 20): ").strip() or 20)
        except ValueError: n_samp = 20
        # Use 90d slice for integrity check
        max_dt_chk = df_candles['datetime'].max()
        df_chk     = df_candles[df_candles['datetime'] >= max_dt_chk - timedelta(days=90)].copy()
        skels_chk  = extract_trade_skeletons(
            df_chk, fixed_time, mode, time_exit_type,
            time_exit_minutes, eod_time, df_m1=df_1m)
        ok = run_integrity_check(skels_chk, risk_per_trade, account_size,
                                 mode, n_samples=n_samp, budget_mode=budget_mode)
        if not ok:
            print("\n❌  Integrity check FAILED. Aborting.")
            sys.exit(1)

    # ── Run all periods ──
    max_date           = df_candles['datetime'].max()
    out_dir            = os.path.dirname(os.path.abspath(__file__))
    output_files       = []
    all_period_results = {}

    for label, days in period_days.items():
        cutoff      = max_date - timedelta(days=days)
        df_slice    = df_candles[df_candles['datetime'] >= cutoff].copy()
        df_m1_slice = df_1m[df_1m['datetime'] >= cutoff].copy()

        log.info(f"\n{'='*55}")
        log.info(f"Period: {label} | {len(df_slice):,} bars | {mode_label}")

        day_skeletons = extract_trade_skeletons(
            df_slice, fixed_time, mode, time_exit_type,
            time_exit_minutes, eod_time, df_m1=df_m1_slice)

        results_df, all_trades = run_grid(
            day_skeletons, account_size, risk_per_trade,
            label=label, budget_mode=budget_mode)

        all_period_results[label] = results_df
        passing = results_df[results_df['passed']==True]
        log.info(f"Combos: {len(results_df):,} | Passing PF≥1.00: {len(passing):,}")
        if len(passing) > 0:
            best = passing.sort_values('combined_edge', ascending=False).iloc[0]
            log.info(f"Best: SL={best['sl_pct']}% TP={best['tp_pct']}% CE={best['combined_edge']:.4f} PF={best['profit_factor']:.3f}")
            log.info(f"  MAE mode={best.get('mae_mode')}% ({best.get('mae_mode_freq')}% of wins) | median={best.get('mae_median')}%")
            log.info(f"  MFE mode={best.get('mfe_mode')}% ({best.get('mfe_mode_freq')}% of losers) | median={best.get('mfe_median')}%")

        path = build_excel(results_df, all_trades, label,
                           account_size, risk_per_trade, mode_label, out_dir)
        output_files.append(path)

    # ── Master consensus ──
    if len(all_period_results) >= 2:
        log.info(f"\n{'='*55}")
        log.info("Building MASTER CONSENSUS...")
        master_path = build_master_consensus(
            all_period_results, period_days, mode_label, out_dir)
        output_files.append(master_path)

    te_desc = (f"{time_exit_minutes}min from signal candle"
               if time_exit_type == 'minutes' else f"EOD {eod_time}")
    print(f"\n{'='*60}")
    print(f"  DONE — {mode_label} | {candle_min}min | Fixed: {fixed_time} | Exit: {te_desc}")
    print("  Output files:")
    for f in output_files: print(f"    {f}")
    print("=" * 60)


if __name__ == '__main__':
    main()
